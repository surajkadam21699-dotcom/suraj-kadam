"""Portfolio review reports from a filled-in holdings sheet.

    python tools/portfolio_report.py --template Portfolio_Report_Input.xlsx
    python tools/portfolio_report.py --input Portfolio_Report_Input.xlsx

--template writes the blank intake sheet to fill in.
--input reads a filled one and writes two workbooks beside it:

    <Client>_Portfolio_Analysis.xlsx   full review, 5 sheets
    <Client>_Portfolio_1Pager.xlsx     one-sheet summary, prints on one page

Only Fund Name, Category and Current Value are required per holding.
Allocation %, ranking, concentration bands, category rollups, findings and
the keep/exit proposal are all derived. Anything typed into the optional
Keep/Exit and Target % columns overrides the derived proposal.

House rules, all editable on the input sheet:
    largest holding   max 12%      top 3   max 30%      top 5   max 45%
    tail holding      under 1%     a category is crowded at 3+ funds
    a fund is "best"  3Y return at or above 13%

This is the general form of the two one-off scripts written for the
September 2026 report (build_portfolio_analysis_xlsx.py, which reproduces
that PDF's own narrative verbatim, and build_portfolio_onepager_xlsx.py).
"""

import argparse
import datetime
import os
import re
import sys
from collections import defaultdict

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

FONT = "Arial"
NAVY, BLUE, GREY, WHITE, YELLOW, LINE = "1F3864", "2E75B6", "F2F2F2", "FFFFFF", "FFF2CC", "BFBFBF"
RED_FILL, AMBER_FILL, GREEN_FILL, QUIET = "FFC7CE", "FFEB9C", "C6EFCE", "E7E9E8"

BODY = Font(name=FONT, size=10)
BOLD = Font(name=FONT, size=10, bold=True)
CALC = Font(name=FONT, size=10, color="000000")
SRC = Font(name=FONT, size=10, color="0000FF")
HEAD = Font(name=FONT, size=10, bold=True, color=WHITE)
RED = Font(name=FONT, size=10, color="9C0006")
AMBER = Font(name=FONT, size=10, color="9C5700")
GREEN = Font(name=FONT, size=10, color="006100")
ITAL = Font(name=FONT, size=9, italic=True, color="404040")
SECT = Font(name=FONT, size=11, bold=True, color=NAVY)

INR = r'[>=10000000]₹ ##\,##\,##\,##0;[>=100000]₹ ##\,##\,##0;₹ #,##0'
PCT = '0.00%;(0.00%);-'
PCT0 = '0%;(0%);-'
DELTA = '+0.00%;-0.00%;-'
CNT = '#,##0;(#,##0);-'
DATE = 'DD-MMM-YY'
THIN = Side(style="thin", color=LINE)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CATEGORIES = [
    "Equity - Large Cap", "Equity - Large & Mid Cap", "Equity - Mid Cap",
    "Equity - Small Cap", "Equity - Multi Cap", "Equity - Flexi Cap",
    "Equity - Focused", "Equity - Value", "Equity - Contra",
    "Equity - Dividend Yield", "Equity - ELSS", "Equity - Sectoral/Thematic",
    "Equity - Quant", "Hybrid - Aggressive", "Hybrid - Balanced Advantage",
    "Hybrid - Multi Asset", "Hybrid - Conservative", "Hybrid - Equity Savings",
    "Hybrid - Arbitrage", "Debt - Overnight", "Debt - Liquid",
    "Debt - Ultra Short Duration", "Debt - Low Duration", "Debt - Money Market",
    "Debt - Short Duration", "Debt - Corporate Bond", "Debt - Banking and PSU",
    "Debt - Credit Risk", "Debt - Dynamic Bond", "Debt - Gilt",
    "Others - Index", "Others - Gold ETF", "Others - Other ETFs",
    "Others - FoF Domestic", "Others - FoF Overseas", "Others - Solution Oriented",
]

# input sheet layout
IN_FIRST, IN_LAST = 12, 211
DEFAULTS = dict(cap1=0.12, cap3=0.30, cap5=0.45, tail=0.01, crowded=3, best=0.13,
                tcap=0.25)


def put(ws, ref, value, *, font=BODY, fmt=None, fill=None, align=None,
        wrap=False, border=False):
    c = ws[ref]
    c.value = value
    c.font = font
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    if align or wrap:
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if border:
        c.border = BOX
    return c


def banner(ws, title, subtitle, last_col):
    ws.merge_cells(f"A1:{last_col}1")
    put(ws, "A1", title, font=Font(name=FONT, size=13, bold=True, color=WHITE),
        fill=NAVY, align="left")
    ws.row_dimensions[1].height = 28
    ws.merge_cells(f"A2:{last_col}2")
    put(ws, "A2", subtitle, font=ITAL, fill=GREY, align="left")
    ws.row_dimensions[2].height = 17


def heads(ws, row, labels, widths=None, left=()):
    for i, label in enumerate(labels, start=1):
        col = get_column_letter(i)
        put(ws, f"{col}{row}", label, font=HEAD, fill=BLUE,
            align="left" if col in left else "center", wrap=True, border=True)
        if widths:
            ws.column_dimensions[col].width = widths[i - 1]
    ws.row_dimensions[row].height = 30


# =========================================================== intake sheet ====
def write_template(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Input"
    ws.sheet_view.showGridLines = False
    banner(ws, "PORTFOLIO REVIEW  -  DATA INPUT",
           "Fill the blue cells. Only Fund Name, Category and Current Value are required "
           "for each holding - everything else is worked out for you.", "I")

    put(ws, "A4", "Client name", font=BOLD)
    ws.merge_cells("B4:D4")
    put(ws, "B4", "", font=SRC, fill=YELLOW, border=True)
    put(ws, "E4", "As on date", font=BOLD)
    put(ws, "F4", "", font=SRC, fmt=DATE, fill=YELLOW, align="center", border=True)
    put(ws, "A5", "Prepared by", font=BOLD)
    ws.merge_cells("B5:D5")
    put(ws, "B5", "", font=SRC, fill=YELLOW, border=True)

    put(ws, "E5", "House rules", font=BOLD)
    rules = [("G4", "Max largest holding", "H4", DEFAULTS["cap1"], PCT),
             ("G5", "Max top 3", "H5", DEFAULTS["cap3"], PCT),
             ("G6", "Max top 5", "H6", DEFAULTS["cap5"], PCT),
             ("G7", "Tail holding below", "H7", DEFAULTS["tail"], PCT),
             ("G8", "Best fund: 3Y at or above", "H8", DEFAULTS["best"], PCT),
             ("G9", "Max weight per fund in target", "H9", DEFAULTS["tcap"], PCT)]
    for lab_ref, label, val_ref, val, fmt in rules:
        put(ws, lab_ref, label, font=Font(name=FONT, size=9))
        put(ws, val_ref, val, font=SRC, fmt=fmt, fill=YELLOW, align="center", border=True)

    ws.merge_cells("A10:I10")
    put(ws, "A10", "▼  Required: Fund Name, Category, Current Value.    "
                  "Optional: 3Y Return (drives the best/worst call), and Keep/Exit + "
                  "Target % if you want to override the proposal this tool works out.",
        font=ITAL, fill=GREY)

    labels = ["#", "Fund Name", "Category", "Current Value (Rs)", "Allocation %",
              "3Y Return", "Keep / Exit", "Target %", "Remark"]
    widths = [5, 46, 28, 18, 13, 12, 13, 11, 40]
    heads(ws, 11, labels, widths, left=("B", "C", "I"))
    ws.freeze_panes = "C12"

    for r in range(IN_FIRST, IN_LAST + 1):
        put(ws, f"A{r}", f'=IF($B{r}="","",COUNTA($B${IN_FIRST}:$B{r}))',
            font=CALC, fmt=CNT, fill=GREY, align="center", border=True)
        put(ws, f"B{r}", None, font=SRC, border=True)
        put(ws, f"C{r}", None, font=SRC, border=True)
        put(ws, f"D{r}", None, font=SRC, fmt=INR, border=True)
        put(ws, f"E{r}", f'=IF($B{r}="","",IF(SUM($D${IN_FIRST}:$D${IN_LAST})=0,"",'
                         f'$D{r}/SUM($D${IN_FIRST}:$D${IN_LAST})))',
            font=CALC, fmt=PCT, fill=GREY, align="center", border=True)
        put(ws, f"F{r}", None, font=SRC, fmt=PCT, align="center", border=True)
        put(ws, f"G{r}", None, font=SRC, align="center", border=True)
        put(ws, f"H{r}", None, font=SRC, fmt=PCT0, align="center", border=True)
        put(ws, f"I{r}", None, font=SRC, border=True)

    t = IN_LAST + 1
    put(ws, f"A{t}", "", fill=NAVY, border=True)
    put(ws, f"B{t}", "TOTAL", font=HEAD, fill=NAVY, border=True)
    put(ws, f"C{t}", f'=COUNTA($B${IN_FIRST}:$B${IN_LAST})&" holdings"',
        font=HEAD, fill=NAVY, border=True)
    put(ws, f"D{t}", f"=SUM($D${IN_FIRST}:$D${IN_LAST})", font=HEAD, fmt=INR,
        fill=NAVY, border=True)
    put(ws, f"E{t}", f"=SUM($E${IN_FIRST}:$E${IN_LAST})", font=HEAD, fmt=PCT,
        fill=NAVY, align="center", border=True)
    for col in "FGHI":
        put(ws, f"{col}{t}", "", fill=NAVY, border=True)

    for i, cat in enumerate(CATEGORIES, start=1):
        put(ws, f"L{i}", cat)
    for i, v in enumerate(["Keep", "Exit"], start=1):
        put(ws, f"M{i}", v)
    ws.column_dimensions["L"].hidden = True
    ws.column_dimensions["M"].hidden = True

    dv = DataValidation(type="list", formula1=f"=$L$1:$L${len(CATEGORIES)}",
                        allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(f"C{IN_FIRST}:C{IN_LAST}")
    dv2 = DataValidation(type="list", formula1="=$M$1:$M$2", allow_blank=True,
                         showDropDown=False)
    ws.add_data_validation(dv2)
    dv2.add(f"G{IN_FIRST}:G{IN_LAST}")

    # one example row, marked for deletion
    for col, v in (("B", "Bluechip Flexi Cap Fund - Growth (EXAMPLE)"),
                   ("C", "Equity - Flexi Cap"), ("D", 265430), ("F", 0.1512),
                   ("I", "EXAMPLE ROW - delete it before you send the file.")):
        ws[f"{col}{IN_FIRST}"].value = v

    wb.save(path)
    return path


# ================================================================ analysis ===
class Holding:
    __slots__ = ("name", "cat", "value", "alloc", "r3y", "keep", "target", "remark", "rank")

    def __init__(self, name, cat, value, r3y, keep, target, remark):
        self.name, self.cat, self.value = name, cat, value
        self.r3y, self.keep, self.target, self.remark = r3y, keep, target, remark
        self.alloc = 0.0
        self.rank = 0

    @property
    def band(self):
        return "Over 10%" if self.alloc > 0.10 else "5 - 10%" if self.alloc >= 0.05 else "Under 5%"


def read_input(path):
    wb = load_workbook(path, data_only=True)
    ws = wb["Input"] if "Input" in wb.sheetnames else wb.worksheets[0]

    def cell(ref, default=None):
        v = ws[ref].value
        return default if v in (None, "") else v

    meta = dict(
        client=str(cell("B4", "Client")).strip(),
        ason=cell("F4") or datetime.date.today(),
        by=str(cell("B5", "")).strip(),
        cap1=float(cell("H4", DEFAULTS["cap1"])),
        cap3=float(cell("H5", DEFAULTS["cap3"])),
        cap5=float(cell("H6", DEFAULTS["cap5"])),
        tail=float(cell("H7", DEFAULTS["tail"])),
        best=float(cell("H8", DEFAULTS["best"])),
        tcap=float(cell("H9", DEFAULTS["tcap"])),
        crowded=DEFAULTS["crowded"],
    )
    if isinstance(meta["ason"], datetime.datetime):
        meta["ason"] = meta["ason"].date()

    rows = []
    for r in range(IN_FIRST, IN_LAST + 1):
        name = ws[f"B{r}"].value
        if not name or not str(name).strip():
            continue
        name = str(name).strip()
        if "(EXAMPLE)" in name.upper():
            continue
        value = ws[f"D{r}"].value
        alloc = ws[f"E{r}"].value
        if value in (None, "") and alloc in (None, ""):
            raise SystemExit(f"row {r}: {name!r} has neither Current Value nor Allocation %")
        keep = ws[f"G{r}"].value
        keep = str(keep).strip().lower() if keep else None
        h = Holding(name, str(ws[f"C{r}"].value or "Uncategorised").strip(),
                    float(value or 0), ws[f"F{r}"].value,
                    True if keep == "keep" else False if keep == "exit" else None,
                    ws[f"H{r}"].value, ws[f"I{r}"].value or "")
        if value in (None, "") and alloc:
            h.alloc = float(alloc)          # percentage given directly
        rows.append(h)
    if not rows:
        raise SystemExit("no holdings found - fill column B on the Input sheet")

    total = sum(h.value for h in rows)
    if total > 0:
        for h in rows:
            h.alloc = h.value / total
    else:                                    # allocations supplied directly
        s = sum(h.alloc for h in rows)
        if abs(s - 1) > 0.02:
            raise SystemExit(f"allocations sum to {s:.2%}, not 100% - check column E")
    rows.sort(key=lambda h: -h.alloc)
    for i, h in enumerate(rows, start=1):
        h.rank = i
    return meta, rows


def limit_checks(meta, rows):
    """The concentration checks that actually mean something for this portfolio.

    The largest of n holdings can never be below 1/n, so a 12% ceiling is
    unreachable with 8 funds and flagging it would be a fake breach. Each check
    is therefore only applied when the cap is arithmetically achievable.
    """
    n = len(rows)
    out = []
    for k, cap, who in ((1, meta["cap1"], rows[0].name),
                        (3, meta["cap3"], "Three largest holdings"),
                        (5, meta["cap5"], "Five largest holdings")):
        if n >= k and k / n <= cap:
            label = "Largest holding" if k == 1 else f"Top {k} funds"
            out.append((k, label, who, cap))
    return out


def top_n(rows, n):
    return sum(h.alloc for h in rows[:n])


def by_category(rows):
    agg = defaultdict(lambda: [0, 0.0])
    for h in rows:
        agg[h.cat][0] += 1
        agg[h.cat][1] += h.alloc
    return sorted(((c, n, a) for c, (n, a) in agg.items()), key=lambda t: -t[2])


def analyse(meta, rows):
    """Derive findings and the keep/exit proposal. Explicit input always wins."""
    cats = by_category(rows)
    findings = []
    pct = lambda v: f"{v * 100:.2f}%"

    for k, label, who, cap in limit_checks(meta, rows):
        actual = top_n(rows, k)
        if actual <= cap:
            continue
        if k == 1:
            findings.append(("Concentration",
                             f"{who} alone holds {pct(actual)} against a {pct(cap)} ceiling."))
        else:
            findings.append((f"Top {k} too heavy",
                             f"The {k} largest holdings take {pct(actual)} of the portfolio; "
                             f"the ceiling is {pct(cap)}."))
    for cat, n, a in cats:
        if n >= meta["crowded"]:
            findings.append((f"{n} funds in {cat}",
                             f"{pct(a)} split across {n} overlapping funds - no single "
                             f"holding large enough to matter."))
    for cat, n, a in cats:
        if n == 2 and a > 0.20:
            findings.append((f"The same bet twice - {cat}",
                             f"Two funds in one category hold {pct(a)} between them."))
    tail = [h for h in rows if h.alloc < meta["tail"]]
    if tail:
        findings.append((f"{len(tail)} holding{'s' if len(tail) > 1 else ''} below "
                         f"{pct(meta['tail'])}",
                         f"The smallest is {pct(tail[-1].alloc)}. These add tracking work "
                         f"and no meaningful return."))
    weak = [h for h in rows if h.r3y is not None and h.r3y < meta["best"]]
    if weak:
        worst = min(weak, key=lambda h: h.r3y)
        findings.append((f"{len(weak)} fund{'s' if len(weak) > 1 else ''} below the "
                         f"{pct(meta['best'])} bar",
                         f"Weakest is {worst.name} at {pct(worst.r3y)} over three years."))

    # --- proposal: one fund per category, largest (or best-returning) wins ----
    best_of = {}
    for h in rows:
        cur = best_of.get(h.cat)
        better = cur is None or (
            (h.r3y or -1, h.alloc) > (cur.r3y or -1, cur.alloc))
        if better:
            best_of[h.cat] = h
    for h in rows:
        if h.keep is None:                    # not overridden on the input sheet
            h.keep = (h.alloc >= meta["tail"] and best_of[h.cat] is h)

    keeps = [h for h in rows if h.keep]
    if not keeps:                             # never propose an empty portfolio
        keeps = rows[:1]
        keeps[0].keep = True

    fixed = sum(h.target for h in keeps if h.target)
    free = [h for h in keeps if not h.target]
    pool = max(0.0, 1.0 - fixed)

    # Weight the keepers by their current size, but no fund may exceed the target
    # cap. Capping one fund frees weight that has to land on the others, so fill
    # in rounds until nothing is over the cap - a single pass would silently lose
    # whatever the capped funds gave up.
    cap = meta["tcap"]
    if free:
        if cap * len(free) < pool:            # cap too tight to reach 100%
            cap = pool / len(free)            # fall back to equal weights
        share = {h: h.alloc for h in free}
        settled, remaining = {}, pool
        pending = list(free)
        while pending:
            base = sum(share[h] for h in pending) or len(pending)
            over = [h for h in pending
                    if remaining * ((share[h] / base) if base else 1 / len(pending)) > cap]
            if not over:
                for h in pending:
                    part = (share[h] / base) if base else 1 / len(pending)
                    settled[h] = remaining * part
                break
            for h in over:
                settled[h] = cap
                remaining -= cap
                pending.remove(h)
        for h, v in settled.items():
            h.target = round(v, 4)

    drift = round(1.0 - sum(h.target for h in keeps), 4)
    if free and abs(drift) > 1e-9:            # put rounding on the smallest keeper
        smallest = min(free, key=lambda h: h.target)
        smallest.target = round(smallest.target + drift, 4)
    return findings, cats, keeps


# ========================================================== full analysis ====
def write_analysis(path, meta, rows, findings, cats, keeps):
    wb = Workbook()
    wb.remove(wb.active)
    ason = meta["ason"].strftime("%d %B %Y")
    n = len(rows)
    h_first = 5
    h_last = h_first + n - 1
    alloc_rng = f"Holdings!$D${h_first}:$D${h_last}"

    # ------------------------------------------------------------ Holdings --
    ws = wb.create_sheet("Holdings")
    ws.sheet_view.showGridLines = False
    banner(ws, f"HOLDINGS  -  {meta['client'].upper()}",
           f"As on {ason}.  {n} holdings.  Allocation computed from current value.", "H")
    heads(ws, 4, ["Rank", "Fund Name", "Category", "Allocation %", "Current Value",
                  "3Y Return", "Keep / Exit", "Target %"],
          [6, 46, 26, 13, 17, 12, 13, 12], left=("B", "C"))
    ws.freeze_panes = "C5"

    for i, h in enumerate(rows):
        r = h_first + i
        font, fill = ((RED, RED_FILL) if h.band == "Over 10%"
                      else (AMBER, AMBER_FILL) if h.band == "5 - 10%" else (BODY, QUIET))
        put(ws, f"A{r}", h.rank, fmt=CNT, align="center", border=True)
        put(ws, f"B{r}", h.name, font=SRC, border=True)
        put(ws, f"C{r}", h.cat, border=True)
        put(ws, f"D{r}", h.alloc, font=font, fmt=PCT, fill=fill, align="center", border=True)
        put(ws, f"E{r}", h.value or None, font=SRC, fmt=INR, border=True)
        put(ws, f"F{r}", h.r3y, font=SRC, fmt=PCT, align="center", border=True)
        put(ws, f"G{r}", "KEEP" if h.keep else "Exit",
            font=GREEN if h.keep else RED, fill=GREEN_FILL if h.keep else None,
            align="center", border=True)
        put(ws, f"H{r}", h.target if h.keep else 0, font=SRC, fmt=PCT0,
            align="center", border=True)
    t = h_last + 1
    for col in "ABCEFG":
        put(ws, f"{col}{t}", "", fill=NAVY, border=True)
    put(ws, f"B{t}", f"TOTAL  ({n} holdings)", font=HEAD, fill=NAVY, border=True)
    put(ws, f"D{t}", f"=SUM({alloc_rng})", font=HEAD, fmt=PCT, fill=NAVY,
        align="center", border=True)
    put(ws, f"E{t}", f"=SUM(E{h_first}:E{h_last})", font=HEAD, fmt=INR, fill=NAVY, border=True)
    put(ws, f"G{t}", f'=COUNTIF(G{h_first}:G{h_last},"KEEP")&" keep / "&'
                     f'COUNTIF(G{h_first}:G{h_last},"Exit")&" exit"',
        font=HEAD, fill=NAVY, align="center", border=True)
    put(ws, f"H{t}", f"=SUM(H{h_first}:H{h_last})", font=HEAD, fmt=PCT0, fill=NAVY,
        align="center", border=True)
    ws.auto_filter.ref = f"A4:H{h_last}"

    # ------------------------------------------------------------ Category --
    ws = wb.create_sheet("Category")
    ws.sheet_view.showGridLines = False
    banner(ws, "PORTFOLIO BY CATEGORY", "Computed live from the Holdings sheet.", "D")
    heads(ws, 4, ["Category", "Funds", "Allocation %", "Crowded?"],
          [30, 10, 15, 26], left=("A", "D"))
    first = 5
    for i, (cat, cnt, _) in enumerate(cats):
        r = first + i
        put(ws, f"A{r}", cat, border=True)
        put(ws, f"B{r}", f'=COUNTIF(Holdings!$C${h_first}:$C${h_last},$A{r})',
            font=CALC, fmt=CNT, align="center", border=True)
        put(ws, f"C{r}", f'=SUMIF(Holdings!$C${h_first}:$C${h_last},$A{r},{alloc_rng})',
            font=CALC, fmt=PCT, align="center", border=True)
        put(ws, f"D{r}", f'=IF($B{r}>={meta["crowded"]},"Yes - consolidate","")',
            font=RED, align="center", border=True)
    ct = first + len(cats)
    put(ws, f"A{ct}", "TOTAL", font=HEAD, fill=NAVY, border=True)
    put(ws, f"B{ct}", f"=SUM(B{first}:B{ct - 1})", font=HEAD, fmt=CNT, fill=NAVY,
        align="center", border=True)
    put(ws, f"C{ct}", f"=SUM(C{first}:C{ct - 1})", font=HEAD, fmt=PCT, fill=NAVY,
        align="center", border=True)
    put(ws, f"D{ct}", "", fill=NAVY, border=True)
    chart = BarChart()
    chart.type = "bar"
    chart.title = "Allocation by category"
    chart.height, chart.width = max(7, 0.5 * len(cats) + 3), 13
    chart.add_data(Reference(ws, min_col=3, min_row=4, max_row=ct - 1), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=first, max_row=ct - 1))
    chart.legend = None
    ws.add_chart(chart, "F4")

    # -------------------------------------------------- Proposed Portfolio --
    ws = wb.create_sheet("Proposed Portfolio")
    ws.sheet_view.showGridLines = False
    banner(ws, f"PROPOSED PORTFOLIO  -  {len(keeps)} FUNDS",
           f"From {n} holdings down to {len(keeps)}. Target weights are proportional to "
           f"current size, capped at {meta['tcap']:.0%} per fund - override any of them "
           f"in the Target % column of the input sheet.", "E")
    heads(ws, 4, ["Fund", "Category", "Now", "Target", "Change"],
          [46, 28, 13, 13, 13], left=("A", "B"))
    kf = 5
    for i, h in enumerate(keeps):
        r = kf + i
        put(ws, f"A{r}", h.name, font=SRC, border=True)
        put(ws, f"B{r}", h.cat, border=True)
        put(ws, f"C{r}", h.alloc, font=SRC, fmt=PCT, align="center", border=True)
        put(ws, f"D{r}", h.target, font=SRC, fmt=PCT0, fill=GREEN_FILL,
            align="center", border=True)
        put(ws, f"E{r}", f"=$D{r}-$C{r}", font=CALC, fmt=DELTA, align="center", border=True)
    kt = kf + len(keeps)
    put(ws, f"A{kt}", f"TOTAL  ({len(keeps)} funds)", font=HEAD, fill=NAVY, border=True)
    put(ws, f"B{kt}", "", fill=NAVY, border=True)
    put(ws, f"C{kt}", f"=SUM(C{kf}:C{kt - 1})", font=HEAD, fmt=PCT, fill=NAVY,
        align="center", border=True)
    put(ws, f"D{kt}", f"=SUM(D{kf}:D{kt - 1})", font=HEAD, fmt=PCT0, fill=NAVY,
        align="center", border=True)
    put(ws, f"E{kt}", f"=$D{kt}-$C{kt}", font=HEAD, fmt=DELTA, fill=NAVY,
        align="center", border=True)
    r = kt + 2
    ws.merge_cells(f"A{r}:E{r + 1}")
    put(ws, f"A{r}", f"The {n - len(keeps)} funds not listed here are exited. Start by "
                     f"redirecting every SIP from those funds into these {len(keeps)}: it "
                     f"rebalances with new money before a single unit is sold, so nothing "
                     f"is taxed.", font=GREEN, fill=GREEN_FILL, wrap=True, border=True)

    # ---------------------------------------------------------- Action Plan --
    ws = wb.create_sheet("Action Plan")
    ws.sheet_view.showGridLines = False
    banner(ws, "ACTION PLAN", "Owner and Done are yours to fill in.", "E")
    for col, w in zip("ABCDE", (14, 30, 56, 18, 10)):
        ws.column_dimensions[col].width = w
    plan = [
        ("Week 1", "Tax position", "Confirm holding periods and cost basis with the CA before any exit."),
        ("Week 1-2", "Stop the bleeding", f"Redirect all SIPs from the {n - len(keeps)} exiting funds into the {len(keeps)} keepers."),
        ("Week 3-4", "Clear the tail", f"Exit holdings below {meta['tail']:.0%} - they cost attention and return nothing."),
        ("Month 2-4", "Remove duplicates", "Exit the redundant funds inside crowded categories, one category at a time."),
        ("Month 4-6", "Rebalance", "Bring the keepers to their target weights and document the gains for filing."),
        ("Ongoing", "Review", "Check quarterly, rebalance once a year."),
    ]
    heads(ws, 4, ["When", "Step", "What happens", "Owner", "Done"], left=("B", "C"))
    r = 5
    for when, step, what in plan:
        put(ws, f"A{r}", when, font=BOLD, align="center", border=True)
        put(ws, f"B{r}", step, font=BOLD, border=True)
        put(ws, f"C{r}", what, wrap=True, border=True)
        put(ws, f"D{r}", None, font=SRC, border=True)
        put(ws, f"E{r}", None, font=SRC, align="center", border=True)
        ws.row_dimensions[r].height = 26
        r += 1
    r += 1
    ws.merge_cells(f"A{r}:E{r + 1}")
    put(ws, f"A{r}", "This analysis is based on current allocation data and general "
                     "investment principles. Consult your financial adviser and CA for tax "
                     "planning and suitability before acting. Mutual fund investments are "
                     "subject to market risks; read all scheme-related documents carefully.",
        font=ITAL, wrap=True)

    # -------------------------------------------------------------- Summary --
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    banner(ws, f"PORTFOLIO REVIEW  -  {meta['client'].upper()}",
           f"As on {ason}." + (f"  Prepared by {meta['by']}." if meta["by"] else ""), "E")
    for col, w in zip("ABCDE", (30, 16, 16, 16, 58)):
        ws.column_dimensions[col].width = w

    put(ws, "A4", "CONCENTRATION AGAINST LIMITS", font=SECT)
    heads(ws, 5, ["Measure", "Actual", "Limit", "", "Verdict"], left=("A", "E"))
    checks = limit_checks(meta, rows)
    r = 6
    for k, label, _who, cap in checks:
        formula = "=" + "+".join(f"LARGE({alloc_rng},{i})" for i in range(1, k + 1))
        put(ws, f"A{r}", label, font=BOLD, border=True)
        put(ws, f"B{r}", formula, font=CALC, fmt=PCT, align="center", border=True)
        put(ws, f"C{r}", cap, font=SRC, fmt=PCT, align="center", border=True)
        put(ws, f"D{r}", "", border=True)
        put(ws, f"E{r}", f'=IF($B{r}>$C{r},"OVER by "&TEXT(($B{r}-$C{r})*100,"0.00")&'
                         f'" pts","Within limit")',
            font=RED, border=True)
        r += 1

    if not checks:
        ws.merge_cells(f"A{r}:E{r}")
        put(ws, f"A{r}", f"With {n} holdings the largest is at least "
                         f"{1 / n:.2%}, so these ceilings cannot be met by arithmetic "
                         f"and are not applied.", font=ITAL, fill=YELLOW, wrap=True, border=True)
        r += 1

    r += 1
    put(ws, f"A{r}", "HEADLINE", font=SECT)
    r += 1
    lines = [("Holdings", f"=COUNTA(Holdings!$B${h_first}:$B${h_last})", CNT),
             ("Portfolio value", f"=SUM(Holdings!$E${h_first}:$E${h_last})", INR),
             ("Funds kept", f'=COUNTIF(Holdings!$G${h_first}:$G${h_last},"KEEP")', CNT),
             ("Funds exited", f'=COUNTIF(Holdings!$G${h_first}:$G${h_last},"Exit")', CNT),
             ("Kept funds hold today",
              f'=SUMIF(Holdings!$G${h_first}:$G${h_last},"KEEP",{alloc_rng})', PCT)]
    for label, formula, fmt in lines:
        put(ws, f"A{r}", label, font=BOLD, border=True)
        ws.merge_cells(f"B{r}:C{r}")
        put(ws, f"B{r}", formula, font=CALC, fmt=fmt, fill=GREY, align="center", border=True)
        put(ws, f"D{r}", "", border=True)
        put(ws, f"E{r}", "", border=True)
        r += 1

    r += 1
    put(ws, f"A{r}", "WHAT IS WRONG", font=SECT)
    r += 1
    if not findings:
        ws.merge_cells(f"A{r}:E{r}")
        put(ws, f"A{r}", "No house rule is breached. The portfolio is within every "
                         "concentration limit and carries no crowded category or tail "
                         "holding.", font=GREEN, fill=GREEN_FILL, wrap=True, border=True)
        r += 1
    for title, detail in findings:
        put(ws, f"A{r}", title, font=BOLD, wrap=True, border=True)
        ws.merge_cells(f"B{r}:E{r}")
        put(ws, f"B{r}", detail, wrap=True, border=True)
        ws.row_dimensions[r].height = 28
        r += 1

    wb.move_sheet("Summary", offset=-4)
    wb.active = 0
    wb.save(path)
    return path


# =============================================================== one-pager ===
def write_onepager(path, meta, rows, findings, cats, keeps):
    wb = Workbook()
    ws = wb.active
    ws.title = "Review"
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEF", (6, 52, 22, 13, 13, 44)):
        ws.column_dimensions[col].width = w
    n = len(rows)
    ason = meta["ason"].strftime("%d %B %Y")
    top4 = sum(h.alloc for h in rows if h.alloc > 0.10)
    n_over = sum(1 for h in rows if h.alloc > 0.10)
    tail = [h for h in rows if h.alloc < meta["tail"]]

    banner(ws, f"PORTFOLIO CONCENTRATION REVIEW  -  {meta['client'].upper()}",
           f"As on {ason}   |   {n} holdings   |   Allocation basis 100.00%"
           + (f"   |   {meta['by']}" if meta["by"] else ""), "F")

    ws.merge_cells("A4:F5")
    lead = (f"The portfolio holds {n} funds. "
            + (f"{n_over} of them take {top4:.2%} of the money. " if n_over else "")
            + (f"{len(tail)} holdings at the tail contribute almost nothing. " if tail else "")
            + f"The proposal is to consolidate into {len(keeps)}.")
    put(ws, "A4", lead, font=Font(name=FONT, size=11), wrap=True)
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 20

    h_first = 16
    h_last = h_first + n - 1
    alloc = f"$D${h_first}:$D${h_last}"

    put(ws, "A7", "CONCENTRATION AGAINST LIMITS", font=SECT)
    heads(ws, 8, ["", "Measure", "Holding", "Actual", "Limit", "Verdict"],
          left=("B", "C", "F"))
    r = 9
    for k, label, who, cap in limit_checks(meta, rows):
        formula = "=" + "+".join(f"LARGE({alloc},{i})" for i in range(1, k + 1))
        put(ws, f"A{r}", "", border=True)
        put(ws, f"B{r}", label, font=BOLD, border=True)
        put(ws, f"C{r}", who, border=True)
        put(ws, f"D{r}", formula, font=CALC, fmt=PCT, fill=RED_FILL, align="center", border=True)
        put(ws, f"E{r}", cap, font=SRC, fmt=PCT, align="center", border=True)
        put(ws, f"F{r}", f'=IF($D{r}>$E{r},"OVER by "&TEXT(($D{r}-$E{r})*100,"0.00")&'
                         f'" pts","Within limit")', font=RED, fill=RED_FILL, border=True)
        r += 1

    put(ws, "A13", f"WHERE THE MONEY SITS  -  ALL {n} HOLDINGS", font=SECT)
    ws.merge_cells("A14:F14")
    put(ws, "A14", "Bars are drawn to scale.", font=ITAL)
    heads(ws, 15, ["#", "Fund", "Category", "Allocation", "Band", "Action"],
          left=("B", "C", "F"))
    for i, h in enumerate(rows):
        rr = h_first + i
        font, fill = ((RED, RED_FILL) if h.band == "Over 10%"
                      else (AMBER, AMBER_FILL) if h.band == "5 - 10%" else (BODY, QUIET))
        put(ws, f"A{rr}", h.rank, fmt=CNT, align="center", border=True)
        put(ws, f"B{rr}", h.name, font=SRC, border=True)
        put(ws, f"C{rr}", h.cat, border=True)
        put(ws, f"D{rr}", h.alloc, font=SRC, fmt=PCT, align="center", border=True)
        put(ws, f"E{rr}", h.band, font=font, fill=fill, align="center", border=True)
        put(ws, f"F{rr}", "KEEP" if h.keep else "Exit / consolidate",
            font=GREEN if h.keep else RED, fill=GREEN_FILL if h.keep else None, border=True)
    t = h_last + 1
    for col in "ACE":
        put(ws, f"{col}{t}", "", fill=NAVY, border=True)
    put(ws, f"B{t}", "TOTAL", font=HEAD, fill=NAVY, border=True)
    put(ws, f"D{t}", f"=SUM({alloc})", font=HEAD, fmt=PCT, fill=NAVY,
        align="center", border=True)
    put(ws, f"F{t}", f'=COUNTIF($F${h_first}:$F${h_last},"KEEP")&" keep  /  "&'
                     f'COUNTIF($F${h_first}:$F${h_last},"Exit / consolidate")&" exit"',
        font=HEAD, fill=NAVY, border=True)
    ws.conditional_formatting.add(
        f"D{h_first}:D{h_last}",
        DataBarRule(start_type="num", start_value=0,
                    end_type="num", end_value=round(rows[0].alloc * 1.05, 4),
                    color="A6362C", showValue=True))

    r = t + 2
    put(ws, f"A{r}", "WHAT IS WRONG", font=SECT)
    r += 1
    if not findings:
        ws.merge_cells(f"A{r}:F{r}")
        put(ws, f"A{r}", "No house rule is breached.", font=GREEN, fill=GREEN_FILL, border=True)
        r += 1
    for title, detail in findings:
        put(ws, f"A{r}", "", border=True)
        put(ws, f"B{r}", title, font=BOLD, wrap=True, border=True)
        ws.merge_cells(f"C{r}:F{r}")
        put(ws, f"C{r}", detail, wrap=True, border=True)
        ws.row_dimensions[r].height = 26
        r += 1

    r += 1
    put(ws, f"A{r}", f"THE FIX  -  {n} FUNDS DOWN TO {len(keeps)}", font=SECT)
    r += 1
    heads(ws, r, ["", "Fund", "Category", "Now", "Target", "Change"], left=("B", "C"))
    r += 1
    kf = r
    for h in keeps:
        put(ws, f"A{r}", "", border=True)
        put(ws, f"B{r}", h.name, font=SRC, border=True)
        put(ws, f"C{r}", h.cat, border=True)
        put(ws, f"D{r}", h.alloc, font=SRC, fmt=PCT, align="center", border=True)
        put(ws, f"E{r}", h.target, font=SRC, fmt=PCT0, fill=GREEN_FILL,
            align="center", border=True)
        put(ws, f"F{r}", f"=$E{r}-$D{r}", font=CALC, fmt=DELTA, align="center", border=True)
        r += 1
    for col in ("A", "C"):
        put(ws, f"{col}{r}", "", fill=NAVY, border=True)
    put(ws, f"B{r}", f"{len(keeps)} holdings", font=HEAD, fill=NAVY, border=True)
    put(ws, f"D{r}", f"=SUM(D{kf}:D{r - 1})", font=HEAD, fmt=PCT, fill=NAVY,
        align="center", border=True)
    put(ws, f"E{r}", f"=SUM(E{kf}:E{r - 1})", font=HEAD, fmt=PCT0, fill=NAVY,
        align="center", border=True)
    put(ws, f"F{r}", f"=$E{r}-$D{r}", font=HEAD, fmt=DELTA, fill=NAVY,
        align="center", border=True)
    r += 2

    ws.merge_cells(f"A{r}:F{r + 1}")
    put(ws, f"A{r}", f"FIRST MOVE, THIS MONTH:  stop every SIP going into the "
                     f"{n - len(keeps)} funds being exited and redirect them to these "
                     f"{len(keeps)}. It rebalances with new money before a single unit is "
                     f"sold, so nothing is taxed.",
        font=GREEN, fill=GREEN_FILL, wrap=True, border=True)
    ws.row_dimensions[r].height = 20
    ws.row_dimensions[r + 1].height = 20
    r += 3

    ws.merge_cells(f"A{r}:F{r + 1}")
    put(ws, f"A{r}", "This analysis is based on current allocation data and general "
                     "investment principles. Before implementing any changes, consult your "
                     "financial adviser and CA for tax planning and suitability assessment. "
                     "Mutual fund investments are subject to market risks; read all "
                     "scheme-related documents carefully.", font=ITAL, wrap=True)

    ws.print_area = f"A1:F{r + 1}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    wb.save(path)
    return path


def slug(name):
    return re.sub(r"[^A-Za-z0-9]+", "_", name).strip("_") or "Client"


def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    g = ap.add_mutually_exclusive_group(required=True)
    g.add_argument("--template", metavar="OUT", help="write the blank intake sheet")
    g.add_argument("--input", metavar="FILE", help="read a filled intake sheet")
    ap.add_argument("--outdir", default=".", help="where to write the reports")
    a = ap.parse_args(argv)

    if a.template:
        print(f"wrote {write_template(a.template)}")
        return

    meta, rows = read_input(a.input)
    findings, cats, keeps = analyse(meta, rows)
    stem = os.path.join(a.outdir, slug(meta["client"]))
    p1 = write_analysis(f"{stem}_Portfolio_Analysis.xlsx", meta, rows, findings, cats, keeps)
    p2 = write_onepager(f"{stem}_Portfolio_1Pager.xlsx", meta, rows, findings, cats, keeps)
    print(f"{len(rows)} holdings | {len(keeps)} keep / {len(rows) - len(keeps)} exit "
          f"| {len(findings)} findings")
    print(f"wrote {p1}\nwrote {p2}")


if __name__ == "__main__":
    main()

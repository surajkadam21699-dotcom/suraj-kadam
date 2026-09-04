"""Convert Profit & Loss reports from the portfolio software into the
SCHEME-WISE INVESTMENT REGISTER layout of Client_Investment_Tracker.xlsx.

Usage:
    python tools/convert_pl_report.py OUTPUT.xlsx REPORT.xlsx [REPORT2.xlsx ...]

Each report is one client. The scheme summary line under every category
heading becomes one register row; the transaction sub-tables, the per-scheme
"Total" lines and the category / grand total lines are skipped, because the
register recomputes those itself.

The output is a filled copy of the tracker: the Clients sheet carries the
report header block, the Investments sheet carries the scheme rows, and the
Total Review sheet adds them up. A .csv of the same rows is written beside
it for pasting into an existing register.

Every parsed client is checked against the report's own Grand Total line;
any mismatch beyond one rupee is reported and the exit status is non-zero.
"""

import csv
import re
import subprocess
import sys
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font

BUILDER = Path(__file__).with_name("build_client_investment_template.py")
CATEGORY_RE = re.compile(r"^(Equity|Debt|Hybrid|Others)\s+-\s+.+")
FOLIO_RE = re.compile(r"\s*Folio\s*:\s*", re.IGNORECASE)

# report column -> meaning, from the "Scheme Name / Inv. Since / Purchase ..." header
COL = {"since": 4, "purchase": 5, "switch_in": 6, "div_reinv": 7, "redemption": 8,
       "switch_out": 9, "div_pay": 10, "cur_value": 11, "cur_units": 12,
       "cur_nav": 13, "gain": 14, "abs_rtn": 15, "xirr": 16}

# the transaction sub-table printed under each scheme
TXN = {"sr": 1, "type": 2, "date": 5, "amount": 7, "pur_nav": 9, "units": 11,
       "sensex": 13, "cur_value": 15, "days": 17, "stt": 19}
TXN_HEADER = ["Client Name", "Category", "Scheme Name", "Folio No.", "Sr.",
              "Transaction Type", "Date", "Amount", "Pur. NAV", "Units",
              "Sensex", "Cur. Value", "Days", "STT"]

# register column (Investments sheet) for each parsed field
# Investments columns A-O, in order - the same 15 columns the CSV carries, so
# a CSV row and a register row line up one to one.
REGISTER = [
    ("client", 1), ("category", 2), ("scheme", 3), ("folio", 4), ("since", 5),
    ("purchase", 6), ("switch_in", 7), ("div_reinv", 8), ("redemption", 9),
    ("switch_out", 10), ("div_pay", 11), ("cur_value", 12), ("cur_units", 13),
    ("cur_nav", 14), ("xirr", 15),
]
CSV_HEADER = ["Client Name", "Category", "Scheme Name", "Folio No.", "Inv. Since",
              "Purchase", "Switch In", "Div Reinv", "Redemption", "Switch Out",
              "Div Pay", "Cur. Value", "Cur. Units", "Cur. NAV", "XIRR"]


def num(value):
    """'5,00,000.00' / 5000.0 / '' -> float."""
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "").replace("₹", "")
    if text in ("", "-", "--"):
        return 0.0
    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()").rstrip("%")
    try:
        out = float(text)
    except ValueError:
        return 0.0
    return -out if negative else out


def pct(value):
    """'3.80%' -> 0.038 (the software prints percentages as text)."""
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value) / 100 if abs(float(value)) > 1 else float(value)
    return num(value) / 100


def parse_date(value):
    """'01-06-21' -> date(2021, 6, 1). The software prints dd-mm-yy."""
    if isinstance(value, date):
        return value
    if not value:
        return None
    text = str(value).strip()
    for fmt_len, century in ((2, 2000), (4, None)):
        parts = text.split("-")
        if len(parts) == 3 and len(parts[2]) == fmt_len:
            day, month, year = (int(p) for p in parts)
            if century:
                year += century
            try:
                return date(year, month, day)
            except ValueError:
                return None
    return None


def header_field(sheet, row, prefix=""):
    """Text of a header cell, with its 'Mob. No.:' style label stripped off."""
    text = str(sheet.cell(row=row, column=1).value or "").strip()
    if prefix and prefix in text:
        text = text.split(prefix, 1)[1]
    return text.strip()


def read_transactions(sheet, scheme_row, client, category, scheme, folio):
    """Read the Sr./Transaction Type/Date/Amount block printed under a scheme.

    The block starts on the row after the scheme summary line and ends at the
    sub-table's own "Total" line or at the next scheme / category heading.
    """
    out = []
    r = scheme_row + 1
    if str(sheet.cell(row=r, column=1).value or "").strip() != "Sr.":
        return out
    r += 1
    while r <= sheet.max_row:
        sr = sheet.cell(row=r, column=TXN["sr"]).value
        if not isinstance(sr, (int, float)):
            break                      # "Total", a new scheme, or a heading
        out.append({
            "client": client, "category": category, "scheme": scheme, "folio": folio,
            "sr": int(sr),
            "type": str(sheet.cell(row=r, column=TXN["type"]).value or "").strip(),
            "date": parse_date(sheet.cell(row=r, column=TXN["date"]).value),
            **{k: num(sheet.cell(row=r, column=TXN[k]).value) for k in
               ("amount", "pur_nav", "units", "sensex", "cur_value", "days", "stt")},
        })
        r += 1
    return out


def parse_report(path):
    """-> (client dict, [scheme row dicts], grand total dict or None)"""
    sheet = load_workbook(path, data_only=True).worksheets[0]

    # the header prints the name as "NAME (PAN)" or "NAME ()" - drop the bracket
    raw_name = re.sub(r"\s*\([^)]*\)\s*$", "", str(sheet["A2"].value or "").strip())
    name, rep_by = raw_name, ""
    match = re.split(r"\s+REP\s+BY\s+", raw_name, flags=re.IGNORECASE)
    if len(match) == 2:
        name, rep_by = match[0].strip(), match[1].strip()

    client = {
        "name": name, "rep_by": rep_by,
        "pan": header_field(sheet, 3, "Pan :"),
        "address": header_field(sheet, 4, ""),
        "mobile": header_field(sheet, 5, "Mob. No.:"),
        "email": header_field(sheet, 6, "Email:"),
        "dob": parse_date(header_field(sheet, 7, "DOB:")),
    }

    rows, txns, grand, category = [], [], None, None
    for r in range(9, sheet.max_row + 1):
        label = sheet.cell(row=r, column=1).value
        if not isinstance(label, str) or not label.strip():
            continue
        label = label.strip()

        if label == "Grand Total":
            grand = {k: num(sheet.cell(row=r, column=c).value) for k, c in COL.items()}
            continue
        if CATEGORY_RE.match(label) and "Folio" not in label:
            category = label
            continue
        if "Folio" not in label or sheet.cell(row=r, column=COL["since"]).value in (None, ""):
            continue  # transaction line, per-scheme Total, or category total

        parts = FOLIO_RE.split(label.replace("\n", " "), 1)
        scheme, folio = parts[0], (parts[1] if len(parts) > 1 else "")
        scheme_name, folio_no = " ".join(scheme.split()), folio.strip()
        txns.extend(read_transactions(sheet, r, name, category or "",
                                      scheme_name, folio_no))
        rows.append({
            "client": name,
            "category": category or "",
            "scheme": scheme_name,
            "folio": folio_no,
            "since": parse_date(sheet.cell(row=r, column=COL["since"]).value),
            **{k: num(sheet.cell(row=r, column=COL[k]).value) for k in
               ("purchase", "switch_in", "div_reinv", "redemption", "switch_out",
                "div_pay", "cur_value", "cur_units", "cur_nav")},
            "xirr": pct(sheet.cell(row=r, column=COL["xirr"]).value),
            "reported_gain": num(sheet.cell(row=r, column=COL["gain"]).value),
            "reported_abs": pct(sheet.cell(row=r, column=COL["abs_rtn"]).value),
        })
    return client, rows, grand, txns


def check(client, rows, grand, tolerance=1.0):
    """Compare the parsed rows with the report's own Grand Total line."""
    invested = sum(r["purchase"] + r["switch_in"] + r["div_reinv"] for r in rows)
    withdrawn = sum(r["redemption"] + r["switch_out"] + r["div_pay"] for r in rows)
    current = sum(r["cur_value"] for r in rows)
    gain = (current + withdrawn) - invested

    print(f"  schemes parsed        : {len(rows)}")
    print(f"  total invested        : {invested:,.2f}")
    print(f"  withdrawn             : {withdrawn:,.2f}")
    print(f"  current value         : {current:,.2f}")
    print(f"  gain / loss           : {gain:,.2f}"
          f"   ({gain / invested:.2%})" if invested else "")
    if not grand:
        print("  ! no Grand Total line found - totals not cross-checked")
        return True

    ok = True
    for label, got, exp in (
        ("purchase", sum(r["purchase"] for r in rows), grand["purchase"]),
        ("switch in", sum(r["switch_in"] for r in rows), grand["switch_in"]),
        ("redemption", sum(r["redemption"] for r in rows), grand["redemption"]),
        ("switch out", sum(r["switch_out"] for r in rows), grand["switch_out"]),
        ("current value", current, grand["cur_value"]),
        ("gain / loss", gain, grand["gain"]),
    ):
        if abs(got - exp) > tolerance:
            print(f"  MISMATCH {label}: parsed {got:,.2f} vs report {exp:,.2f}")
            ok = False
    print("  cross-check vs report Grand Total: " + ("OK" if ok else "FAILED"))
    return ok


INPUT_FONT = Font(name="Arial", size=10, color="0000FF")


# Input columns only - the calculated columns must keep their formulas.
CLIENT_INPUT_COLS = tuple(range(1, 13))
REGISTER_INPUT_COLS = tuple(range(1, 16))


def blank_row(sheet, row, cols):
    """Wipe a row's input cells and drop the template's italic EXAMPLE styling.

    Needed because openpyxl's ws.cell(value=None) leaves the old value in
    place - writing an empty field would otherwise keep the example text.
    Never touches a calculated column.
    """
    for col in cols:
        cell = sheet.cell(row=row, column=col)
        cell.value = None
        if cell.font.italic:
            cell.font = INPUT_FONT


def add_transactions_sheet(wb, txns):
    """Append every parsed transaction as a plain, sortable table."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    ws = wb.create_sheet("Transactions")
    ws.sheet_view.showGridLines = False
    thin = Side(style="thin", color="BFBFBF")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, text in enumerate(TXN_HEADER, start=1):
        c = ws.cell(row=1, column=i, value=text)
        c.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F3864")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = box
    ws.row_dimensions[1].height = 28

    money = '#,##0.00;(#,##0.00);-'
    for i, t in enumerate(txns):
        r = 2 + i
        for col, val in enumerate(
            [t["client"], t["category"], t["scheme"], t["folio"], t["sr"], t["type"],
             t["date"], t["amount"], t["pur_nav"], t["units"], t["sensex"],
             t["cur_value"], t["days"], t["stt"]], start=1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = Font(name="Arial", size=10)
            c.border = box
            if col == 7:
                c.number_format = 'DD-MMM-YY'
            elif col in (8, 9, 10, 11, 12, 14):
                c.number_format = money
            elif col in (5, 13):
                c.number_format = '#,##0'

    for col, width in zip(range(1, 15),
                          (30, 32, 46, 18, 6, 16, 12, 16, 12, 14, 11, 16, 8, 9)):
        ws.column_dimensions[chr(64 + col) if col <= 26 else "A"].width = width
    ws.freeze_panes = "E2"
    ws.auto_filter.ref = f"A1:N{max(2, len(txns) + 1)}"
    return ws


def write_workbook(out_path, clients, rows, txns=()):
    subprocess.run([sys.executable, str(BUILDER), str(out_path)], check=True,
                   stdout=subprocess.DEVNULL)
    wb = load_workbook(out_path)
    cs, iv = wb["Clients"], wb["Investments"]

    blank_row(cs, 4, CLIENT_INPUT_COLS)
    blank_row(iv, 6, REGISTER_INPUT_COLS)

    for i, c in enumerate(clients):
        r = 4 + i
        blank_row(cs, r, CLIENT_INPUT_COLS)
        for col, val in ((1, f"CL-{i + 1:04d}"), (2, c["name"]), (3, c["pan"]),
                         (4, c["rep_by"]), (6, c["mobile"]), (7, c["email"]),
                         (8, c["dob"]), (9, c["address"])):
            if val:
                cs.cell(row=r, column=col).value = val

    if clients:
        # point the Client Report at a client that now exists
        wb["Client Report"]["C4"].value = clients[0]["name"]

    for i, row in enumerate(rows):
        r = 6 + i
        blank_row(iv, r, REGISTER_INPUT_COLS)
        for key, col in REGISTER:
            val = row[key]
            keep = key in ("client", "category", "scheme", "folio") or val not in (None, "", 0, 0.0)
            if keep and val not in (None, ""):
                iv.cell(row=r, column=col).value = val

    if txns:
        add_transactions_sheet(wb, txns)
    wb.save(out_path)


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        return 2
    out_path = Path(sys.argv[1])
    clients, all_rows, all_txns, all_ok = [], [], [], True

    for report in sys.argv[2:]:
        print(f"\n{Path(report).name}")
        client, rows, grand, txns = parse_report(report)
        print(f"  client                : {client['name']}"
              + (f"  (rep by {client['rep_by']})" if client["rep_by"] else ""))
        all_ok &= check(client, rows, grand)
        print(f"  transactions parsed   : {len(txns)}")
        clients.append(client)
        all_rows.extend(rows)
        all_txns.extend(txns)

    write_workbook(out_path, clients, all_rows, all_txns)

    csv_path = out_path.with_suffix(".csv")
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.writer(fh)
        writer.writerow(CSV_HEADER)
        for row in all_rows:
            writer.writerow([
                row["client"], row["category"], row["scheme"], row["folio"],
                row["since"].strftime("%d-%b-%y") if row["since"] else "",
                row["purchase"], row["switch_in"], row["div_reinv"], row["redemption"],
                row["switch_out"], row["div_pay"], row["cur_value"], row["cur_units"],
                row["cur_nav"], "" if row["xirr"] is None else round(row["xirr"] * 100, 2),
            ])

    print(f"\nwritten: {out_path}  ({len(all_rows)} scheme rows, "
          f"{len(all_txns)} transactions, {len(clients)} client(s))")
    print(f"written: {csv_path}")
    print(f"\nRecalculate {out_path.name} (open and save in Excel, or run the "
          "xlsx recalc script) so the Total Review sheet shows the new figures.")
    return 0 if all_ok else 1


if __name__ == "__main__":
    sys.exit(main())

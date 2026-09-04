"""Build the Mutual Fund Research & Performance Review Excel template.

Usage: python tools/build_fund_research_template.py [output.xlsx]

One research row per scheme per client. Every scheme carries its live
(current) investment value alongside its 3-year return, and the workbook
sorts itself into three verdict tabs:

    3Y return >= Best threshold      -> Best Fund
    3Y return >= Improve threshold   -> Need to Improve
    3Y return <  Improve threshold   -> Not Performing

Both thresholds are editable cells on the Instructions sheet
(default 13% and 10%), so the bands can be re-cut without touching a
formula. Styling, number formats and Indian digit grouping follow
tools/build_client_investment_template.py so the two files sit together.

Sheets: Instructions | Fund Research | Best Funds | Need to Improve |
        Not Performing | Summary | Lists
"""

import datetime
import sys

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------- constants --
FONT = "Arial"
NAVY, BLUE, LIGHT_BLUE = "1F3864", "2E75B6", "DDEBF7"
YELLOW, GREY, WHITE, LINE = "FFF2CC", "F2F2F2", "FFFFFF", "BFBFBF"
GREEN_FILL, AMBER_FILL, RED_FILL = "C6EFCE", "FFEB9C", "FFC7CE"

INPUT_FONT = Font(name=FONT, size=10, color="0000FF")
CALC_FONT = Font(name=FONT, size=10, color="000000")
LINK_FONT = Font(name=FONT, size=10, color="008000")
BODY_FONT = Font(name=FONT, size=10)
BOLD_FONT = Font(name=FONT, size=10, bold=True)
GREEN_FONT = Font(name=FONT, size=10, color="006100")
RED_FONT = Font(name=FONT, size=10, color="9C0006")
AMBER_FONT = Font(name=FONT, size=10, color="9C5700")

# Indian digit grouping, matching the software's 26,65,000.00 style
INR = r'[>=10000000]₹ ##\,##\,##\,##0.00;[>=100000]₹ ##\,##\,##0.00;₹ #,##0.00'
PCT = '0.00%;(0.00%);-'
PCT1 = '0.0%;(0.0%);-'
CNT = '#,##0;(#,##0);-'
DATE = 'DD-MMM-YY'
TEXT = '@'

THIN = Side(style="thin", color=LINE)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

RES_FIRST, RES_LAST = 6, 505      # Fund Research data rows (500 schemes)
OUT_FIRST, OUT_LAST = 6, 205      # verdict tab rows (200 schemes each)

BEST, IMPROVE, POOR, PENDING = (
    "Best Fund", "Need to Improve", "Not Performing", "Data Pending")

R = "'Fund Research'"              # quoted - the sheet name has a space
R_VERDICT = f"{R}!$R${RES_FIRST}:$R${RES_LAST}"
R_3Y = f"{R}!$L${RES_FIRST}:$L${RES_LAST}"
R_INV = f"{R}!$G${RES_FIRST}:$G${RES_LAST}"
R_CUR = f"{R}!$H${RES_FIRST}:$H${RES_LAST}"

ASSET_CLASSES = ["Equity", "Debt", "Hybrid", "Others"]
CATEGORIES = [
    "Equity - Large Cap Fund", "Equity - Large & Mid Cap Fund", "Equity - Mid Cap Fund",
    "Equity - Small Cap Fund", "Equity - Multi Cap Fund", "Equity - Flexi Cap Fund",
    "Equity - Focused Fund", "Equity - Value Fund", "Equity - Contra Fund",
    "Equity - Dividend Yield Fund", "Equity - ELSS", "Equity - Sectoral/ Thematic",
    "Debt - Overnight Fund", "Debt - Liquid Fund", "Debt - Ultra Short Duration Fund",
    "Debt - Low Duration Fund", "Debt - Money Market Fund", "Debt - Short Duration Fund",
    "Debt - Corporate Bond Fund", "Debt - Banking and PSU Fund", "Debt - Credit Risk Fund",
    "Debt - Dynamic Bond", "Debt - Gilt Fund",
    "Hybrid - Aggressive Hybrid Fund",
    "Hybrid - Dynamic Asset Allocation or Balanced Advantage",
    "Hybrid - Multi Asset Allocation", "Hybrid - Conservative Hybrid Fund",
    "Hybrid - Equity Savings", "Hybrid - Arbitrage Fund",
    "Others - Index Funds", "Others - FoF Domestic", "Others - FoF Overseas",
    "Others - Gold ETF", "Others - Other ETFs", "Others - Solution Oriented",
]
ACTIONS = ["Hold", "Top-Up", "Watch", "Switch", "Redeem", "SIP Stop", "Reviewed - No Change"]


# ------------------------------------------------------------------ helpers --
def put(ws, ref, value, *, font=BODY_FONT, fmt=None, fill=None,
        align=None, wrap=False, border=False):
    """Write one cell and style it."""
    c = ws[ref]
    c.value = value
    c.font = font
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    if align or wrap:
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if border:
        c.border = BOX
    return c


def banner(ws, title, subtitle, last_col):
    """Navy title bar + grey subtitle across the sheet width."""
    span = f"A1:{last_col}1"
    ws.merge_cells(span)
    put(ws, "A1", title, font=Font(name=FONT, size=13, bold=True, color=WHITE),
        fill=NAVY, align="left")
    ws.row_dimensions[1].height = 26
    ws.merge_cells(f"A2:{last_col}2")
    put(ws, "A2", subtitle, font=Font(name=FONT, size=9, italic=True, color="404040"),
        fill=GREY, align="left")
    ws.row_dimensions[2].height = 16


def header_row(ws, row, headers, widths):
    """Blue column-header band with borders and set column widths."""
    for i, (text, width) in enumerate(zip(headers, widths), start=1):
        col = get_column_letter(i)
        put(ws, f"{col}{row}", text,
            font=Font(name=FONT, size=10, bold=True, color=WHITE),
            fill=BLUE, align="center", wrap=True, border=True)
        ws.column_dimensions[col].width = width
    ws.row_dimensions[row].height = 32


def verdict_rule(ws, rng, text, fill, font):
    """Colour a range when the row's verdict column equals `text`."""
    ws.conditional_formatting.add(
        rng,
        FormulaRule(formula=[f'$R{RES_FIRST}="{text}"'],
                    fill=PatternFill("solid", fgColor=fill), font=font))


# ------------------------------------------------------------- Instructions --
def build_instructions(wb):
    ws = wb.create_sheet("Instructions")
    ws.sheet_view.showGridLines = False
    banner(ws, "MUTUAL FUND RESEARCH  -  PERFORMANCE REVIEW",
           "Enter every scheme once on Fund Research. The three verdict tabs and the "
           "Summary fill themselves in.", "F")
    for col, width in zip("ABCDEF", (22, 34, 16, 14, 30, 14)):
        ws.column_dimensions[col].width = width

    put(ws, "A4", "HOW TO USE", font=Font(name=FONT, size=11, bold=True, color=NAVY))
    steps = [
        ("1.  Lists", "Nothing to do - it holds the Category and Action dropdowns."),
        ("2.  Fund Research", "The only sheet you type into. One row per scheme per "
                              "client. Fill the blue columns; black columns calculate."),
        ("3.  Best Funds", "Read-only. Every scheme whose 3-year return clears the "
                           "Best threshold, ranked best first."),
        ("4.  Need to Improve", "Read-only. Schemes in the middle band, worst first - "
                                "this is your call-list."),
        ("5.  Not Performing", "Read-only. Schemes below the Improve threshold, worst "
                               "first. Review these with the client."),
        ("6.  Summary", "Read-only dashboard: how many schemes and how much money sit "
                        "in each verdict."),
    ]
    r = 5
    for name, what in steps:
        put(ws, f"A{r}", name, font=BOLD_FONT)
        ws.merge_cells(f"B{r}:F{r}")
        put(ws, f"B{r}", what, wrap=True)
        ws.row_dimensions[r].height = 26
        r += 1

    r += 1
    put(ws, f"A{r}", "COLOUR LEGEND", font=Font(name=FONT, size=11, bold=True, color=NAVY))
    r += 1
    legend = [
        ("Blue text on white", "You type here. These are the only cells to fill in.", INPUT_FONT, WHITE),
        ("Black text on grey", "Calculated. Do not overwrite - it will break the tabs.", CALC_FONT, GREY),
        ("Yellow fill", "Key setting. Safe to change - the whole workbook follows it.", CALC_FONT, YELLOW),
    ]
    for label, what, font, fill in legend:
        put(ws, f"A{r}", label, font=font, fill=fill, border=True)
        ws.merge_cells(f"B{r}:F{r}")
        put(ws, f"B{r}", what, wrap=True)
        r += 1

    # ------------------------------------------------- the rating rule block --
    r += 1
    rule_row = r
    put(ws, f"A{r}", "RATING RULE", font=Font(name=FONT, size=11, bold=True, color=NAVY))
    ws.merge_cells(f"B{r}:F{r}")
    put(ws, f"B{r}", "A fund is judged on its 3-year return. Change the two yellow "
                     "cells to re-cut the bands - every tab re-sorts itself.",
        font=Font(name=FONT, size=9, italic=True, color="404040"), wrap=True)
    r += 1
    for h, w in (("Verdict", None), ("Rule on 3-Year Return", None), ("", None), ("Threshold", None)):
        pass
    put(ws, f"A{r}", "Verdict", font=Font(name=FONT, size=10, bold=True, color=WHITE),
        fill=BLUE, align="center", border=True)
    ws.merge_cells(f"B{r}:C{r}")
    put(ws, f"B{r}", "Rule on 3-Year Return", font=Font(name=FONT, size=10, bold=True, color=WHITE),
        fill=BLUE, align="center", border=True)
    put(ws, f"D{r}", "Threshold", font=Font(name=FONT, size=10, bold=True, color=WHITE),
        fill=BLUE, align="center", border=True)
    r += 1

    best_row = r
    put(ws, f"A{r}", BEST, font=GREEN_FONT, fill=GREEN_FILL, border=True)
    ws.merge_cells(f"B{r}:C{r}")
    put(ws, f"B{r}", "3-year return is greater than or equal to", border=True)
    put(ws, f"D{r}", 0.13, font=INPUT_FONT, fmt=PCT, fill=YELLOW, align="center", border=True)
    r += 1

    imp_row = r
    put(ws, f"A{r}", IMPROVE, font=AMBER_FONT, fill=AMBER_FILL, border=True)
    ws.merge_cells(f"B{r}:C{r}")
    put(ws, f"B{r}", "3-year return is greater than or equal to", border=True)
    put(ws, f"D{r}", 0.10, font=INPUT_FONT, fmt=PCT, fill=YELLOW, align="center", border=True)
    r += 1

    put(ws, f"A{r}", POOR, font=RED_FONT, fill=RED_FILL, border=True)
    ws.merge_cells(f"B{r}:C{r}")
    put(ws, f"B{r}", "3-year return is below", border=True)
    put(ws, f"D{r}", f"=$D${imp_row}", font=CALC_FONT, fmt=PCT, fill=GREY,
        align="center", border=True)
    r += 2

    put(ws, f"A{r}", "Set by", font=BOLD_FONT)
    ws.merge_cells(f"B{r}:F{r}")
    put(ws, f"B{r}", "The 13% Best threshold is the client's own brief: a fund "
                     "returning 13% or more over 3 years is treated as best. The 10% "
                     "Improve floor separates a fund that is lagging from one that is "
                     "genuinely under-performing. Neither is a SEBI or AMFI standard - "
                     "both are house rules, so change them when the view changes.",
        wrap=True)
    ws.row_dimensions[r].height = 46
    r += 2

    put(ws, f"A{r}", "NOTE ON RETURNS", font=Font(name=FONT, size=11, bold=True, color=NAVY))
    r += 1
    ws.merge_cells(f"A{r}:F{r}")
    put(ws, f"A{r}", "Enter 1Y / 3Y / 5Y returns as the scheme's published CAGR. The "
                     "cells are percent-formatted, so typing 13.5 gives 13.50%. Returns "
                     "are not calculated from the invested and current value in this "
                     "workbook - a portfolio holding has its own cash-flow dates, and "
                     "its XIRR is not the scheme's published return. Paste the scheme "
                     "figure from the factsheet, AMFI or your research terminal, and "
                     "note the source date in cell D3 of Fund Research.",
        wrap=True)
    ws.row_dimensions[r].height = 62

    return best_row, imp_row


# ----------------------------------------------------------- Fund Research --
RES_COLS = [
    # header,                     width, kind
    ("Client Name",                 26, "in"),
    ("Scheme Name",                 42, "in"),
    ("Category",                    30, "in"),
    ("Asset Class",                 12, "calc"),
    ("Folio No.",                   16, "in"),
    ("Inv. Since",                  12, "in"),
    ("Total Invested",              16, "in"),
    ("Current Value\n(Live)",       16, "in"),
    ("Gain / Loss",                 15, "calc"),
    ("Abs. Rtn.",                   11, "calc"),
    ("1Y Return",                   11, "in"),
    ("3Y Return\n(CAGR)",           12, "in"),
    ("5Y Return",                   11, "in"),
    ("Category Avg\n3Y",            13, "in"),
    ("Benchmark\n3Y",               12, "in"),
    ("vs Category\n(3Y)",           12, "calc"),
    ("Status",                      10, "calc"),
    ("Verdict",                     17, "calc"),
    ("Suggested Action",            26, "calc"),
    ("Remark",                      46, "in"),
    ("_best",                        3, "calc"),
    ("_improve",                     3, "calc"),
    ("_poor",                        3, "calc"),
]


def build_research(wb, best_row, imp_row):
    ws = wb.create_sheet("Fund Research")
    ws.sheet_view.showGridLines = False
    banner(ws, "FUND RESEARCH REGISTER",
           "One row per scheme per client. Fill the blue columns - A to C, E to H, "
           "K to O and T. Everything else calculates.", "T")

    put(ws, "A3", "Returns as on", font=BOLD_FONT)
    put(ws, "B3", "", font=INPUT_FONT, fmt=DATE, fill=YELLOW, align="center", border=True)
    put(ws, "C3", "Source of returns", font=BOLD_FONT)
    put(ws, "D3", "", font=INPUT_FONT, fill=YELLOW, border=True)
    ws.merge_cells("D3:G3")
    put(ws, "H3", "Schemes entered", font=BOLD_FONT, align="right")
    put(ws, "I3", f"=COUNTA($B${RES_FIRST}:$B${RES_LAST})", font=CALC_FONT,
        fmt=CNT, fill=GREY, align="center", border=True)
    put(ws, "K3", "Rating bands", font=BOLD_FONT, align="right")
    put(ws, "L3", f"=Instructions!$D${best_row}", font=LINK_FONT, fmt=PCT,
        fill=GREY, align="center", border=True)
    put(ws, "M3", "Best", font=Font(name=FONT, size=9, italic=True))
    put(ws, "N3", f"=Instructions!$D${imp_row}", font=LINK_FONT, fmt=PCT,
        fill=GREY, align="center", border=True)
    put(ws, "O3", "Improve floor", font=Font(name=FONT, size=9, italic=True))

    ws.merge_cells("A4:T4")
    put(ws, "A4", "▼  BLUE = you type    ▼  BLACK on grey = calculated, do not overwrite",
        font=Font(name=FONT, size=9, italic=True, color="404040"), fill=GREY)

    header_row(ws, 5, [h for h, _, _ in RES_COLS], [w for _, w, _ in RES_COLS])
    ws.freeze_panes = "C6"

    money = {"G", "H", "I"}
    pcts = {"J", "K", "L", "M", "N", "O", "P"}

    for row in range(RES_FIRST, RES_LAST + 1):
        f = {
            "D": f'=IF($C{row}="","",IFERROR(TRIM(LEFT($C{row},FIND("-",$C{row})-1)),""))',
            "I": f'=IF($B{row}="","",$H{row}-$G{row})',
            "J": f'=IF(OR($B{row}="",$G{row}=0,$G{row}=""),"",$I{row}/$G{row})',
            "P": f'=IF(OR($L{row}="",$N{row}=""),"",$L{row}-$N{row})',
            "Q": f'=IF($B{row}="","",IF(N($H{row})>0,"Live","Exited"))',
            "R": (f'=IF($B{row}="","",'
                  f'IF($L{row}="","{PENDING}",'
                  f'IF($L{row}>=Instructions!$D${best_row},"{BEST}",'
                  f'IF($L{row}>=Instructions!$D${imp_row},"{IMPROVE}","{POOR}"))))'),
            "S": (f'=IF($R{row}="","",'
                  f'IF($R{row}="{PENDING}","Enter the 3Y return",'
                  f'IF($R{row}="{BEST}","Hold / consider top-up",'
                  f'IF($R{row}="{IMPROVE}","Watch 2 quarters, then decide",'
                  f'"Review with client - switch or exit"))))'),
            # unique 1-based rank inside each bucket; the second COUNTIFS breaks ties
            # by row order so no two rows share a rank and none is dropped
            "U": (f'=IF($R{row}<>"{BEST}","",'
                  f'COUNTIFS($R${RES_FIRST}:$R${RES_LAST},"{BEST}",'
                  f'$L${RES_FIRST}:$L${RES_LAST},">"&$L{row})'
                  f'+COUNTIFS($R${RES_FIRST}:$R{row},"{BEST}",'
                  f'$L${RES_FIRST}:$L{row},$L{row}))'),
            "V": (f'=IF($R{row}<>"{IMPROVE}","",'
                  f'COUNTIFS($R${RES_FIRST}:$R${RES_LAST},"{IMPROVE}",'
                  f'$L${RES_FIRST}:$L${RES_LAST},"<"&$L{row})'
                  f'+COUNTIFS($R${RES_FIRST}:$R{row},"{IMPROVE}",'
                  f'$L${RES_FIRST}:$L{row},$L{row}))'),
            "W": (f'=IF($R{row}<>"{POOR}","",'
                  f'COUNTIFS($R${RES_FIRST}:$R${RES_LAST},"{POOR}",'
                  f'$L${RES_FIRST}:$L${RES_LAST},"<"&$L{row})'
                  f'+COUNTIFS($R${RES_FIRST}:$R{row},"{POOR}",'
                  f'$L${RES_FIRST}:$L{row},$L{row}))'),
        }
        for i, (_, _, kind) in enumerate(RES_COLS, start=1):
            col = get_column_letter(i)
            fmt = INR if col in money else PCT if col in pcts else None
            if col == "F":
                fmt = DATE
            if col in f:
                put(ws, f"{col}{row}", f[col], font=CALC_FONT, fmt=fmt,
                    fill=GREY, border=True,
                    align="center" if col in ("D", "Q", "R") else None)
            else:
                put(ws, f"{col}{row}", None, font=INPUT_FONT, fmt=fmt, border=True,
                    align="center" if col in ("F",) else None)
        ws.row_dimensions[row].height = 15

    # one worked example so the expected format is unambiguous - the remark
    # tells the user to delete it, matching the Client Investment Tracker
    example = {
        "A": "Sample Client (EXAMPLE)",
        "B": "Bluechip Flexi Cap Fund - Growth",
        "C": "Equity - Flexi Cap Fund",
        "E": "1910625/49",
        "F": datetime.date(2021, 6, 1),
        "G": 190000,
        "H": 265430,
        "K": 0.1840,
        "L": 0.1512,
        "M": 0.1725,
        "N": 0.1410,
        "O": 0.1365,
        "T": "EXAMPLE ROW - delete it before you start. Returns are the scheme's "
             "published CAGR, not this holding's XIRR.",
    }
    for col, value in example.items():
        ws[f"{col}{RES_FIRST}"].value = value

    # verdict colour bands across the whole row
    band = f"A{RES_FIRST}:T{RES_LAST}"
    verdict_rule(ws, band, BEST, GREEN_FILL, GREEN_FONT)
    verdict_rule(ws, band, IMPROVE, AMBER_FILL, AMBER_FONT)
    verdict_rule(ws, band, POOR, RED_FILL, RED_FONT)

    # dropdowns
    cat = DataValidation(type="list", formula1=f"=Lists!$A$2:$A${len(CATEGORIES) + 1}",
                         allow_blank=True, showDropDown=False)
    ws.add_data_validation(cat)
    cat.add(f"C{RES_FIRST}:C{RES_LAST}")

    for col in "UVW":
        ws.column_dimensions[col].hidden = True

    ws.auto_filter.ref = f"A5:T{RES_LAST}"
    return ws


# ------------------------------------------------------------- verdict tabs --
# (header, source column on Fund Research, number format, width)
OUT_COLS = [
    ("#",                  None, CNT,  5),
    ("Client Name",        "A",  None, 26),
    ("Scheme Name",        "B",  None, 44),
    ("Category",           "C",  None, 30),
    ("Total Invested",     "G",  INR,  16),
    ("Current Value\n(Live)", "H", INR, 16),
    ("Gain / Loss",        "I",  INR,  15),
    ("Abs. Rtn.",          "J",  PCT,  11),
    ("3Y Return\n(CAGR)",  "L",  PCT,  12),
    ("Category Avg\n3Y",   "N",  PCT,  13),
    ("vs Category\n(3Y)",  "P",  PCT,  12),
    ("Suggested Action",   "S",  None, 26),
    ("Remark",             "T",  None, 46),
]


def build_verdict_tab(wb, title, subtitle, rank_col, tint, tint_font, order_note):
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    last = get_column_letter(len(OUT_COLS))
    banner(ws, title.upper(), subtitle, last)

    put(ws, "A3", "Schemes", font=BOLD_FONT)
    put(ws, "B3", f'=COUNTIF({R_VERDICT},"{title_verdict(title)}")',
        font=CALC_FONT, fmt=CNT, fill=GREY, align="center", border=True)
    put(ws, "C3", "Current value", font=BOLD_FONT, align="right")
    put(ws, "D3", f'=SUMIF({R_VERDICT},"{title_verdict(title)}",{R_CUR})',
        font=CALC_FONT, fmt=INR, fill=GREY, border=True)
    put(ws, "E3", "Invested", font=BOLD_FONT, align="right")
    put(ws, "F3", f'=SUMIF({R_VERDICT},"{title_verdict(title)}",{R_INV})',
        font=CALC_FONT, fmt=INR, fill=GREY, border=True)
    put(ws, "G3", "Avg 3Y", font=BOLD_FONT, align="right")
    put(ws, "H3", f'=IFERROR(AVERAGEIF({R_VERDICT},"{title_verdict(title)}",{R_3Y}),"")',
        font=CALC_FONT, fmt=PCT, fill=GREY, align="center", border=True)

    ws.merge_cells(f"A4:{last}4")
    put(ws, "A4", order_note,
        font=Font(name=FONT, size=9, italic=True, color="404040"), fill=GREY)

    header_row(ws, 5, [h for h, _, _, _ in OUT_COLS], [w for _, _, _, w in OUT_COLS])
    ws.freeze_panes = "C6"

    for row in range(OUT_FIRST, OUT_LAST + 1):
        n = row - OUT_FIRST + 1
        for i, (_, src, fmt, _) in enumerate(OUT_COLS, start=1):
            col = get_column_letter(i)
            if src is None:
                v = (f'=IF(INDEX({R}!$B${RES_FIRST}:$B${RES_LAST},'
                     f'MATCH({n},{R}!${rank_col}${RES_FIRST}:${rank_col}${RES_LAST},0))="","",{n})')
                v = f'=IFERROR({v[1:]},"")'
            else:
                v = (f'=IFERROR(INDEX({R}!${src}${RES_FIRST}:${src}${RES_LAST},'
                     f'MATCH({n},{R}!${rank_col}${RES_FIRST}:${rank_col}${RES_LAST},0)),"")')
            put(ws, f"{col}{row}", v, font=CALC_FONT, fmt=fmt, border=True,
                align="center" if col == "A" else None)
        ws.row_dimensions[row].height = 15

    # tint the 3Y return column so the ranking reads at a glance
    for row in range(OUT_FIRST, OUT_LAST + 1):
        c = ws[f"I{row}"]
        c.fill = PatternFill("solid", fgColor=tint)
        c.font = tint_font

    ws.auto_filter.ref = f"A5:{last}{OUT_LAST}"
    return ws


def title_verdict(title):
    return {"Best Funds": BEST, "Need to Improve": IMPROVE,
            "Not Performing": POOR}[title]


# ----------------------------------------------------------------- Summary --
def build_summary(wb, best_row, imp_row):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    banner(ws, "PERFORMANCE SUMMARY",
           "Every figure reads from Fund Research. Nothing here is typed.", "H")
    for col, width in zip("ABCDEFGH", (4, 22, 12, 20, 20, 20, 12, 12)):
        ws.column_dimensions[col].width = width

    put(ws, "B4", "PORTFOLIO BY VERDICT",
        font=Font(name=FONT, size=11, bold=True, color=NAVY))

    heads = ["Verdict", "Schemes", "Total Invested", "Current Value (Live)",
             "Gain / Loss", "% of Value", "Avg 3Y"]
    for i, h in enumerate(heads):
        col = get_column_letter(2 + i)
        put(ws, f"{col}5", h, font=Font(name=FONT, size=10, bold=True, color=WHITE),
            fill=BLUE, align="center", wrap=True, border=True)
    ws.row_dimensions[5].height = 30

    rows = [(BEST, GREEN_FILL, GREEN_FONT), (IMPROVE, AMBER_FILL, AMBER_FONT),
            (POOR, RED_FILL, RED_FONT), (PENDING, GREY, BODY_FONT)]
    first = 6
    for j, (verdict, fill, font) in enumerate(rows):
        r = first + j
        put(ws, f"B{r}", verdict, font=font, fill=fill, border=True)
        put(ws, f"C{r}", f'=COUNTIF({R_VERDICT},$B{r})', font=CALC_FONT,
            fmt=CNT, align="center", border=True)
        put(ws, f"D{r}", f'=SUMIF({R_VERDICT},$B{r},{R_INV})',
            font=CALC_FONT, fmt=INR, border=True)
        put(ws, f"E{r}", f'=SUMIF({R_VERDICT},$B{r},{R_CUR})',
            font=CALC_FONT, fmt=INR, border=True)
        put(ws, f"F{r}", f'=IF($C{r}=0,"",$E{r}-$D{r})', font=CALC_FONT,
            fmt=INR, border=True)
        put(ws, f"G{r}", f'=IF($E${first + 4}=0,"",$E{r}/$E${first + 4})',
            font=CALC_FONT, fmt=PCT1, align="center", border=True)
        put(ws, f"H{r}", f'=IFERROR(AVERAGEIF({R_VERDICT},$B{r},{R_3Y}),"")',
            font=CALC_FONT, fmt=PCT, align="center", border=True)

    tot = first + 4
    put(ws, f"B{tot}", "TOTAL", font=Font(name=FONT, size=10, bold=True, color=WHITE),
        fill=NAVY, border=True)
    for col, fmt in (("C", CNT), ("D", INR), ("E", INR), ("F", INR)):
        put(ws, f"{col}{tot}", f"=SUM({col}{first}:{col}{tot - 1})",
            font=Font(name=FONT, size=10, bold=True, color=WHITE), fmt=fmt,
            fill=NAVY, border=True,
            align="center" if col == "C" else None)
    put(ws, f"G{tot}", f'=IF($E{tot}=0,"",1)',
        font=Font(name=FONT, size=10, bold=True, color=WHITE), fmt=PCT1,
        fill=NAVY, align="center", border=True)
    put(ws, f"H{tot}", f'=IFERROR(AVERAGE({R_3Y}),"")',
        font=Font(name=FONT, size=10, bold=True, color=WHITE), fmt=PCT,
        fill=NAVY, align="center", border=True)

    r = tot + 2
    put(ws, f"B{r}", "HEADLINE NUMBERS",
        font=Font(name=FONT, size=11, bold=True, color=NAVY))
    r += 1
    lines = [
        ("Schemes reviewed", f'=COUNTA({R}!$B${RES_FIRST}:$B${RES_LAST})', CNT),
        ("Live schemes", f'=COUNTIF({R}!$Q${RES_FIRST}:$Q${RES_LAST},"Live")', CNT),
        ("Best-fund hit rate",
         f'=IF($C{tot}=0,"",$C{first}/$C{tot})', PCT1),
        ("Money needing action",
         f'=$E{first + 1}+$E{first + 2}', INR),
        ("Share of value needing action",
         f'=IF($E{tot}=0,"",($E{first + 1}+$E{first + 2})/$E{tot})', PCT1),
    ]
    for label, formula, fmt in lines:
        put(ws, f"B{r}", label, font=BOLD_FONT, border=True)
        ws.merge_cells(f"B{r}:C{r}")
        put(ws, f"D{r}", formula, font=CALC_FONT, fmt=fmt, fill=GREY,
            align="center", border=True)
        r += 1

    r += 1
    put(ws, f"B{r}", "Rule in force", font=BOLD_FONT)
    ws.merge_cells(f"C{r}:H{r}")
    put(ws, f"C{r}",
        f'=_xlfn.CONCAT("Best = 3Y return at or above ",'
        f'TEXT(Instructions!$D${best_row},"0.0%"),'
        f'"   |   Need to Improve = ",TEXT(Instructions!$D${imp_row},"0.0%"),'
        f'" to ",TEXT(Instructions!$D${best_row},"0.0%"),'
        f'"   |   Not Performing = below ",TEXT(Instructions!$D${imp_row},"0.0%"))',
        font=LINK_FONT)

    chart = BarChart()
    chart.type = "col"
    chart.title = "Current value by verdict"
    chart.y_axis.title = "₹"
    chart.height, chart.width = 7.5, 14
    data = Reference(ws, min_col=5, min_row=5, max_row=tot - 1)
    cats = Reference(ws, min_col=2, min_row=first, max_row=tot - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None
    ws.add_chart(chart, f"B{r + 2}")
    return ws


# ------------------------------------------------------------------- Lists --
def build_lists(wb):
    ws = wb.create_sheet("Lists")
    heads = ["Category  (Asset - Category)", "Asset Class", "Verdict", "Action"]
    for i, h in enumerate(heads, start=1):
        col = get_column_letter(i)
        put(ws, f"{col}1", h, font=Font(name=FONT, size=10, bold=True, color=WHITE),
            fill=BLUE, border=True)
        ws.column_dimensions[col].width = 34 if i == 1 else 22
    for i, v in enumerate(CATEGORIES, start=2):
        put(ws, f"A{i}", v)
    for i, v in enumerate(ASSET_CLASSES, start=2):
        put(ws, f"B{i}", v)
    for i, v in enumerate([BEST, IMPROVE, POOR, PENDING], start=2):
        put(ws, f"C{i}", v)
    for i, v in enumerate(ACTIONS, start=2):
        put(ws, f"D{i}", v)
    return ws


# -------------------------------------------------------------------- main --
def main(out="Mutual_Fund_Research_Template.xlsx"):
    wb = Workbook()
    wb.remove(wb.active)

    best_row, imp_row = build_instructions(wb)
    build_research(wb, best_row, imp_row)
    build_verdict_tab(
        wb, "Best Funds",
        "Every scheme clearing the Best threshold on 3-year return.",
        "U", GREEN_FILL, GREEN_FONT,
        "Ranked best first by 3-year return. Read-only - edit on Fund Research.")
    build_verdict_tab(
        wb, "Need to Improve",
        "Schemes in the middle band - lagging, not yet failing. Your call-list.",
        "V", AMBER_FILL, AMBER_FONT,
        "Ranked worst first by 3-year return. Read-only - edit on Fund Research.")
    build_verdict_tab(
        wb, "Not Performing",
        "Schemes below the Improve floor on 3-year return. Review with the client.",
        "W", RED_FILL, RED_FONT,
        "Ranked worst first by 3-year return. Read-only - edit on Fund Research.")
    build_summary(wb, best_row, imp_row)
    build_lists(wb)

    wb.active = 0
    wb.save(out)
    print(f"wrote {out}")


if __name__ == "__main__":
    main(*(sys.argv[1:2] or []))

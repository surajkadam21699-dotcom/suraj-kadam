"""Build the Client Investment - Total Review Excel template.

Usage: python tools/build_client_investment_template.py [output.xlsx]

Column set and maths follow the Opulencia Capital scheme-wise Profit & Loss
export, so a report pulled from the software pastes straight into the
Investments sheet:

    Total Invested = Purchase + Switch In + Div Reinv
    Withdrawn      = Redemption + Switch Out + Div Pay
    Gain / Loss    = (Cur. Value + Withdrawn) - Total Invested
    Abs. Rtn.      = Gain / Loss / Total Invested

(Verified against the sample report's Debt / Equity / Grand Total rows.)

Sheets: Instructions | Clients | Investments | Total Review | Client Report | Lists
"""

import sys

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------- constants --
FONT = "Arial"
NAVY, BLUE, LIGHT_BLUE = "1F3864", "2E75B6", "DDEBF7"
YELLOW, GREY, WHITE, LINE = "FFF2CC", "F2F2F2", "FFFFFF", "BFBFBF"

INPUT_FONT = Font(name=FONT, size=10, color="0000FF")
CALC_FONT = Font(name=FONT, size=10, color="000000")
BODY_FONT = Font(name=FONT, size=10)
GREEN_FONT = Font(name=FONT, size=10, color="006100")
RED_FONT = Font(name=FONT, size=10, color="9C0006")

# Indian digit grouping, matching the software's 26,65,000.00 style
INR = r'[>=10000000]₹ ##\,##\,##\,##0.00;[>=100000]₹ ##\,##\,##0.00;₹ #,##0.00'
UNITS = '#,##0.0000;(#,##0.0000);-'
NAV = '#,##0.0000;(#,##0.0000);-'
PCT = '0.00%;(0.00%);-'
CNT = '#,##0;(#,##0);-'
DECIMAL = '#,##0.00;(#,##0.00);-'
DATE = 'DD-MMM-YY'
TEXT = '@'

THIN = Side(style="thin", color=LINE)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

import os
CLI_FIRST, CLI_LAST = 4, int(os.environ.get("N_CLI", 203))      # Clients data rows (200 clients)
INV_FIRST, INV_LAST = 6, int(os.environ.get("N_INV", 1005))     # Investments data rows (1000 schemes)
RPT_FIRST, RPT_LAST = 22, int(os.environ.get("N_RPT", 121))     # Client Report scheme rows (100 schemes)

ASSET_CLASSES = ["Equity", "Debt", "Hybrid", "Others"]
CATEGORIES = [
    "Equity - Large Cap Fund", "Equity - Large & Mid Cap Fund", "Equity - Mid Cap Fund",
    "Equity - Small Cap Fund", "Equity - Multi Cap Fund", "Equity - Flexi Cap Fund",
    "Equity - Focused Fund", "Equity - Value Fund", "Equity - Contra Fund",
    "Equity - Dividend Yield Fund", "Equity - ELSS", "Equity - Sectoral/ Thematic",
    "Debt - Overnight Fund", "Debt - Liquid Fund", "Debt - Ultra Short Duration Fund",
    "Debt - Low Duration Fund", "Debt - Money Market Fund", "Debt - Short Duration Fund",
    "Debt - Corporate Bond Fund", "Debt - Banking and PSU Fund", "Debt - Credit Risk Fund",
    "Debt - Dynamic Bond", "Debt - Gilt Fund",
    "Hybrid - Aggressive Hybrid Fund",
    "Hybrid - Dynamic Asset Allocation or Balanced Advantage",
    "Hybrid - Multi Asset Allocation", "Hybrid - Conservative Hybrid Fund",
    "Hybrid - Equity Savings", "Hybrid - Arbitrage Fund",
    "Others - Index Funds", "Others - FoF Domestic", "Others - FoF Overseas",
    "Others - Gold ETF", "Others - Other ETFs", "Others - Solution Oriented",
]
STATUSES = ["Live", "Exited"]
SEGMENTS = ["Retail", "HNI", "Ultra HNI", "Corporate", "NRI"]

# Named ranges used by every rollup formula
C_NAME = f"Clients!$B${CLI_FIRST}:$B${CLI_LAST}"
I_CLI = f"Investments!$A${INV_FIRST}:$A${INV_LAST}"
I_CAT = f"Investments!$B${INV_FIRST}:$B${INV_LAST}"
I_PUR = f"Investments!$F${INV_FIRST}:$F${INV_LAST}"
I_SWIN = f"Investments!$G${INV_FIRST}:$G${INV_LAST}"
I_DIVR = f"Investments!$H${INV_FIRST}:$H${INV_LAST}"
I_RED = f"Investments!$I${INV_FIRST}:$I${INV_LAST}"
I_SWOUT = f"Investments!$J${INV_FIRST}:$J${INV_LAST}"
I_DIVP = f"Investments!$K${INV_FIRST}:$K${INV_LAST}"
I_CUR = f"Investments!$L${INV_FIRST}:$L${INV_LAST}"
I_XIRR = f"Investments!$O${INV_FIRST}:$O${INV_LAST}"
I_ASSET = f"Investments!$P${INV_FIRST}:$P${INV_LAST}"
I_INVSD = f"Investments!$Q${INV_FIRST}:$Q${INV_LAST}"
I_STAT = f"Investments!$U${INV_FIRST}:$U${INV_LAST}"

wb = Workbook()


# ------------------------------------------------------------------ helpers --
def banner(ws, text, subtitle, last_col):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    c = ws.cell(row=1, column=1, value=text)
    c.font = Font(name=FONT, size=14, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 26
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    s = ws.cell(row=2, column=1, value=subtitle)
    s.font = Font(name=FONT, size=9, italic=True, color="595959")
    s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 16


def head(ws, row, headers, start_col=1, fill=BLUE, height=34):
    for i, text in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=text)
        c.font = Font(name=FONT, size=9, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BOX
    ws.row_dimensions[row].height = height


def widths(ws, mapping):
    for col, w in mapping.items():
        ws.column_dimensions[col].width = w


def note(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name=FONT, size=9, italic=True, color="808080")
    return c


def style_body(cell, fmt, kind="calc", stripe=None):
    cell.font = INPUT_FONT if kind == "input" else CALC_FONT
    cell.fill = PatternFill(
        "solid", fgColor=YELLOW if kind == "input" else (stripe or GREY))
    cell.number_format = fmt
    cell.border = BOX
    cell.alignment = Alignment(vertical="center")


# ============================================================== Instructions ==
ins = wb.active
ins.title = "Instructions"
ins.sheet_view.showGridLines = False
banner(ins, "CLIENT INVESTMENT TRACKER  -  TOTAL REVIEW",
       "Built to match the scheme-wise Profit & Loss export from the portfolio software.", 6)
widths(ins, {"A": 3, "B": 30, "C": 88, "D": 12, "E": 12, "F": 12})

CONTENT = [
    ("HOW TO USE", None),
    ("1. Clients", "Enter every client once - code, name, PAN, address, mobile, email, DOB, 'Rep By' / family head. This list drives the dropdowns, the client-wise review and the Client Report header."),
    ("2. Import Report", "The easy way to load data. Paste a whole Profit & Loss export on the left of that sheet; clean register rows appear on the right, ready to copy into Investments. Category headings, transaction sub-tables and total lines are dropped for you, and the scheme name and folio are split apart."),
    ("3. Investments", "One row per scheme per client. Columns A to O are one unbroken block, in the same order as the Import Report output, so a whole report goes in with a single Paste Special > Values on cell A6. Columns P to V are formulas - never paste over them."),
    ("4. Total Review", "Read-only dashboard: firm-level snapshot, client-wise review, asset-class review and category-wise review, with charts."),
    ("5. Client Report", "Pick one client from the dropdown to get that client's header block, summary and scheme list - the single-client view your software prints."),
    (None, None),
    ("HOW THE MATHS WORKS", None),
    ("Total Invested", "Purchase + Switch In + Div Reinv"),
    ("Withdrawn", "Redemption + Switch Out + Div Pay"),
    ("Gain / Loss", "(Cur. Value + Withdrawn) - Total Invested"),
    ("Abs. Rtn.", "Gain / Loss ÷ Total Invested"),
    ("Net Invested", "Total Invested - Withdrawn (money still at work; shown for reference)"),
    ("Status", "Live when Cur. Value is above zero, otherwise Exited."),
    ("XIRR", "Typed in from the software - it needs every cash-flow date, which this sheet does not hold. Client, category and firm level XIRR is a current-value-weighted average of the scheme XIRRs, so treat it as indicative, not as a recomputed XIRR."),
    (None, None),
    ("WHAT YOU EDIT", None),
    ("Blue text on yellow fill", "Input cells - type or paste here."),
    ("Black text on grey fill", "Formula cells - do not overwrite, they will stop calculating."),
    ("EXAMPLE rows", "Clients row 4 and Investments row 6 are filled in as a format guide. Delete both before real use."),
    ("Paste Special > Values", "Use this rather than a plain Ctrl+V when pasting into Investments: a plain paste drags the source formatting across and can wreck the column widths, fills and dropdowns."),
    (None, None),
    ("GOOD TO KNOW", None),
    ("Capacity", f"{CLI_LAST - CLI_FIRST + 1} clients, {INV_LAST - INV_FIRST + 1} scheme rows, {RPT_LAST - RPT_FIRST + 1} schemes per client in the Client Report. To add rows, insert them INSIDE the existing table so the formulas extend with them."),
    ("Asset Class", "Derived from the text before the ' - ' in Category, which is why categories are named 'Equity - Flexi Cap Fund' exactly as in the software."),
    ("New categories", "Add them on the Lists sheet; the dropdown and the category-wise review pick them up. Keep the 'Asset - Category' naming."),
    ("Currency", "Indian Rupee with lakh / crore grouping. To switch, select the amount columns > Format Cells > Custom and edit the symbol."),
    ("Blank amounts", "Leave a column blank where the software prints 0.00 - blanks and zeros both total correctly."),
]

r = 4
for label, text in CONTENT:
    if label is None:
        r += 1
        continue
    if text is None:
        c = ins.cell(row=r, column=2, value=label)
        c.font = Font(name=FONT, size=11, bold=True, color=NAVY)
        r += 1
        continue
    lc = ins.cell(row=r, column=2, value=label)
    lc.font = Font(name=FONT, size=10, bold=True)
    lc.alignment = Alignment(vertical="top")
    tc = ins.cell(row=r, column=3, value=text)
    tc.font = BODY_FONT
    tc.alignment = Alignment(vertical="top", wrap_text=True)
    ins.row_dimensions[r].height = 30
    r += 1

for swatch_row, colour in ((r - 8, YELLOW), (r - 7, GREY)):
    sw = ins.cell(row=swatch_row, column=4)
    sw.fill = PatternFill("solid", fgColor=colour)
    sw.border = BOX

# ==================================================================== Lists ==
ls = wb.create_sheet("Lists")
ls.sheet_view.showGridLines = False
head(ls, 1, ["Category  (Asset - Category)", "Asset Class", "Status", "Client Segment"])
for i, v in enumerate(CATEGORIES):
    ls.cell(row=2 + i, column=1, value=v).font = BODY_FONT
for i, v in enumerate(ASSET_CLASSES):
    ls.cell(row=2 + i, column=2, value=v).font = BODY_FONT
for i, v in enumerate(STATUSES):
    ls.cell(row=2 + i, column=3, value=v).font = BODY_FONT
for i, v in enumerate(SEGMENTS):
    ls.cell(row=2 + i, column=4, value=v).font = BODY_FONT
widths(ls, {"A": 50, "B": 16, "C": 14, "D": 18})
note(ls, len(CATEGORIES) + 3, 1,
     "Add or rename values here - dropdowns and the category-wise review follow this list. Keep the 'Asset - Category' naming so the Asset Class is picked up.")

CAT_RANGE = f"Lists!$A$2:$A${1 + len(CATEGORIES)}"
ASSET_RANGE = f"Lists!$B$2:$B${1 + len(ASSET_CLASSES)}"
SEG_RANGE = f"Lists!$D$2:$D${1 + len(SEGMENTS)}"

# ================================================================== Clients ==
cs = wb.create_sheet("Clients")
cs.sheet_view.showGridLines = False
banner(cs, "CLIENT MASTER",
       "One row per client. The name entered here is the key used by every other sheet.", 12)
head(cs, 3, ["Client Code", "Client Name", "PAN", "Rep By / Family Head", "Segment",
             "Mobile No.", "Email", "DOB", "Address", "Advisor / RM",
             "Onboarding Date", "Remarks"])
widths(cs, {"A": 13, "B": 34, "C": 14, "D": 30, "E": 12, "F": 15, "G": 28,
            "H": 13, "I": 46, "J": 20, "K": 15, "L": 26})
cs.freeze_panes = "C4"

for row in range(CLI_FIRST, CLI_LAST + 1):
    for col in range(1, 13):
        style_body(cs.cell(row=row, column=col), TEXT, "input")
    cs.cell(row=row, column=8).number_format = DATE
    cs.cell(row=row, column=11).number_format = DATE

EX_CLIENT = ["CL-0001", "Sample Client (EXAMPLE)", "ABCDE1234F",
             "Rep By Family Head Name", "HNI", "98xxxxxx61",
             "client@example.com", "2008-06-05",
             "Flat 301, Sample Residency, Kothrud, Pune 411038", "S. Kadam",
             "2020-08-31", "Delete this EXAMPLE row before use"]
for col, val in enumerate(EX_CLIENT, start=1):
    c = cs.cell(row=CLI_FIRST, column=col, value=val)
    c.font = Font(name=FONT, size=10, color="0000FF", italic=True)
cs.cell(row=CLI_FIRST, column=8).number_format = DATE
cs.cell(row=CLI_FIRST, column=11).number_format = DATE

dv_seg = DataValidation(type="list", formula1=SEG_RANGE, allow_blank=True)
cs.add_data_validation(dv_seg)
dv_seg.add(f"E{CLI_FIRST}:E{CLI_LAST}")

# ============================================================== Investments ==
iv = wb.create_sheet("Investments")
iv.sheet_view.showGridLines = False
banner(iv, "SCHEME-WISE INVESTMENT REGISTER",
       "One row per scheme per client. Columns A to O are one unbroken paste block; P to V calculate themselves.", 22)
head(iv, 5, [
    "Client Name", "Category", "Scheme Name", "Folio No.", "Inv. Since",
    "Purchase", "Switch In", "Div Reinv", "Redemption", "Switch Out",
    "Div Pay", "Cur. Value", "Cur. Units", "Cur. NAV", "XIRR",
    "Asset Class", "Total Invested", "Withdrawn", "Gain / Loss", "Abs. Rtn.",
    "Status", "Rpt. Idx",
])
widths(iv, {"A": 30, "B": 32, "C": 46, "D": 18, "E": 12, "F": 16, "G": 16,
            "H": 13, "I": 15, "J": 15, "K": 12, "L": 17, "M": 14, "N": 13,
            "O": 10, "P": 12, "Q": 17, "R": 16, "S": 16, "T": 11, "U": 10, "V": 9})
iv.freeze_panes = "C6"

iv.cell(row=4, column=1, value="▼ A to O are one block - paste all 15 columns in one go "
                               "(the Import Report sheet prepares them for you)").font = \
    Font(name=FONT, size=8, italic=True, color="808080")
iv.cell(row=4, column=16, value="▼ calculated - never paste over these").font = \
    Font(name=FONT, size=8, italic=True, color="808080")

# A-O are typed or pasted; P-V are formulas. Keeping every input column in one
# unbroken block is what lets a whole report paste in with a single Ctrl+V.
INPUT_SPEC = {1: TEXT, 2: TEXT, 3: TEXT, 4: TEXT, 5: DATE, 6: INR, 7: INR,
              8: INR, 9: INR, 10: INR, 11: INR, 12: INR, 13: UNITS, 14: NAV,
              15: PCT}
CALC_SPEC = {16: TEXT, 17: INR, 18: INR, 19: INR, 20: PCT, 21: TEXT, 22: CNT}

for row in range(INV_FIRST, INV_LAST + 1):
    for col, fmt in INPUT_SPEC.items():
        style_body(iv.cell(row=row, column=col), fmt, "input")
    for col, fmt in CALC_SPEC.items():
        style_body(iv.cell(row=row, column=col), fmt, "calc")

    iv.cell(row=row, column=16,
            value=f'=IF($B{row}="","",IFERROR(TRIM(LEFT($B{row},FIND(" - ",$B{row})-1)),"Others"))')
    iv.cell(row=row, column=17,
            value=f'=IF($A{row}="","",SUM($F{row}:$H{row}))')
    iv.cell(row=row, column=18,
            value=f'=IF($A{row}="","",SUM($I{row}:$K{row}))')
    iv.cell(row=row, column=19,
            value=f'=IF($A{row}="","",($L{row}+$R{row})-$Q{row})')
    iv.cell(row=row, column=20,
            value=f'=IF(OR($A{row}="",$Q{row}=0),"",$S{row}/$Q{row})')
    iv.cell(row=row, column=21,
            value=f'=IF($A{row}="","",IF($L{row}>0,"Live","Exited"))')
    iv.cell(row=row, column=22,
            value=(f'=IF($A{row}="","",IF($A{row}=\'Client Report\'!$C$4,'
                   f'COUNTIFS($A${INV_FIRST}:$A{row},\'Client Report\'!$C$4),""))'))
    iv.cell(row=row, column=22).font = Font(name=FONT, size=9, color="A6A6A6")

dv_cli = DataValidation(type="list", formula1=C_NAME, allow_blank=True,
                        showErrorMessage=True, errorTitle="Unknown client",
                        error="Add this client on the Clients sheet first.")
dv_cat = DataValidation(type="list", formula1=CAT_RANGE, allow_blank=True,
                        showErrorMessage=False)
iv.add_data_validation(dv_cli)
iv.add_data_validation(dv_cat)
dv_cli.add(f"A{INV_FIRST}:A{INV_LAST}")
dv_cat.add(f"B{INV_FIRST}:B{INV_LAST}")

for col in ("O", "S", "T"):
    rng = f"{col}{INV_FIRST}:{col}{INV_LAST}"
    iv.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["0"], font=RED_FONT))
    iv.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["0"], font=GREEN_FONT))

EX_INV = {1: "Sample Client (EXAMPLE)", 2: "Debt - Low Duration Fund",
          3: "Bluechip Low Duration Fund - Regular Plan - Growth", 4: "1910625/49",
          5: "2021-06-01", 7: 190000, 10: 195243.79, 14: 41.2489, 15: 0.038}
for col, val in EX_INV.items():
    c = iv.cell(row=INV_FIRST, column=col, value=val)
    c.font = Font(name=FONT, size=10, color="0000FF", italic=True)
iv.cell(row=INV_FIRST, column=5).number_format = DATE
note(iv, INV_LAST + 2, 1,
     "EXAMPLE row above - delete it before real use. Leave a cell blank wherever the software prints 0.00.")

# ============================================================= Total Review ==
tr = wb.create_sheet("Total Review")
tr.sheet_view.showGridLines = False
tr.sheet_view.zoomScale = 90
banner(tr, "TOTAL INVESTMENT REVIEW  -  ALL CLIENTS",
       "Every figure is calculated from the Investments sheet. Nothing on this sheet needs typing.", 12)
widths(tr, {"A": 3, "B": 34, "C": 13, "D": 17, "E": 17, "F": 18, "G": 17,
            "H": 18, "I": 18, "J": 11, "K": 11, "L": 11})

MONEY_COLS = [(4, INR), (5, INR), (6, INR), (7, INR), (8, INR), (9, INR),
              (10, PCT), (11, PCT), (12, PCT)]
TABLE_HEAD = ["Client Name", "Schemes", "Purchase", "Switch In", "Total Invested",
              "Withdrawn", "Cur. Value", "Gain / Loss", "Abs. Rtn.", "Wtd. XIRR",
              "% of AUM"]


def section(row, text, last_col=12):
    tr.merge_cells(start_row=row, start_column=2, end_row=row, end_column=last_col)
    c = tr.cell(row=row, column=2, value=text)
    c.font = Font(name=FONT, size=11, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(vertical="center", indent=1)
    tr.row_dimensions[row].height = 20


def group_table(head_row, source_ref, key_range, first_src,
                n_rows, label_head, total_label):
    """Build a rollup table keyed on `key_range` of the Investments sheet."""
    head(tr, head_row, [label_head] + TABLE_HEAD[1:], start_col=2, height=30)
    first = head_row + 1
    last = first + n_rows - 1
    for i in range(n_rows):
        row = first + i
        src = first_src + i
        stripe = WHITE if i % 2 == 0 else GREY
        tr.cell(row=row, column=2, value=f'=IF({source_ref}{src}="","",{source_ref}{src})')
        tr.cell(row=row, column=3, value=f'=IF($B{row}="","",COUNTIFS({key_range},$B{row}))')
        tr.cell(row=row, column=4, value=f'=IF($B{row}="","",SUMIFS({I_PUR},{key_range},$B{row}))')
        tr.cell(row=row, column=5, value=f'=IF($B{row}="","",SUMIFS({I_SWIN},{key_range},$B{row}))')
        tr.cell(row=row, column=6, value=f'=IF($B{row}="","",SUMIFS({I_INVSD},{key_range},$B{row}))')
        tr.cell(row=row, column=7, value=(
            f'=IF($B{row}="","",SUMIFS({I_RED},{key_range},$B{row})'
            f'+SUMIFS({I_SWOUT},{key_range},$B{row})+SUMIFS({I_DIVP},{key_range},$B{row}))'))
        tr.cell(row=row, column=8, value=f'=IF($B{row}="","",SUMIFS({I_CUR},{key_range},$B{row}))')
        tr.cell(row=row, column=9, value=f'=IF($B{row}="","",($H{row}+$G{row})-$F{row})')
        tr.cell(row=row, column=10, value=f'=IF(OR($B{row}="",$F{row}=0),"",$I{row}/$F{row})')
        tr.cell(row=row, column=11, value=(
            f'=IF(OR($B{row}="",$H{row}=0),"",'
            f'SUMPRODUCT(({key_range}=$B{row})*{I_CUR}*{I_XIRR})/$H{row})'))
        tr.cell(row=row, column=12, value=f'=IF(OR($B{row}="",$C$12=0),"",$H{row}/$C$12)')
        style_body(tr.cell(row=row, column=2), TEXT, "calc", stripe)
        style_body(tr.cell(row=row, column=3), CNT, "calc", stripe)
        for col, fmt in MONEY_COLS:
            style_body(tr.cell(row=row, column=col), fmt, "calc", stripe)

    total = last + 1
    tr.cell(row=total, column=2, value=total_label)
    for col in range(3, 9):
        letter = tr.cell(row=first, column=col).column_letter
        tr.cell(row=total, column=col, value=f'=SUM(${letter}${first}:${letter}${last})')
    tr.cell(row=total, column=9, value=f'=($H${total}+$G${total})-$F${total}')
    tr.cell(row=total, column=10, value=f'=IF($F${total}=0,"",$I${total}/$F${total})')
    tr.cell(row=total, column=11, value=f'=IF($H${total}=0,"",SUMPRODUCT({I_CUR},{I_XIRR})/$H${total})')
    tr.cell(row=total, column=12, value=f'=IF($C$12=0,"",$H${total}/$C$12)')
    for col, fmt in [(2, TEXT), (3, CNT)] + MONEY_COLS:
        c = tr.cell(row=total, column=col)
        c.font = Font(name=FONT, size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.number_format = fmt
        c.border = BOX
    for col in ("I", "J", "K"):
        rng = f"{col}{first}:{col}{last}"
        tr.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["0"], font=RED_FONT))
        tr.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["0"], font=GREEN_FONT))
    return first, last, total


# --- snapshot ---------------------------------------------------------------
section(4, "PORTFOLIO SNAPSHOT  -  ALL CLIENTS")
SNAPSHOT = [
    ("Total Clients", f'=COUNTA({C_NAME})', CNT),
    ("Total Schemes", f'=COUNTIF({I_CLI},"?*")', CNT),
    ("Live Schemes", f'=COUNTIF({I_STAT},"Live")', CNT),
    ("Total Purchase", f'=SUM({I_PUR})', INR),
    ("Total Switch In", f'=SUM({I_SWIN})', INR),
    ("Total Invested  (Purchase + Switch In + Div Reinv)", f'=SUM({I_INVSD})', INR),
    ("Total Withdrawn  (Redemption + Switch Out + Div Pay)", f'=SUM({I_RED})+SUM({I_SWOUT})+SUM({I_DIVP})', INR),
    ("Current Value  (AUM)", f'=SUM({I_CUR})', INR),
    ("Total Gain / Loss", '=($C$12+$C$11)-$C$10', INR),
    ("Overall Abs. Rtn.", '=IF($C$10=0,"",$C$13/$C$10)', PCT),
    ("Wtd. XIRR  (by current value)", f'=IF($C$12=0,"",SUMPRODUCT({I_CUR},{I_XIRR})/$C$12)', PCT),
    ("Top Client by Current Value", f'=IF($C$12=0,"",IFERROR(INDEX($B$19:$B${19+CLI_LAST-CLI_FIRST},MATCH(MAX($H$19:$H${19+CLI_LAST-CLI_FIRST}),$H$19:$H${19+CLI_LAST-CLI_FIRST},0)),""))', TEXT),
]
for i, (label, formula, fmt) in enumerate(SNAPSHOT):
    row = 5 + i
    lc = tr.cell(row=row, column=2, value=label)
    lc.font = Font(name=FONT, size=10, bold=True)
    lc.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    lc.border = BOX
    lc.alignment = Alignment(vertical="center", indent=1)
    tr.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    vc = tr.cell(row=row, column=3, value=formula)
    vc.font = Font(name=FONT, size=10, bold=True)
    vc.fill = PatternFill("solid", fgColor=GREY)
    vc.number_format = fmt
    vc.border = BOX
    vc.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    tr.cell(row=row, column=4).border = BOX

# --- client-wise ------------------------------------------------------------
section(17, "CLIENT-WISE REVIEW")
CL_F, CL_L, CL_T = group_table(18, "Clients!$B", I_CLI,
                               CLI_FIRST, CLI_LAST - CLI_FIRST + 1,
                               "Client Name", "GRAND TOTAL")

AC_HEAD = CL_T + 2
section(AC_HEAD, "ASSET CLASS-WISE REVIEW")
AC_F, AC_L, AC_T = group_table(AC_HEAD + 1, "Lists!$B", I_ASSET,
                               2, len(ASSET_CLASSES), "Asset Class", "TOTAL")

CAT_HEAD = AC_T + 2
section(CAT_HEAD, "CATEGORY-WISE REVIEW")
CAT_F, CAT_L, CAT_T = group_table(CAT_HEAD + 1, "Lists!$A", I_CAT,
                                  2, len(CATEGORIES), "Category", "TOTAL")

note(tr, CAT_T + 2, 2,
     "Gain / Loss = (Cur. Value + Withdrawn) - Total Invested   |   Abs. Rtn. = Gain / Loss ÷ Total Invested   |   "
     "Wtd. XIRR is a current-value-weighted average of the scheme XIRRs (indicative).")
note(tr, CAT_T + 3, 2,
     "Client rows follow the Clients sheet order; asset class and category rows follow the Lists sheet order.")

# --- charts -----------------------------------------------------------------
pie = PieChart()
pie.title = "Current Value by Asset Class"
pie.height, pie.width = 8.5, 13
pie.add_data(Reference(tr, min_col=8, min_row=AC_HEAD + 1, max_row=AC_L), titles_from_data=True)
pie.set_categories(Reference(tr, min_col=2, min_row=AC_F, max_row=AC_L))
tr.add_chart(pie, "N5")

bar = BarChart()
bar.type = "bar"
bar.title = "Total Invested vs Current Value  (first 20 clients)"
bar.height, bar.width = 13, 13
bar.add_data(Reference(tr, min_col=6, max_col=6, min_row=18, max_row=min(CL_F + 19, CL_L)), titles_from_data=True)
bar.add_data(Reference(tr, min_col=8, max_col=8, min_row=18, max_row=min(CL_F + 19, CL_L)), titles_from_data=True)
bar.set_categories(Reference(tr, min_col=2, min_row=CL_F, max_row=min(CL_F + 19, CL_L)))
bar.y_axis.title = "Amount"
tr.add_chart(bar, "N25")

tr.freeze_panes = "A4"

# ============================================================ Client Report ==
cr = wb.create_sheet("Client Report")
cr.sheet_view.showGridLines = False
banner(cr, "CLIENT PROFIT & LOSS REVIEW",
       "Pick a client below - the header, summary and scheme list fill in from the Clients and Investments sheets.", 11)
widths(cr, {"A": 6, "B": 22, "C": 42, "D": 16, "E": 14, "F": 17, "G": 17,
            "H": 17, "I": 17, "J": 11, "K": 11})

sel_lbl = cr.cell(row=4, column=2, value="Select Client")
sel_lbl.font = Font(name=FONT, size=11, bold=True, color=WHITE)
sel_lbl.fill = PatternFill("solid", fgColor=NAVY)
sel_lbl.alignment = Alignment(vertical="center", indent=1)
sel = cr.cell(row=4, column=3, value="Sample Client (EXAMPLE)")
sel.font = Font(name=FONT, size=11, bold=True, color="0000FF")
sel.fill = PatternFill("solid", fgColor=YELLOW)
sel.border = BOX
sel.alignment = Alignment(vertical="center", indent=1)
cr.row_dimensions[4].height = 22
dv_sel = DataValidation(type="list", formula1=C_NAME, allow_blank=True)
cr.add_data_validation(dv_sel)
dv_sel.add("C4")

CLI_COLS = {"Client Code": 1, "PAN": 3, "Rep By / Family Head": 4, "Segment": 5,
            "Mobile No.": 6, "Email": 7, "DOB": 8, "Address": 9, "Advisor / RM": 10}
DETAILS = [("Client Code", CNT), ("PAN", TEXT), ("Rep By / Family Head", TEXT),
           ("Mobile No.", TEXT), ("Email", TEXT), ("DOB", DATE),
           ("Address", TEXT), ("Advisor / RM", TEXT)]
for i, (label, fmt) in enumerate(DETAILS):
    row = 6 + i
    lc = cr.cell(row=row, column=2, value=label)
    lc.font = Font(name=FONT, size=10, bold=True)
    lc.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    lc.border = BOX
    lc.alignment = Alignment(vertical="center", indent=1)
    col = CLI_COLS[label]
    lookup = (f'INDEX(Clients!${chr(64 + col)}${CLI_FIRST}:${chr(64 + col)}'
              f'${CLI_LAST},MATCH($C$4,{C_NAME},0))')
    vc = cr.cell(row=row, column=3, value=(
        f'=IF($C$4="","",IFERROR(IF({lookup}="","",{lookup}),""))'))
    vc.font = CALC_FONT
    vc.fill = PatternFill("solid", fgColor=GREY)
    vc.number_format = TEXT if fmt is CNT else fmt
    vc.border = BOX
    vc.alignment = Alignment(vertical="center", indent=1)

# firm block (typed once, prints alongside the client block)
FIRM = [("Firm Name", "OPULENCIA CAPITAL PRIVATE LIMITED"),
        ("Registration", "AMFI-Registered Mutual Fund Distributor"),
        ("Office Address", "Flat No. 501, 5th Floor, Piyusha Society, Law College Road, Erandwane, Pune - 411004"),
        ("Tel. No.", ""), ("Email", ""), ("Website", "")]
for i, (label, val) in enumerate(FIRM):
    row = 6 + i
    lc = cr.cell(row=row, column=5, value=label)
    lc.font = Font(name=FONT, size=10, bold=True)
    lc.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    lc.border = BOX
    lc.alignment = Alignment(vertical="center", indent=1)
    cr.merge_cells(start_row=row, start_column=6, end_row=row, end_column=11)
    vc = cr.cell(row=row, column=6, value=val or None)
    vc.font = INPUT_FONT
    vc.fill = PatternFill("solid", fgColor=YELLOW)
    vc.border = BOX
    vc.alignment = Alignment(vertical="center", indent=1)
    for col in range(7, 12):
        cr.cell(row=row, column=col).border = BOX
note(cr, 12, 5, "Firm block - type your own details once; they stay put when the client changes.")

# summary strip
section_row = 15
cr.merge_cells(start_row=section_row, start_column=1, end_row=section_row, end_column=11)
sc = cr.cell(row=section_row, column=1, value="SUMMARY FOR SELECTED CLIENT")
sc.font = Font(name=FONT, size=11, bold=True, color=WHITE)
sc.fill = PatternFill("solid", fgColor=NAVY)
sc.alignment = Alignment(vertical="center", indent=1)

head(cr, 16, ["Schemes", "Purchase", "Switch In", "Total Invested", "Withdrawn",
              "Cur. Value", "Gain / Loss", "Abs. Rtn.", "Wtd. XIRR"],
     start_col=2, height=30)
SUMMARY = [
    (2, f'=IF($C$4="","",COUNTIFS({I_CLI},$C$4))', CNT),
    (3, f'=IF($C$4="","",SUMIFS({I_PUR},{I_CLI},$C$4))', INR),
    (4, f'=IF($C$4="","",SUMIFS({I_SWIN},{I_CLI},$C$4))', INR),
    (5, f'=IF($C$4="","",SUMIFS({I_INVSD},{I_CLI},$C$4))', INR),
    (6, (f'=IF($C$4="","",SUMIFS({I_RED},{I_CLI},$C$4)+SUMIFS({I_SWOUT},{I_CLI},$C$4)'
         f'+SUMIFS({I_DIVP},{I_CLI},$C$4))'), INR),
    (7, f'=IF($C$4="","",SUMIFS({I_CUR},{I_CLI},$C$4))', INR),
    (8, '=IF($C$4="","",($G$17+$F$17)-$E$17)', INR),
    (9, '=IF(OR($C$4="",$E$17=0),"",$H$17/$E$17)', PCT),
    (10, (f'=IF(OR($C$4="",$G$17=0),"",SUMPRODUCT(({I_CLI}=$C$4)*{I_CUR}*{I_XIRR})/$G$17)'), PCT),
]
for col, formula, fmt in SUMMARY:
    c = cr.cell(row=17, column=col, value=formula)
    c.font = Font(name=FONT, size=10, bold=True)
    c.fill = PatternFill("solid", fgColor=GREY)
    c.number_format = fmt
    c.border = BOX
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)

# scheme detail
cr.merge_cells(start_row=19, start_column=1, end_row=19, end_column=11)
dc = cr.cell(row=19, column=1, value="SCHEME-WISE DETAIL")
dc.font = Font(name=FONT, size=11, bold=True, color=WHITE)
dc.fill = PatternFill("solid", fgColor=NAVY)
dc.alignment = Alignment(vertical="center", indent=1)

head(cr, 21, ["Sr.", "Category", "Scheme Name", "Folio No.", "Inv. Since",
              "Total Invested", "Withdrawn", "Cur. Value", "Gain / Loss",
              "Abs. Rtn.", "XIRR"], start_col=1, height=30)

DETAIL_SRC = [(2, "B", TEXT), (3, "C", TEXT), (4, "D", TEXT), (5, "E", DATE),
              (6, "Q", INR), (7, "R", INR), (8, "L", INR), (9, "S", INR),
              (10, "T", PCT), (11, "O", PCT)]
for i in range(RPT_LAST - RPT_FIRST + 1):
    row = RPT_FIRST + i
    stripe = WHITE if i % 2 == 0 else GREY
    idx = i + 1
    sr = cr.cell(row=row, column=1, value=(
        f'=IF($C$4="","",IF({idx}>COUNTIFS({I_CLI},$C$4),"",{idx}))'))
    style_body(sr, CNT, "calc", stripe)
    sr.alignment = Alignment(horizontal="center", vertical="center")
    for col, src, fmt in DETAIL_SRC:
        lookup = (f'INDEX(Investments!${src}${INV_FIRST}:${src}${INV_LAST},'
                  f'MATCH($A{row},Investments!$V${INV_FIRST}:$V${INV_LAST},0))')
        c = cr.cell(row=row, column=col, value=(
            f'=IF($A{row}="","",IFERROR(IF({lookup}="","",{lookup}),""))'))
        style_body(c, fmt, "calc", stripe)

TOT = RPT_LAST + 1
cr.cell(row=TOT, column=1, value="")
cr.cell(row=TOT, column=2, value="TOTAL")
for col, letter in ((6, "F"), (7, "G"), (8, "H")):
    cr.cell(row=TOT, column=col,
            value=f'=SUM(${letter}${RPT_FIRST}:${letter}${RPT_LAST})')
cr.cell(row=TOT, column=9, value=f'=($H${TOT}+$G${TOT})-$F${TOT}')
cr.cell(row=TOT, column=10, value=f'=IF($F${TOT}=0,"",$I${TOT}/$F${TOT})')
cr.cell(row=TOT, column=11, value='=$J$17')
for col, fmt in ((1, TEXT), (2, TEXT), (3, TEXT), (4, TEXT), (5, TEXT), (6, INR),
                 (7, INR), (8, INR), (9, INR), (10, PCT), (11, PCT)):
    c = cr.cell(row=TOT, column=col)
    c.font = Font(name=FONT, size=10, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.number_format = fmt
    c.border = BOX

for col in ("I", "J", "K"):
    rng = f"{col}{RPT_FIRST}:{col}{RPT_LAST}"
    cr.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["0"], font=RED_FONT))
    cr.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["0"], font=GREEN_FONT))

note(cr, TOT + 2, 1,
     f"Lists up to {RPT_LAST - RPT_FIRST + 1} schemes for the selected client, in Investments-sheet order. "
     "If a client holds more, the count in the summary strip is still complete.")
cr.freeze_panes = "A22"

# =========================================================== Import Report ==
# Paste a whole Profit & Loss export in on the left; read register-shaped rows
# out on the right. The software prints category headings, per-scheme
# transaction sub-tables and total lines, none of which line up with the
# register - this sheet drops them and reshapes what is left.
im = wb.create_sheet("Import Report")
im.sheet_view.showGridLines = False
banner(im, "IMPORT A PROFIT & LOSS REPORT",
       "Paste the report on the left - ready-to-paste register rows appear on the right.", 16)

SRC_FIRST, SRC_LAST = 9, 2008        # paste area rows
OUT_FIRST, OUT_LAST = 9, 408         # extracted rows (400 schemes)
OUT_COL_FIRST, OUT_COL_LAST = 24, 38  # X to AL - same 15 columns as Investments A:O
POS_COL = 23                          # W: row position of each extracted scheme
SRC = f"$A${SRC_FIRST}:$A${SRC_LAST}"
OUT_A = get_column_letter(OUT_COL_FIRST)
OUT_Z = get_column_letter(OUT_COL_LAST)

STEPS = [
    ("STEP 1", "Open the client's Profit & Loss report, press Ctrl+A then Ctrl+C to copy the whole sheet."),
    ("STEP 2", f"Click cell A{SRC_FIRST} below and press Ctrl+V. Category headings, transaction "
               "rows and total lines are ignored automatically."),
    ("STEP 3", "Check the Client Name box on the right. It is read from the report header - "
               "type over it, or pick from the list, so it matches the Clients sheet exactly."),
    ("STEP 4", f"Select {OUT_A}{OUT_FIRST}:{OUT_Z}... down to the last filled row, press Ctrl+C, then on "
               "the Investments sheet right-click cell A6 and choose Paste Special > Values."),
]
for i, (tag, text) in enumerate(STEPS):
    r = 4 + i
    tc = im.cell(row=r, column=1, value=tag)
    tc.font = Font(name=FONT, size=9, bold=True, color=WHITE)
    tc.fill = PatternFill("solid", fgColor=BLUE)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    im.merge_cells(start_row=r, start_column=2, end_row=r, end_column=16)
    bc = im.cell(row=r, column=2, value=text)
    bc.font = BODY_FONT
    bc.alignment = Alignment(vertical="center", indent=1)

# Client name: detected from the report header in W5, cleaned in R6, and
# overridable in R5 - the extracted rows use the override when it is filled.
lbl = im.cell(row=5, column=17, value="Client Name")
lbl.font = Font(name=FONT, size=10, bold=True)
lbl.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
lbl.border = BOX
lbl.alignment = Alignment(vertical="center", indent=1)
sel = im.cell(row=5, column=18)
sel.font = INPUT_FONT
sel.fill = PatternFill("solid", fgColor=YELLOW)
sel.border = BOX
sel.alignment = Alignment(vertical="center", indent=1)
dv_imp = DataValidation(type="list", formula1=C_NAME, allow_blank=True)
im.add_data_validation(dv_imp)
dv_imp.add("R5")

hint = im.cell(row=6, column=17, value="read from report:")
hint.font = Font(name=FONT, size=9, italic=True, color="808080")
hint.alignment = Alignment(horizontal="right", vertical="center")
im.cell(row=5, column=POS_COL, value=(
    f'=IFERROR(INDEX({SRC},MATCH(1,$V${SRC_FIRST}:$V${SRC_LAST},0)),"")'))
det = im.cell(row=6, column=18, value=(
    '=IF($W$5="","",TRIM(IF(ISNUMBER(SEARCH(" REP BY ",$W$5)),'
    'LEFT($W$5,SEARCH(" REP BY ",$W$5)-1),SUBSTITUTE($W$5,"()",""))))'))
det.font = Font(name=FONT, size=9, italic=True, color="595959")
det.alignment = Alignment(vertical="center", indent=1)

# --- paste area: the report's own column layout, left unstyled so a paste
#     lands cleanly and the file stays small
head(im, 8, ["Scheme Name  (+ Folio)", "", "", "Inv. Since", "Purchase", "Switch In",
             "Div Reinv", "Redemption", "Switch Out", "Div Pay", "Cur. Value",
             "Cur. Units", "Cur. NAV", "Gain/Loss", "Abs. Rtn.", "XIRR"], height=30)
widths(im, {"A": 46, "B": 9, "C": 9, "D": 11, "E": 13, "F": 13, "G": 11, "H": 13,
            "I": 13, "J": 11, "K": 13, "L": 12, "M": 11, "N": 13, "O": 10, "P": 9,
            "Q": 17, "R": 30})
im.freeze_panes = "A9"

# --- helper columns S-W: category carry-down, scheme flag, running index,
#     client-name flag, and the matched source row for each output line
for r in range(SRC_FIRST, SRC_LAST + 1):
    is_category = (f'AND(ISTEXT($A{r}),ISNUMBER(SEARCH(" - ",$A{r})),'
                   f'ISERROR(SEARCH("Folio",$A{r})),$D{r}="")')
    carry = '""' if r == SRC_FIRST else f"$S{r - 1}"
    im.cell(row=r, column=19, value=f'=IF({is_category},TRIM($A{r}),{carry})')
    im.cell(row=r, column=20,
            value=f'=IF(AND(ISNUMBER(SEARCH("Folio",$A{r})),$D{r}<>""),1,"")')
    im.cell(row=r, column=21,
            value=f'=IF($T{r}=1,COUNTIF($T${SRC_FIRST}:$T{r},1),"")')
    im.cell(row=r, column=22,
            value=f'=IF(AND(ISTEXT($A{r}),LEFT($A{r + 1},3)="Pan"),1,"")')

# --- extracted rows, in Investments A:O order -------------------------------
im.merge_cells(start_row=7, start_column=OUT_COL_FIRST, end_row=7, end_column=OUT_COL_LAST)
rdy = im.cell(row=7, column=OUT_COL_FIRST,
              value="READY TO PASTE  →  copy these rows, then Investments!A6 → "
                    "Paste Special → Values")
rdy.font = Font(name=FONT, size=10, bold=True, color=WHITE)
rdy.fill = PatternFill("solid", fgColor=NAVY)
rdy.alignment = Alignment(vertical="center", indent=1)

head(im, 8, ["Client Name", "Category", "Scheme Name", "Folio No.", "Inv. Since",
             "Purchase", "Switch In", "Div Reinv", "Redemption", "Switch Out",
             "Div Pay", "Cur. Value", "Cur. Units", "Cur. NAV", "XIRR"],
     start_col=OUT_COL_FIRST, fill=NAVY, height=30)
for col, width in zip(range(OUT_COL_FIRST, OUT_COL_LAST + 1),
                      (30, 32, 46, 16, 12, 14, 14, 12, 14, 14, 11, 15, 13, 12, 9)):
    im.column_dimensions[get_column_letter(col)].width = width

# output column -> source column on the pasted report
AMOUNT_SRC = {29: "E", 30: "F", 31: "G", 32: "H", 33: "I", 34: "J", 35: "K",
              36: "L", 37: "M"}
OUT_FMT = {24: TEXT, 25: TEXT, 26: TEXT, 27: TEXT, 28: DATE, 36: UNITS, 37: NAV,
           38: PCT}


def src_index(letter, pos):
    return f'INDEX(${letter}${SRC_FIRST}:${letter}${SRC_LAST},{pos})'


for i in range(OUT_LAST - OUT_FIRST + 1):
    r = OUT_FIRST + i
    pos = f"$W{r}"
    idx = im.cell(row=r, column=POS_COL,
                  value=f'=IFERROR(MATCH({i + 1},$U${SRC_FIRST}:$U${SRC_LAST},0),"")')
    idx.font = Font(name=FONT, size=9, color="A6A6A6")

    clean = f'SUBSTITUTE({src_index("A", pos)},CHAR(10)," ")'
    since = src_index("D", pos)
    values = {
        24: f'=IF({pos}="","",IF($R$5<>"",$R$5,$R$6))',
        25: f'=IF({pos}="","",{src_index("S", pos)})',
        26: (f'=IF({pos}="","",IFERROR(TRIM(LEFT({clean},SEARCH("Folio",{clean})-1)),'
             f'TRIM({clean})))'),
        27: (f'=IF({pos}="","",IFERROR(TRIM(SUBSTITUTE(MID({clean},'
             f'SEARCH("Folio",{clean})+5,200),":","")),""))'),
        28: (f'=IF({pos}="","",IF(ISNUMBER({since}),{since},'
             f'IFERROR(DATE(2000+VALUE(RIGHT({since},2)),VALUE(MID({since},4,2)),'
             f'VALUE(LEFT({since},2))),"")))'),
    }
    for col, letter in AMOUNT_SRC.items():
        ref = src_index(letter, pos)
        values[col] = (f'=IF({pos}="","",IFERROR(IF(ISNUMBER({ref}),{ref},'
                       f'VALUE(SUBSTITUTE({ref},",",""))),""))')
    xirr = src_index("P", pos)
    values[38] = (f'=IF({pos}="","",IFERROR(IF(ISNUMBER({xirr}),{xirr},'
                  f'VALUE(SUBSTITUTE({xirr},"%",""))/100),""))')

    stripe = WHITE if i % 2 == 0 else GREY
    for col in range(OUT_COL_FIRST, OUT_COL_LAST + 1):
        cell = im.cell(row=r, column=col, value=values[col])
        style_body(cell, OUT_FMT.get(col, DECIMAL), "calc", stripe)

for col in range(19, POS_COL + 1):
    im.column_dimensions[get_column_letter(col)].hidden = True

note(im, OUT_LAST + 2, OUT_COL_FIRST,
     f"Takes up to {OUT_LAST - OUT_FIRST + 1} schemes from a report of up to "
     f"{SRC_LAST - SRC_FIRST + 1} rows. Clear the paste area before importing the next client.")

# helper column is machinery, not data - keep it out of the way
iv.column_dimensions["V"].hidden = True

wb._sheets = [wb[name] for name in
              ["Instructions", "Clients", "Import Report", "Investments",
               "Total Review", "Client Report", "Lists"]]
wb.active = wb.sheetnames.index("Total Review")
out = sys.argv[1] if len(sys.argv) > 1 else "Client_Investment_Tracker.xlsx"
wb.save(out)
print("written:", out)

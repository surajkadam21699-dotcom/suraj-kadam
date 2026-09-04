"""Build the one-sheet Mutual Fund Review template.

Usage: python tools/build_simple_fund_template.py [output.xlsx]

Everything on a single tab: the rating rule, a live summary, and the
scheme register. A fund is graded on its 3-year return -

    3Y >= Best threshold      -> Best Fund
    3Y >= Improve threshold   -> Need to Improve
    3Y <  Improve threshold   -> Not Performing
    3Y blank                  -> Data Pending

Both thresholds are editable cells, so the bands re-cut without touching a
formula. The header row carries an AutoFilter: click the Verdict arrow to
see just the best funds, just the ones needing work, or just the laggards.

Number formats and styling follow build_client_investment_template.py.
"""

import datetime
import sys

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

FONT = "Arial"
NAVY, BLUE, GREY, WHITE, YELLOW, LINE = "1F3864", "2E75B6", "F2F2F2", "FFFFFF", "FFF2CC", "BFBFBF"
GREEN_FILL, AMBER_FILL, RED_FILL = "C6EFCE", "FFEB9C", "FFC7CE"

INPUT_FONT = Font(name=FONT, size=10, color="0000FF")
CALC_FONT = Font(name=FONT, size=10, color="000000")
BODY_FONT = Font(name=FONT, size=10)
BOLD_FONT = Font(name=FONT, size=10, bold=True)
GREEN_FONT = Font(name=FONT, size=10, color="006100")
AMBER_FONT = Font(name=FONT, size=10, color="9C5700")
RED_FONT = Font(name=FONT, size=10, color="9C0006")
HEAD_FONT = Font(name=FONT, size=10, bold=True, color=WHITE)

INR = r'[>=10000000]₹ ##\,##\,##\,##0.00;[>=100000]₹ ##\,##\,##0.00;₹ #,##0.00'
PCT = '0.00%;(0.00%);-'
CNT = '#,##0;(#,##0);-'

THIN = Side(style="thin", color=LINE)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

BEST, IMPROVE, POOR, PENDING = (
    "Best Fund", "Need to Improve", "Not Performing", "Data Pending")

HDR = 10                      # header row
FIRST, LAST = 11, 310         # 300 scheme rows
BEST_CELL, IMP_CELL = "$J$4", "$J$5"

# header, width, calculated?
COLS = [
    ("Client Name",            24, False),
    ("Scheme Name",            40, False),
    ("Category",               28, False),
    ("Total Invested",         16, False),
    ("Current Value (Live)",   18, False),
    ("Gain / Loss",            15, True),
    ("Abs. Rtn.",              11, True),
    ("3Y Return (CAGR)",       14, False),
    ("Category Avg 3Y",        14, False),
    ("vs Category",            12, True),
    ("Verdict",                17, True),
    ("Suggested Action",       30, True),
    ("Remark",                 44, False),
]

CATEGORIES = [
    "Equity - Large Cap Fund", "Equity - Large & Mid Cap Fund", "Equity - Mid Cap Fund",
    "Equity - Small Cap Fund", "Equity - Multi Cap Fund", "Equity - Flexi Cap Fund",
    "Equity - Focused Fund", "Equity - Value Fund", "Equity - Contra Fund",
    "Equity - Dividend Yield Fund", "Equity - ELSS", "Equity - Sectoral/ Thematic",
    "Debt - Overnight Fund", "Debt - Liquid Fund", "Debt - Ultra Short Duration Fund",
    "Debt - Low Duration Fund", "Debt - Money Market Fund", "Debt - Short Duration Fund",
    "Debt - Corporate Bond Fund", "Debt - Banking and PSU Fund", "Debt - Credit Risk Fund",
    "Debt - Dynamic Bond", "Debt - Gilt Fund", "Hybrid - Aggressive Hybrid Fund",
    "Hybrid - Dynamic Asset Allocation or Balanced Advantage",
    "Hybrid - Multi Asset Allocation", "Hybrid - Conservative Hybrid Fund",
    "Hybrid - Equity Savings", "Hybrid - Arbitrage Fund", "Others - Index Funds",
    "Others - FoF Domestic", "Others - FoF Overseas", "Others - Gold ETF",
    "Others - Other ETFs", "Others - Solution Oriented",
]


def put(ws, ref, value, *, font=BODY_FONT, fmt=None, fill=None,
        align=None, wrap=False, border=False):
    c = ws[ref]
    c.value = value
    c.font = font
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    if align or wrap:
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if border:
        c.border = BOX
    return c


def main(out="Fund_Review_Simple.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Fund Review"
    ws.sheet_view.showGridLines = False

    for i, (_, width, _) in enumerate(COLS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # ------------------------------------------------------------- banner --
    ws.merge_cells("A1:M1")
    put(ws, "A1", "MUTUAL FUND REVIEW", font=Font(name=FONT, size=14, bold=True, color=WHITE),
        fill=NAVY, align="left")
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A2:M2")
    put(ws, "A2", "Type in the blue columns. Verdict, action and the summary fill "
                  "themselves in. Use the Verdict filter arrow to see one group at a time.",
        font=Font(name=FONT, size=9, italic=True, color="404040"), fill=GREY, align="left")
    ws.row_dimensions[2].height = 17

    # ------------------------------------------------- summary (A3:E8) -----
    put(ws, "A3", "SUMMARY", font=HEAD_FONT, fill=BLUE, border=True)
    put(ws, "B3", "", fill=BLUE, border=True)
    ws.merge_cells("A3:B3")
    for ref, text in (("C3", "Schemes"), ("D3", "Total Invested"), ("E3", "Current Value")):
        put(ws, ref, text, font=HEAD_FONT, fill=BLUE, align="center", wrap=True, border=True)
    ws.row_dimensions[3].height = 20

    K_RNG = f"$K${FIRST}:$K${LAST}"
    D_RNG = f"$D${FIRST}:$D${LAST}"
    E_RNG = f"$E${FIRST}:$E${LAST}"

    for j, (verdict, fill, font) in enumerate((
            (BEST, GREEN_FILL, GREEN_FONT), (IMPROVE, AMBER_FILL, AMBER_FONT),
            (POOR, RED_FILL, RED_FONT), (PENDING, GREY, BODY_FONT))):
        r = 4 + j
        put(ws, f"A{r}", verdict, font=font, fill=fill, border=True)
        put(ws, f"B{r}", "", fill=fill, border=True)
        ws.merge_cells(f"A{r}:B{r}")
        put(ws, f"C{r}", f"=COUNTIF({K_RNG},$A{r})", font=CALC_FONT, fmt=CNT,
            align="center", border=True)
        put(ws, f"D{r}", f"=SUMIF({K_RNG},$A{r},{D_RNG})", font=CALC_FONT, fmt=INR, border=True)
        put(ws, f"E{r}", f"=SUMIF({K_RNG},$A{r},{E_RNG})", font=CALC_FONT, fmt=INR, border=True)

    put(ws, "A8", "TOTAL", font=HEAD_FONT, fill=NAVY, border=True)
    put(ws, "B8", "", fill=NAVY, border=True)
    ws.merge_cells("A8:B8")
    for col, fmt in (("C", CNT), ("D", INR), ("E", INR)):
        put(ws, f"{col}8", f"=SUM({col}4:{col}7)", font=HEAD_FONT, fmt=fmt, fill=NAVY,
            border=True, align="center" if col == "C" else None)

    # --------------------------------------------- rating rule (G3:J6) -----
    put(ws, "G3", "RATING RULE  -  edit the yellow cells", font=HEAD_FONT, fill=BLUE, border=True)
    for ref in ("H3", "I3"):
        put(ws, ref, "", fill=BLUE, border=True)
    ws.merge_cells("G3:I3")
    put(ws, "J3", "3Y Return", font=HEAD_FONT, fill=BLUE, align="center", wrap=True, border=True)

    rules = [
        (BEST, "Best Fund  -  3Y return at or above", 0.13, GREEN_FILL, GREEN_FONT),
        (IMPROVE, "Need to Improve  -  3Y return at or above", 0.10, AMBER_FILL, AMBER_FONT),
        (POOR, "Not Performing  -  3Y return below", None, RED_FILL, RED_FONT),
    ]
    for j, (_, label, value, fill, font) in enumerate(rules):
        r = 4 + j
        put(ws, f"G{r}", label, font=font, fill=fill, border=True)
        for ref in (f"H{r}", f"I{r}"):
            put(ws, ref, "", fill=fill, border=True)
        ws.merge_cells(f"G{r}:I{r}")
        if value is None:
            put(ws, f"J{r}", f"={IMP_CELL}", font=CALC_FONT, fmt=PCT, fill=GREY,
                align="center", border=True)
        else:
            put(ws, f"J{r}", value, font=INPUT_FONT, fmt=PCT, fill=YELLOW,
                align="center", border=True)

    ws.merge_cells("G7:M8")
    put(ws, "G7", "Enter 1Y / 3Y / 5Y returns as the scheme's published CAGR from the "
                  "factsheet or AMFI - not the XIRR of this holding, which depends on the "
                  "client's own cash-flow dates. The 13% and 10% bars are house rules, "
                  "not a SEBI or AMFI standard.",
        font=Font(name=FONT, size=9, italic=True, color="404040"), wrap=True)

    # -------------------------------------------------------- header row --
    for i, (text, _, _) in enumerate(COLS, start=1):
        put(ws, f"{get_column_letter(i)}{HDR}", text, font=HEAD_FONT, fill=BLUE,
            align="center", wrap=True, border=True)
    ws.row_dimensions[HDR].height = 32
    ws.freeze_panes = f"C{FIRST}"

    money, pcts = {"D", "E", "F"}, {"G", "H", "I", "J"}
    for r in range(FIRST, LAST + 1):
        f = {
            "F": f'=IF($B{r}="","",$E{r}-$D{r})',
            "G": f'=IF(OR($B{r}="",N($D{r})=0),"",$F{r}/$D{r})',
            "J": f'=IF(OR($H{r}="",$I{r}=""),"",$H{r}-$I{r})',
            "K": (f'=IF($B{r}="","",IF($H{r}="","{PENDING}",'
                  f'IF($H{r}>={BEST_CELL},"{BEST}",'
                  f'IF($H{r}>={IMP_CELL},"{IMPROVE}","{POOR}"))))'),
            "L": (f'=IF($K{r}="","",IF($K{r}="{PENDING}","Enter the 3Y return",'
                  f'IF($K{r}="{BEST}","Hold / consider top-up",'
                  f'IF($K{r}="{IMPROVE}","Watch 2 quarters, then decide",'
                  f'"Review with client - switch or exit"))))'),
        }
        for i, (_, _, is_calc) in enumerate(COLS, start=1):
            col = get_column_letter(i)
            fmt = INR if col in money else PCT if col in pcts else None
            if is_calc:
                put(ws, f"{col}{r}", f[col], font=CALC_FONT, fmt=fmt, fill=GREY,
                    border=True, align="center" if col == "K" else None)
            else:
                put(ws, f"{col}{r}", None, font=INPUT_FONT, fmt=fmt, border=True)
        ws.row_dimensions[r].height = 15

    # one worked example, marked for deletion
    for col, value in (
            ("A", "Sample Client (EXAMPLE)"), ("B", "Bluechip Flexi Cap Fund - Growth"),
            ("C", "Equity - Flexi Cap Fund"), ("D", 190000), ("E", 265430),
            ("H", 0.1512), ("I", 0.1410),
            ("M", "EXAMPLE ROW - delete it before you start.")):
        ws[f"{col}{FIRST}"].value = value

    # colour the whole row by verdict
    band = f"A{FIRST}:M{LAST}"
    for verdict, fill, font in ((BEST, GREEN_FILL, GREEN_FONT),
                                (IMPROVE, AMBER_FILL, AMBER_FONT),
                                (POOR, RED_FILL, RED_FONT)):
        ws.conditional_formatting.add(band, FormulaRule(
            formula=[f'$K{FIRST}="{verdict}"'],
            fill=PatternFill("solid", fgColor=fill), font=font))

    # category dropdown, sourced from a hidden column on this same sheet
    for i, cat in enumerate(CATEGORIES, start=1):
        put(ws, f"P{i}", cat)
    ws.column_dimensions["P"].hidden = True
    dv = DataValidation(type="list", formula1=f"=$P$1:$P${len(CATEGORIES)}",
                        allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(f"C{FIRST}:C{LAST}")

    ws.auto_filter.ref = f"A{HDR}:M{LAST}"
    wb.save(out)
    print(f"wrote {out}")


if __name__ == "__main__":
    main(*(sys.argv[1:2] or []))

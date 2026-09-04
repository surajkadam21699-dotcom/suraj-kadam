"""Fund-versus-benchmark performance report, in the portfolio-review format.

Usage: python tools/build_fund_performance_report.py [outdir]

Writes two workbooks:
    <Client>_Fund_Performance_Analysis.xlsx   Summary | Fund Performance |
                                              By Benchmark | Action Plan
    <Client>_Fund_Performance_1Pager.xlsx     one sheet, prints on one page

Source: the Outperforming / Underperforming tables supplied by the client on
4 September 2026. All 11 stated gaps reconcile against fund minus benchmark.

Reading note carried through both files: the same index appears with different
returns across rows - Nifty 500 TRI at 7.56%, 6.39%, 5.79% and 5.83%, Nifty
Financial Services TRI at -10.36% and -1.20%. One index over one period has one
return, so each row is measured over its own window (each fund against its own
benchmark over that fund's holding period). Every row is therefore sound on its
own, but the gaps are not like-for-like across funds and the sheet does not
present them as a league table.
"""

import os
import re
import sys
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

CLIENT = "Yogesh Sharad Gavankar"
ASON = "4 September 2026"
BEST = 0.13                      # house rule: a fund at or above 13% is "best"

FONT = "Arial"
NAVY, BLUE, GREY, WHITE, YELLOW, LINE = "1F3864", "2E75B6", "F2F2F2", "FFFFFF", "FFF2CC", "BFBFBF"
RED_FILL, AMBER_FILL, GREEN_FILL, QUIET = "FFC7CE", "FFEB9C", "C6EFCE", "E7E9E8"

BODY = Font(name=FONT, size=10)
BOLD = Font(name=FONT, size=10, bold=True)
CALC = Font(name=FONT, size=10, color="000000")
SRC = Font(name=FONT, size=10, color="0000FF")
HEAD = Font(name=FONT, size=10, bold=True, color=WHITE)
RED = Font(name=FONT, size=10, color="9C0006")
AMBER = Font(name=FONT, size=10, color="9C5700")
GREEN = Font(name=FONT, size=10, color="006100")
ITAL = Font(name=FONT, size=9, italic=True, color="404040")
SECT = Font(name=FONT, size=11, bold=True, color=NAVY)

PCT = '0.00%;(0.00%);-'
PTS = '+0.00"  pts";-0.00"  pts";-'
CNT = '#,##0;(#,##0);-'
THIN = Side(style="thin", color=LINE)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# fund, benchmark, fund return, benchmark return  (gap is computed, never pasted)
FUNDS = [
    ("Mahindra Manulife Small Cap Fund - Regular Plan - Growth",       "Nifty Smallcap 250 TRI",          0.5351,  0.1801),
    ("Mahindra Manulife Multi Cap Fund - Regular Plan - Growth",       "Nifty500 Multicap 50:25:25 TRI",  0.3240,  0.0203),
    ("HDFC Innovation Fund - Regular Plan - Growth",                   "Nifty 500 TRI",                   0.2884,  0.0756),
    ("Mahindra Manulife Banking and Financial Services Fund - Regular Plan - Growth", "Nifty Financial Services TRI", 0.0958, -0.1036),
    ("Mahindra Manulife Banking and Financial Services Fund - Regular Plan - Growth", "Nifty Financial Services TRI", 0.0806, -0.0120),
    ("Mahindra Manulife Value Fund - Regular Plan - Growth",           "Nifty 500 TRI",                   0.1524,  0.0639),
    ("HDFC Mid Cap Fund - Regular Plan - Growth",                      "Nifty Midcap 150 TRI",            0.1552,  0.0862),
    ("Mahindra Manulife Large and Mid Cap Fund - Regular Plan - Growth", "NIFTY LargeMidcap 250 TRI",     0.1126,  0.0727),
    ("Mahindra Manulife Focused Fund - Regular Plan - Growth",         "Nifty 500 TRI",                   0.0114,  0.0579),
    ("Mahindra Manulife Flexi Cap Fund - Regular Plan - Growth",       "Nifty 500 TRI",                   0.0431,  0.0583),
    ("Mahindra Manulife Consumption Fund - Regular Plan - Growth",     "(not stated)",                    0.0287,  0.0396),
]
NO_DATA = [("Mahindra Manulife Low Duration Fund - Regular Plan - Growth",
            "No return, benchmark or gap was supplied for this fund.")]

GAPS = [
    ("The periods are not the same",
     "Nifty 500 TRI appears at 7.56%, 6.39%, 5.79% and 5.83%, and Nifty Financial "
     "Services TRI at -10.36% and -1.20%. One index over one period has one return, so "
     "each row covers its own window - most likely each fund's holding period. Compare "
     "each fund with its own benchmark; do not rank the gaps against each other."),
    ("Banking and Financial Services appears twice",
     "Two rows, same fund and same benchmark, different figures (9.58% vs -10.36%, and "
     "8.06% vs -1.20%). Two folios, or two periods - confirm which before this reaches "
     "the client."),
    ("One fund has no data",
     "Mahindra Manulife Low Duration Fund is listed with no return, benchmark or gap."),
    ("One benchmark is missing",
     "Mahindra Manulife Consumption Fund shows 2.87% against 3.96%, but the index it is "
     "measured against was not given."),
    ("The underperformer list may be cut short",
     "The source table ends at Consumption Fund with no total, so confirm nothing "
     "below it was lost."),
]


def put(ws, ref, value, *, font=BODY, fmt=None, fill=None, align=None,
        wrap=False, border=False):
    c = ws[ref]
    c.value = value
    c.font = font
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    if align or wrap:
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if border:
        c.border = BOX
    return c


def banner(ws, title, subtitle, last):
    ws.merge_cells(f"A1:{last}1")
    put(ws, "A1", title, font=Font(name=FONT, size=13, bold=True, color=WHITE),
        fill=NAVY, align="left")
    ws.row_dimensions[1].height = 28
    ws.merge_cells(f"A2:{last}2")
    put(ws, "A2", subtitle, font=ITAL, fill=GREY, align="left")
    ws.row_dimensions[2].height = 17


def heads(ws, row, labels, widths=None, left=()):
    for i, label in enumerate(labels, start=1):
        col = get_column_letter(i)
        put(ws, f"{col}{row}", label, font=HEAD, fill=BLUE,
            align="left" if col in left else "center", wrap=True, border=True)
        if widths:
            ws.column_dimensions[col].width = widths[i - 1]
    ws.row_dimensions[row].height = 30


def ranked():
    return sorted(FUNDS, key=lambda f: -(f[2] - f[3]))


def verdict(fr, bm):
    return "Ahead of benchmark" if fr > bm else "Behind benchmark"


def fund_rows(ws, first, cols="ABCDEFG", data_bar=True):
    """Write the 11 comparisons. Gap and verdict are formulas, never pasted."""
    rows = ranked()
    for i, (name, bench, fr, bm) in enumerate(rows):
        r = first + i
        ahead = fr > bm
        put(ws, f"A{r}", i + 1, fmt=CNT, align="center", border=True)
        put(ws, f"B{r}", name, font=SRC, border=True)
        put(ws, f"C{r}", bench, font=SRC if bench != "(not stated)" else AMBER,
            fill=None if bench != "(not stated)" else AMBER_FILL, border=True)
        put(ws, f"D{r}", fr, font=SRC, fmt=PCT, align="center", border=True)
        put(ws, f"E{r}", bm, font=SRC, fmt=PCT, align="center", border=True)
        put(ws, f"F{r}", f"=($D{r}-$E{r})*100", font=CALC, fmt=PTS,
            fill=GREEN_FILL if ahead else RED_FILL,
            align="center", border=True)
        put(ws, f"G{r}", f'=IF($D{r}>$E{r},"Ahead of benchmark","Behind benchmark")',
            font=GREEN if ahead else RED, border=True)
        put(ws, f"H{r}", f'=IF($D{r}>={BEST},"Yes","No")',
            font=GREEN if fr >= BEST else AMBER,
            fill=GREEN_FILL if fr >= BEST else AMBER_FILL, align="center", border=True)
    last = first + len(rows) - 1
    if data_bar:
        ws.conditional_formatting.add(
            f"D{first}:D{last}",
            DataBarRule(start_type="num", start_value=0, end_type="num",
                        end_value=0.56, color="1F4E5F", showValue=True))
    return last


def totals(ws, first, last, label="AVERAGE  (11 comparisons)"):
    r = last + 1
    for col in "ACGH":
        put(ws, f"{col}{r}", "", fill=NAVY, border=True)
    put(ws, f"B{r}", label, font=HEAD, fill=NAVY, border=True)
    put(ws, f"D{r}", f"=AVERAGE(D{first}:D{last})", font=HEAD, fmt=PCT, fill=NAVY,
        align="center", border=True)
    put(ws, f"E{r}", f"=AVERAGE(E{first}:E{last})", font=HEAD, fmt=PCT, fill=NAVY,
        align="center", border=True)
    put(ws, f"F{r}", f"=($D{r}-$E{r})*100", font=HEAD, fmt=PTS, fill=NAVY,
        align="center", border=True)
    put(ws, f"G{r}", f'=COUNTIF(G{first}:G{last},"Ahead of benchmark")&" ahead / "&'
                     f'COUNTIF(G{first}:G{last},"Behind benchmark")&" behind"',
        font=HEAD, fill=NAVY, border=True)
    put(ws, f"H{r}", f'=COUNTIF(H{first}:H{last},"Yes")&" of "&'
                     f'COUNTA(H{first}:H{last})', font=HEAD, fill=NAVY,
        align="center", border=True)
    return r


HEADERS = ["#", "Fund", "Benchmark", "Fund Return", "Benchmark Return",
           "Gap", "Against benchmark", f"Clears {BEST:.0%}?"]
WIDTHS = [5, 54, 30, 13, 15, 13, 21, 12]


# ============================================================== analysis =====
def write_analysis(path):
    wb = Workbook()
    wb.remove(wb.active)
    first = 5

    # ---------------------------------------------------- Fund Performance --
    ws = wb.create_sheet("Fund Performance")
    ws.sheet_view.showGridLines = False
    banner(ws, "FUND PERFORMANCE AGAINST BENCHMARK",
           f"{CLIENT}.  Each fund against its own benchmark over its own period - "
           f"the gaps are not comparable across funds. Ordered by gap.", "H")
    heads(ws, 4, HEADERS, WIDTHS, left=("B", "C", "G"))
    ws.freeze_panes = "C5"
    last = fund_rows(ws, first)
    tot = totals(ws, first, last)
    r = tot + 2
    put(ws, f"A{r}", "LISTED BUT NOT MEASURED", font=SECT)
    r += 1
    for name, why in NO_DATA:
        put(ws, f"A{r}", "", border=True)
        put(ws, f"B{r}", name, font=SRC, border=True)
        ws.merge_cells(f"C{r}:H{r}")
        put(ws, f"C{r}", why, font=AMBER, fill=AMBER_FILL, wrap=True, border=True)
    ws.auto_filter.ref = f"A4:H{last}"

    # --------------------------------------------------------- By Benchmark --
    ws = wb.create_sheet("By Benchmark")
    ws.sheet_view.showGridLines = False
    banner(ws, "BY BENCHMARK",
           "Where one index shows more than one return, the rows under it cover "
           "different periods. This is the evidence for the reading note.", "E")
    heads(ws, 4, ["Benchmark", "Funds", "Index returns seen", "Same period?", ""],
          [32, 9, 34, 20, 2], left=("A", "C", "D"))
    groups = defaultdict(list)
    for name, bench, fr, bm in FUNDS:
        groups[bench].append((name, fr, bm))
    r = 5
    for bench, items in sorted(groups.items(), key=lambda kv: -len(kv[1])):
        seen = sorted({bm for _, _, bm in items}, reverse=True)
        same = len(seen) == 1
        put(ws, f"A{r}", bench, border=True)
        put(ws, f"B{r}", len(items), fmt=CNT, align="center", border=True)
        put(ws, f"C{r}", ",  ".join(f"{v:+.2%}" for v in seen), align="center", border=True)
        put(ws, f"D{r}", "Yes" if same else f"No - {len(seen)} periods",
            font=GREEN if same else RED, fill=GREEN_FILL if same else RED_FILL,
            align="center", border=True)
        r += 1
    r += 1
    ws.merge_cells(f"A{r}:D{r + 2}")
    put(ws, f"A{r}", "An index cannot post two different returns over one period. Nifty 500 "
                     "TRI appears four times with four different values and Nifty Financial "
                     "Services TRI twice, so the rows beneath them were measured over "
                     "different windows. Each fund-versus-benchmark pair is still a fair "
                     "comparison; the gaps between funds are not.",
        font=AMBER, fill=YELLOW, wrap=True, border=True)

    # ---------------------------------------------------------- Action Plan --
    ws = wb.create_sheet("Action Plan")
    ws.sheet_view.showGridLines = False
    banner(ws, "ACTION PLAN", "Owner and Done are yours to fill in.", "E")
    for col, w in zip("ABCDE", (16, 32, 58, 18, 10)):
        ws.column_dimensions[col].width = w
    heads(ws, 4, ["When", "Step", "What happens", "Owner", "Done"], left=("B", "C"))
    plan = [
        ("Before sending", "Fix the data", "Resolve the duplicated Banking and Financial "
         "Services row, add the missing Consumption benchmark and the Low Duration figures, "
         "and state the period each return covers."),
        ("Week 1", "Review the three laggards", "Focused (-4.65 pts), Flexi Cap (-1.52 pts) "
         "and Consumption (-1.09 pts) trail their benchmarks. Focused is the one to act on: "
         "1.14% absolute."),
        ("Week 1", "Look past the gap", "Banking and Financial Services beat a falling index "
         "on both rows but returned 9.58% and 8.06%. Beating a benchmark that fell is not "
         "the same as making money."),
        ("Week 2", "Confirm the winners", "Small Cap (+35.50 pts), Multi Cap (+30.37 pts) "
         "and HDFC Innovation (+21.28 pts) lead by a wide margin. Check the period before "
         "quoting these to the client."),
        ("Ongoing", "Standardise", "Pull all funds over one common period - 1Y, 3Y and 5Y - "
         "so the next review compares like with like."),
    ]
    r = 5
    for when, step, what in plan:
        put(ws, f"A{r}", when, font=BOLD, align="center", border=True)
        put(ws, f"B{r}", step, font=BOLD, wrap=True, border=True)
        put(ws, f"C{r}", what, wrap=True, border=True)
        put(ws, f"D{r}", None, font=SRC, border=True)
        put(ws, f"E{r}", None, font=SRC, align="center", border=True)
        ws.row_dimensions[r].height = 40
        r += 1
    r += 1
    ws.merge_cells(f"A{r}:E{r + 1}")
    put(ws, f"A{r}", "Past performance is not indicative of future returns. This review "
                     "compares each fund with its stated benchmark only and does not account "
                     "for the amount invested in each. Mutual fund investments are subject to "
                     "market risks; read all scheme-related documents carefully.",
        font=ITAL, wrap=True)

    # -------------------------------------------------------------- Summary --
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    banner(ws, f"FUND PERFORMANCE REVIEW  -  {CLIENT.upper()}",
           f"As on {ASON}.  11 fund-benchmark comparisons.", "E")
    for col, w in zip("ABCDE", (34, 15, 15, 15, 62)):
        ws.column_dimensions[col].width = w

    fp = "'Fund Performance'"
    d = f"{fp}!$D${first}:$D${last}"
    e = f"{fp}!$E${first}:$E${last}"
    g = f"{fp}!$G${first}:$G${last}"

    put(ws, "A4", "HEADLINE", font=SECT)
    heads(ws, 5, ["Measure", "Value", "", "", "Comment"], left=("A", "E"))
    lines = [
        ("Comparisons", f"=COUNTA({fp}!$B${first}:$B${last})", CNT,
         "One fund appears twice, so 11 rows cover 10 distinct funds."),
        ("Ahead of benchmark", f'=COUNTIF({g},"Ahead of benchmark")', CNT,
         "Eight of eleven rows beat the index they are measured against."),
        ("Behind benchmark", f'=COUNTIF({g},"Behind benchmark")', CNT,
         "Focused, Flexi Cap and Consumption."),
        ("Hit rate", f'=COUNTIF({g},"Ahead of benchmark")/COUNTA({g})', PCT,
         "Share of rows ahead of their benchmark."),
        ("Average fund return", f"=AVERAGE({d})", PCT,
         "Simple average across rows - not weighted by amount invested."),
        ("Average benchmark", f"=AVERAGE({e})", PCT,
         "Averaged across different indices and different periods."),
        ("Average gap", f"=(AVERAGE({d})-AVERAGE({e}))*100", PTS,
         "Indicative only, for the reason above."),
        (f"Clear the {BEST:.0%} bar", f'=COUNTIF({d},">="&{BEST})', CNT,
         "Small Cap, Multi Cap, HDFC Innovation, HDFC Mid Cap and Value."),
    ]
    r = 6
    for label, formula, fmt, comment in lines:
        put(ws, f"A{r}", label, font=BOLD, border=True)
        ws.merge_cells(f"B{r}:D{r}")
        put(ws, f"B{r}", formula, font=CALC, fmt=fmt, fill=GREY, align="center", border=True)
        put(ws, f"E{r}", comment, wrap=True, border=True)
        ws.row_dimensions[r].height = 24
        r += 1

    r += 1
    put(ws, f"A{r}", "WHAT STANDS OUT", font=SECT)
    r += 1
    stands = [
        ("Three funds trail their benchmark",
         "Focused -4.65 pts, Flexi Cap -1.52 pts, Consumption -1.09 pts."),
        ("Focused Fund is the weakest holding",
         "1.14% against a benchmark that returned 5.79% - the widest shortfall and the "
         "lowest absolute return in the list."),
        ("A wide gap is not always a good return",
         "Banking and Financial Services beat its index by 19.94 and 9.26 points, but "
         "returned only 9.58% and 8.06% - the index fell 10.36% and 1.20%. Large and Mid "
         "Cap is ahead by 3.99 points on 11.26%. All three sit below the 13% bar."),
        ("Five of eleven clear 13%",
         "Small Cap 53.51%, Multi Cap 32.40%, HDFC Innovation 28.84%, HDFC Mid Cap 15.52% "
         "and Value 15.24%."),
        ("The three leaders lead by a distance",
         "Small Cap +35.50, Multi Cap +30.37 and HDFC Innovation +21.28 points. Confirm "
         "the period before quoting these."),
    ]
    for title, detail in stands:
        put(ws, f"A{r}", title, font=BOLD, wrap=True, border=True)
        ws.merge_cells(f"B{r}:E{r}")
        put(ws, f"B{r}", detail, wrap=True, border=True)
        ws.row_dimensions[r].height = 30
        r += 1

    r += 1
    put(ws, f"A{r}", "BEFORE THIS GOES TO THE CLIENT", font=SECT)
    r += 1
    for title, detail in GAPS:
        put(ws, f"A{r}", title, font=BOLD, fill=YELLOW, wrap=True, border=True)
        ws.merge_cells(f"B{r}:E{r}")
        put(ws, f"B{r}", detail, fill=YELLOW, wrap=True, border=True)
        ws.row_dimensions[r].height = 34
        r += 1

    wb.move_sheet("Summary", offset=-3)
    wb.active = 0
    wb.save(path)
    return path


# =============================================================== one-pager ===
def write_onepager(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Review"
    ws.sheet_view.showGridLines = False
    for i, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    banner(ws, f"FUND PERFORMANCE REVIEW  -  {CLIENT.upper()}",
           f"As on {ASON}   |   11 comparisons, 10 distinct funds   |   "
           f"each fund against its own benchmark", "H")

    ws.merge_cells("A4:H5")
    put(ws, "A4", "Eight of eleven fund-benchmark comparisons are ahead of their index and "
                  "three are behind. Five funds clear the 13% bar. The three laggards are "
                  "Focused, Flexi Cap and Consumption - Focused most of all, at 1.14%.",
        font=Font(name=FONT, size=11), wrap=True)
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 20

    ws.merge_cells("A7:H8")
    put(ws, "A7", "READ THIS FIRST:  the same index appears with different returns across "
                  "rows - Nifty 500 TRI at 7.56%, 6.39%, 5.79% and 5.83%. One index over one "
                  "period has one return, so each row covers its own window. Compare each "
                  "fund with its own benchmark; do not rank the gaps against each other.",
        font=AMBER, fill=YELLOW, wrap=True, border=True)
    ws.row_dimensions[7].height = 20
    ws.row_dimensions[8].height = 20

    put(ws, "A10", "EVERY FUND AGAINST ITS BENCHMARK", font=SECT)
    heads(ws, 11, HEADERS, WIDTHS, left=("B", "C", "G"))
    first = 12
    last = fund_rows(ws, first)
    tot = totals(ws, first, last)

    r = tot + 2
    put(ws, f"A{r}", "WHAT STANDS OUT", font=SECT)
    r += 1
    for title, detail in [
        ("Three funds trail their benchmark",
         "Focused -4.65 pts, Flexi Cap -1.52 pts, Consumption -1.09 pts."),
        ("Focused Fund is the one to act on",
         "1.14% against a benchmark that returned 5.79% - widest shortfall, lowest "
         "absolute return."),
        ("A wide gap is not always a good return",
         "Banking and Financial Services beat its index by 19.94 and 9.26 points but "
         "returned 9.58% and 8.06%; the index itself fell. Below the 13% bar on both rows."),
        ("Five of eleven clear 13%",
         "Small Cap 53.51%, Multi Cap 32.40%, HDFC Innovation 28.84%, HDFC Mid Cap 15.52%, "
         "Value 15.24%."),
    ]:
        put(ws, f"A{r}", "", border=True)
        put(ws, f"B{r}", title, font=BOLD, wrap=True, border=True)
        ws.merge_cells(f"C{r}:H{r}")
        put(ws, f"C{r}", detail, wrap=True, border=True)
        ws.row_dimensions[r].height = 26
        r += 1

    r += 1
    put(ws, f"A{r}", "FIX BEFORE THIS GOES TO THE CLIENT", font=SECT)
    r += 1
    for title, detail in GAPS[1:]:
        put(ws, f"A{r}", "", border=True)
        put(ws, f"B{r}", title, font=BOLD, fill=YELLOW, wrap=True, border=True)
        ws.merge_cells(f"C{r}:H{r}")
        put(ws, f"C{r}", detail, fill=YELLOW, wrap=True, border=True)
        ws.row_dimensions[r].height = 28
        r += 1

    r += 1
    ws.merge_cells(f"A{r}:H{r + 1}")
    put(ws, f"A{r}", "Past performance is not indicative of future returns. This review "
                     "compares each fund with its stated benchmark only and does not account "
                     "for the amount invested in each, so it does not show what the portfolio "
                     "as a whole returned. Mutual fund investments are subject to market "
                     "risks; read all scheme-related documents carefully.",
        font=ITAL, wrap=True)

    ws.print_area = f"A1:H{r + 1}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    wb.save(path)
    return path


def main(outdir="."):
    stem = os.path.join(outdir, re.sub(r"[^A-Za-z0-9]+", "_", CLIENT).strip("_"))
    print("wrote", write_analysis(f"{stem}_Fund_Performance_Analysis.xlsx"))
    print("wrote", write_onepager(f"{stem}_Fund_Performance_1Pager.xlsx"))


if __name__ == "__main__":
    main(*(sys.argv[1:2] or []))

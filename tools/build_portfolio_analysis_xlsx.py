"""Convert the Portfolio Fund Analysis Report (PDF) into a workbook.

Usage: python tools/build_portfolio_analysis_xlsx.py [output.xlsx]

Source: Portfolio Fund Analysis Report, generated 2 September 2026, 9 pages.
The PDF has no text layer, so every figure below was read off the rendered
pages and is reproduced exactly as printed. Where the report's own numbers
do not reconcile against its 19-fund holdings table, the workbook shows
both the stated figure and the figure computed from the holdings, and the
Summary sheet lists the differences rather than quietly correcting them.

Sheets: Summary | Holdings | Category | Proposed Portfolio | Action Plan
"""

import sys

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
NAVY, BLUE, GREY, WHITE, YELLOW, LINE = "1F3864", "2E75B6", "F2F2F2", "FFFFFF", "FFF2CC", "BFBFBF"
GREEN_FILL, AMBER_FILL, RED_FILL = "C6EFCE", "FFEB9C", "FFC7CE"

BODY = Font(name=FONT, size=10)
BOLD = Font(name=FONT, size=10, bold=True)
CALC = Font(name=FONT, size=10, color="000000")
SRC = Font(name=FONT, size=10, color="0000FF")          # read off the PDF
HEAD = Font(name=FONT, size=10, bold=True, color=WHITE)
GREEN = Font(name=FONT, size=10, color="006100")
AMBER = Font(name=FONT, size=10, color="9C5700")
RED = Font(name=FONT, size=10, color="9C0006")
ITAL = Font(name=FONT, size=9, italic=True, color="404040")

PCT = '0.00%;(0.00%);-'
PCT0 = '0%;(0%);-'
DPCT = '+0.00%;-0.00%;-'
CNT = '#,##0;(#,##0);-'
THIN = Side(style="thin", color=LINE)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# rank, fund, AMC, allocation, category (as printed), concentration, tier, proposed
HOLDINGS = [
    (1,  "TRUSTMF Small Cap Fund - Regular (G)",            "TRUSTMF",  0.1519, "Small Cap",        "HIGH",   "", None),
    (2,  "Aditya Birla SL Quant Fund - Regular (G)",        "Aditya Birla", 0.1132, "Multi-Cap/Quant", "HIGH", "", None),
    (3,  "Bandhan Business Cycle Fund - Regular (G)",       "Bandhan",  0.1065, "Thematic",         "HIGH",   "", None),
    (4,  "Tata Business Cycle Fund (G)",                    "Tata",     0.1001, "Thematic",         "HIGH",   "Tier 3", 0.10),
    (5,  "Mahindra Manulife Mid Cap Fund (G)",              "Mahindra Manulife", 0.0699, "Mid Cap", "MEDIUM", "Tier 1", 0.20),
    (6,  "Mahindra Manulife Large & Mid Cap Fund-Reg (G)",  "Mahindra Manulife", 0.0573, "Large & Mid Cap", "MEDIUM", "Tier 2", 0.15),
    (7,  "Mahindra Manulife Small Cap Fund - Regular (G)",  "Mahindra Manulife", 0.0572, "Small Cap", "MEDIUM", "Tier 1", 0.15),
    (8,  "Mahindra Manulife Focused Fund (G)",              "Mahindra Manulife", 0.0561, "Focused/Multi-Cap", "MEDIUM", "Tier 2", 0.07),
    (9,  "Mahindra Manulife Business Cycle Fund - Reg (G)", "Mahindra Manulife", 0.0523, "Thematic", "MEDIUM", "", None),
    (10, "Tata India Innovation Fund - Regular (G)",        "Tata",     0.0514, "Thematic",         "MEDIUM", "", None),
    (11, "TRUSTMF Flexi Cap Fund - Regular (G)",            "TRUSTMF",  0.0451, "Flexi Cap",        "MEDIUM", "", None),
    (12, "Aditya Birla SL Ultra Short to Short Term Fund-Reg (G)", "Aditya Birla", 0.0311, "Debt/Short Term", "LOW", "", None),
    (13, "Bandhan Ultra Short Term Fund - Reg (G)",         "Bandhan",  0.0290, "Debt/Short Term",  "LOW",    "Tier 3", 0.08),
    (14, "Bandhan Flexi Cap Fund (G)",                      "Bandhan",  0.0270, "Flexi Cap",        "LOW",    "", None),
    (15, "Aditya Birla SL Flexi Cap Fund (G)",              "Aditya Birla", 0.0247, "Flexi Cap",    "LOW",    "Tier 1", 0.25),
    (16, "Mahindra Manulife Flexi Cap Fund (G)",            "Mahindra Manulife", 0.0153, "Flexi Cap", "LOW",   "", None),
    (17, "Tata Ultra Short Term Fund - Regular (G)",        "Tata",     0.0086, "Debt/Short Term",  "LOW",    "", None),
    (18, "TRUSTMF Money Market Fund (G)",                   "TRUSTMF",  0.0027, "Money Market",     "LOW",    "", None),
    (19, "Mahindra Manulife Value Fund - Regular (G)",      "Mahindra Manulife", 0.0006, "Value",   "LOW",    "", None),
]

CATEGORIES = ["Small Cap", "Mid Cap", "Large & Mid Cap", "Flexi Cap", "Focused/Multi-Cap",
              "Multi-Cap/Quant", "Thematic", "Debt/Short Term", "Money Market", "Value"]

# the report's own "Portfolio by Category" panel, exactly as printed
REPORTED_CATS = [
    ("Small Cap",                 0.2690, 3, "TRUSTMF, MM, others"),
    ("Business Cycle (Thematic)", 0.2589, 3, "Bandhan, Tata, MM"),
    ("Flexi Cap",                 0.1121, 4, "Aditya Birla, TRUSTMF, Bandhan, MM"),
    ("Multi-Cap & Others",        0.1132, 2, "Aditya Birla Quant, Tata Innovation"),
    ("Large & Mid Cap",           0.1272, 2, "MM Large & Mid Cap, MM Focused"),
    ("Debt & Money Market",       0.0714, 4, "Ultra Short Term, Money Market"),
]

FINDINGS = [
    ("Extreme Concentration Risk", "Top 3 funds = 36.16% of portfolio. Recommended maximum is 10-12% per fund."),
    ("Thematic Overlap", "26% of portfolio in Business Cycle funds (2 different funds with similar strategy)."),
    ("Flexi Cap Redundancy", "4 Flexi Cap funds doing similar work, fragmenting capital across overlapping strategies."),
    ("Tail End Holdings", "4 funds below 1% allocation (essentially dead weight from a portfolio perspective)."),
    ("Fee Inefficiency", "Managing 19 funds creates higher administrative burden and fee drag from lower-quality holdings."),
    ("Diversification Problem", "You have 19 funds but lack true diversification - too concentrated in small cap and thematic bets."),
]

ATTENTION = [
    ("Over-Allocated (>10%)", "TRUSTMF Small Cap Fund", 0.1519, "Concentration risk. Best 1Y returns but questionable as largest holding."),
    ("Over-Allocated (>10%)", "Aditya Birla SL Quant Fund", 0.1132, "Adequate holding but not a \"core\" fund in most portfolios."),
    ("Over-Allocated (>10%)", "Bandhan Business Cycle Fund", 0.1065, "Thematic bet taking up excessive space."),
    ("Over-Allocated (>10%)", "Tata Business Cycle Fund", 0.1001, "Duplicate thematic exposure with Bandhan BC Fund."),
    ("Micro-Position (<1%)", "TRUSTMF Money Market Fund", 0.0027, "Too small to matter, creates tracking burden."),
    ("Micro-Position (<1%)", "Mahindra Manulife Value Fund", 0.0006, "Essentially negligible, not even a rounding error."),
    ("Redundant", "4 Flexi Cap Funds", 0.1121, "Aditya Birla SL (2.47%), TRUSTMF (4.51%), Bandhan (2.70%), MM (1.53%) across overlapping strategies."),
    ("Redundant", "2 Business Cycle Funds", 0.2066, "Bandhan (10.65%) + Tata (10.01%) in the same thematic bet."),
    ("Redundant", "Small Cap Duplication", 0.2091, "TRUSTMF (15.19%) + MM Small Cap (5.72%) concentrated in small-cap segment."),
]

RECS = [
    ("Recommendation 1", "Immediate Fund Consolidation", "Reduce from 19 funds to 7 core holdings to eliminate redundancy and reduce fee drag.",
     ["Consolidate 4 Flexi Cap funds into 1 best-performing one",
      "Keep only 1 Business Cycle thematic bet (keep Tata, exit Bandhan)",
      "Reduce small cap concentration from 26.90% to 15-20%",
      "Exit all micro-positions (<0.5%)"]),
    ("Recommendation 2", "Rebalance Top Holdings", "Bring the largest holdings to healthier concentration levels (10-12% maximum).",
     ["Reduce TRUSTMF Small Cap from 15.19% to 12%",
      "Reduce Aditya Birla Quant from 11.32% to 10%",
      "Reduce Bandhan BC from 10.65% to exit",
      "Reduce Tata BC from 10.01% to 8%"]),
    ("Recommendation 3", "Tax-Efficient Exit Strategy", "Implement over 6 months to minimize capital gains tax impact.",
     ["Phase 1 (Weeks 1-2): Redirect new SIPs away from 12 redundant funds",
      "Phase 2 (Weeks 3-8): Exit micro-positions and worst performers",
      "Phase 3 (Months 2-4): Exit redundant flexi caps and thematic duplicates",
      "Phase 4 (Months 4-6): Finalize rebalancing and tax documentation"]),
    ("Recommendation 4", "Consider Direct Plan Migration", "Switching from Regular to Direct plans can save Rs 2-3 lakh annually in fees.",
     ["Review your relationship with your fund distributor",
      "Calculate tax cost of switching vs. long-term fee savings",
      "Direct plans typically have 0.5-1.5% lower expense ratios",
      "Over 10 years: Direct plans can add Rs 23+ lakh to your portfolio"]),
]

PLAN = [
    ("Week 1", "Consult your CA about tax implications of fund consolidation and switching to Direct plans"),
    ("Week 1", "Review your distributor relationship and fee structure (Regular vs Direct)"),
    ("Week 2", "Document current holdings and cost basis for tax planning"),
    ("Week 2-3", "Begin Phase 1 - Redirect all new SIPs to the 7 core funds only"),
    ("Week 3-4", "Exit micro-positions (funds <0.5% allocation)"),
    ("Week 4+", "Plan Phase 2 exits with CA guidance on timing"),
    ("Ongoing", "Monitor quarterly, rebalance once yearly"),
]

H_FIRST = 5
H_LAST = H_FIRST + len(HOLDINGS) - 1
H_ALLOC = f"Holdings!$D${H_FIRST}:$D${H_LAST}"
H_CAT = f"Holdings!$E${H_FIRST}:$E${H_LAST}"


def put(ws, ref, value, *, font=BODY, fmt=None, fill=None, align=None,
        wrap=False, border=False):
    c = ws[ref]
    c.value = value
    c.font = font
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    if align or wrap:
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if border:
        c.border = BOX
    return c


def banner(ws, title, subtitle, last_col):
    ws.merge_cells(f"A1:{last_col}1")
    put(ws, "A1", title, font=Font(name=FONT, size=13, bold=True, color=WHITE),
        fill=NAVY, align="left")
    ws.row_dimensions[1].height = 26
    ws.merge_cells(f"A2:{last_col}2")
    put(ws, "A2", subtitle, font=ITAL, fill=GREY, align="left")
    ws.row_dimensions[2].height = 16


def heads(ws, row, labels, widths):
    for i, (label, width) in enumerate(zip(labels, widths), start=1):
        col = get_column_letter(i)
        put(ws, f"{col}{row}", label, font=HEAD, fill=BLUE, align="center",
            wrap=True, border=True)
        ws.column_dimensions[col].width = width
    ws.row_dimensions[row].height = 30


def conc_style(level):
    return {"HIGH": (RED, RED_FILL), "MEDIUM": (AMBER, AMBER_FILL),
            "LOW": (GREEN, GREEN_FILL)}[level]


# ----------------------------------------------------------------- Holdings --
def build_holdings(wb):
    ws = wb.create_sheet("Holdings")
    ws.sheet_view.showGridLines = False
    banner(ws, "COMPLETE PORTFOLIO HOLDINGS  -  19 FUNDS",
           "Rank, fund, allocation, category and concentration exactly as printed in the "
           "report. Tier and Proposed % come from the report's 7-fund target portfolio.", "I")
    heads(ws, 4,
          ["Rank", "Fund Name", "AMC", "Allocation %", "Category",
           "Concentration", "Keep / Exit", "Proposed %", "Change"],
          [6, 46, 18, 13, 20, 14, 13, 12, 12])
    ws.freeze_panes = "C5"

    for i, (rank, fund, amc, alloc, cat, conc, tier, prop) in enumerate(HOLDINGS):
        r = H_FIRST + i
        font, fill = conc_style(conc)
        put(ws, f"A{r}", rank, fmt=CNT, align="center", border=True)
        put(ws, f"B{r}", fund, font=SRC, border=True)
        put(ws, f"C{r}", amc, border=True)
        put(ws, f"D{r}", alloc, font=SRC, fmt=PCT, align="center", border=True)
        put(ws, f"E{r}", cat, border=True)
        put(ws, f"F{r}", conc, font=font, fill=fill, align="center", border=True)
        keep = bool(tier)
        put(ws, f"G{r}", "KEEP" if keep else "EXIT",
            font=GREEN if keep else RED, fill=GREEN_FILL if keep else RED_FILL,
            align="center", border=True)
        if keep:
            put(ws, f"H{r}", prop, font=SRC, fmt=PCT0, align="center", border=True)
            put(ws, f"I{r}", f"=$H{r}-$D{r}", font=CALC, fmt=DPCT,
                align="center", border=True)
        else:
            put(ws, f"H{r}", 0, font=SRC, fmt=PCT0, align="center", border=True)
            put(ws, f"I{r}", f"=-$D{r}", font=CALC, fmt=DPCT, align="center", border=True)

    t = H_LAST + 1
    put(ws, f"A{t}", "", fill=NAVY, border=True)
    put(ws, f"B{t}", "TOTAL  (19 funds)", font=HEAD, fill=NAVY, border=True)
    put(ws, f"C{t}", "", fill=NAVY, border=True)
    put(ws, f"D{t}", f"=SUM($D${H_FIRST}:$D${H_LAST})", font=HEAD, fmt=PCT,
        fill=NAVY, align="center", border=True)
    put(ws, f"E{t}", "", fill=NAVY, border=True)
    put(ws, f"F{t}", f'=COUNTIF($G${H_FIRST}:$G${H_LAST},"KEEP")&" keep / "&'
                     f'COUNTIF($G${H_FIRST}:$G${H_LAST},"EXIT")&" exit"',
        font=HEAD, fill=NAVY, align="center", border=True)
    put(ws, f"G{t}", "", fill=NAVY, border=True)
    put(ws, f"H{t}", f"=SUM($H${H_FIRST}:$H${H_LAST})", font=HEAD, fmt=PCT0,
        fill=NAVY, align="center", border=True)
    put(ws, f"I{t}", "", fill=NAVY, border=True)

    ws.auto_filter.ref = f"A4:I{H_LAST}"
    return ws


# ----------------------------------------------------------------- Category --
def build_category(wb):
    ws = wb.create_sheet("Category")
    ws.sheet_view.showGridLines = False
    banner(ws, "PORTFOLIO BY CATEGORY",
           "Top block is computed live from the Holdings sheet. Bottom block is the "
           "report's own panel, reproduced as printed - the two do not agree everywhere.", "E")

    put(ws, "A4", "COMPUTED FROM HOLDINGS", font=Font(name=FONT, size=11, bold=True, color=NAVY))
    heads(ws, 5, ["Category", "Funds", "Allocation %", "Share of Equity", ""],
          [26, 9, 14, 15, 2])
    first = 6
    for i, cat in enumerate(CATEGORIES):
        r = first + i
        put(ws, f"A{r}", cat, border=True)
        put(ws, f"B{r}", f'=COUNTIF({H_CAT},$A{r})', font=CALC, fmt=CNT,
            align="center", border=True)
        put(ws, f"C{r}", f'=SUMIF({H_CAT},$A{r},{H_ALLOC})', font=CALC, fmt=PCT,
            align="center", border=True)
        put(ws, f"D{r}", f'=IF($C${first + len(CATEGORIES)}=0,"",$C{r}/$C${first + len(CATEGORIES)})',
            font=CALC, fmt=PCT, align="center", border=True)
    t = first + len(CATEGORIES)
    put(ws, f"A{t}", "TOTAL", font=HEAD, fill=NAVY, border=True)
    put(ws, f"B{t}", f"=SUM(B{first}:B{t - 1})", font=HEAD, fmt=CNT, fill=NAVY,
        align="center", border=True)
    put(ws, f"C{t}", f"=SUM(C{first}:C{t - 1})", font=HEAD, fmt=PCT, fill=NAVY,
        align="center", border=True)
    put(ws, f"D{t}", "", fill=NAVY, border=True)

    r = t + 2
    put(ws, f"A{r}", "AS STATED IN THE REPORT",
        font=Font(name=FONT, size=11, bold=True, color=NAVY))
    r += 1
    for i, label in enumerate(["Category (report wording)", "Stated %", "Funds",
                               "Report's fund list", "Note"]):
        col = get_column_letter(1 + i)
        put(ws, f"{col}{r}", label, font=HEAD, fill=BLUE, align="center",
            wrap=True, border=True)
    ws.column_dimensions["D"].width = 38
    ws.column_dimensions["E"].width = 52
    rep_first = r + 1
    notes = {
        "Small Cap": "Report's own text elsewhere gives TRUSTMF + MM Small Cap = 20.91%.",
        "Business Cycle (Thematic)": "Reconciles: 10.65 + 10.01 + 5.23 = 25.89%.",
        "Flexi Cap": "Reconciles: 4.51 + 2.70 + 2.47 + 1.53 = 11.21%.",
        "Multi-Cap & Others": "11.32% is Aditya Birla Quant alone; Tata Innovation (5.14%) not included.",
        "Large & Mid Cap": "Named funds total 5.73 + 5.61 = 11.34%.",
        "Debt & Money Market": "Reconciles: 3.11 + 2.90 + 0.86 + 0.27 = 7.14%.",
    }
    for i, (cat, pct, n, funds) in enumerate(REPORTED_CATS):
        rr = rep_first + i
        put(ws, f"A{rr}", cat, border=True)
        put(ws, f"B{rr}", pct, font=SRC, fmt=PCT, align="center", border=True)
        put(ws, f"C{rr}", n, fmt=CNT, align="center", border=True)
        put(ws, f"D{rr}", funds, wrap=True, border=True)
        put(ws, f"E{rr}", notes[cat], font=ITAL, wrap=True, border=True)
        ws.row_dimensions[rr].height = 28
    rt = rep_first + len(REPORTED_CATS)
    put(ws, f"A{rt}", "TOTAL AS STATED", font=HEAD, fill=NAVY, border=True)
    put(ws, f"B{rt}", f"=SUM(B{rep_first}:B{rt - 1})", font=HEAD, fmt=PCT,
        fill=NAVY, align="center", border=True)
    put(ws, f"C{rt}", f"=SUM(C{rep_first}:C{rt - 1})", font=HEAD, fmt=CNT,
        fill=NAVY, align="center", border=True)
    put(ws, f"D{rt}", "", fill=NAVY, border=True)
    put(ws, f"E{rt}", "The six stated buckets do not sum to 100%.",
        font=HEAD, fill=NAVY, border=True)

    chart = BarChart()
    chart.type = "bar"
    chart.title = "Allocation by category (computed)"
    chart.height, chart.width = 9, 13
    chart.add_data(Reference(ws, min_col=3, min_row=5, max_row=t - 1), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=first, max_row=t - 1))
    chart.legend = None
    ws.add_chart(chart, f"G5")
    return ws


# ------------------------------------------------------- Proposed Portfolio --
def build_proposed(wb):
    ws = wb.create_sheet("Proposed Portfolio")
    ws.sheet_view.showGridLines = False
    banner(ws, "PROPOSED OPTIMIZED PORTFOLIO  -  7 FUNDS",
           "The report's target portfolio: 19 funds down to 7, in three tiers.", "F")
    heads(ws, 4, ["Tier", "Fund", "Current %", "Proposed %", "Change", "Rationale"],
          [10, 40, 12, 12, 12, 52])
    tiers = [
        ("Tier 1", "Core Holdings", 0.60, [
            ("Aditya Birla SL Flexi Cap Fund", 0.0247, 0.25, "Best quality, consolidate other flexi caps here"),
            ("Mahindra Manulife Small Cap Fund", 0.0572, 0.15, "Consolidate TRUSTMF SC here, reduce concentration"),
            ("Mahindra Manulife Mid Cap Fund", 0.0699, 0.20, "Strong performer, expand position")]),
        ("Tier 2", "Balanced Holdings", 0.22, [
            ("Mahindra Manulife Large & Mid Cap Fund", 0.0573, 0.15, "Steady performer, lower volatility"),
            ("Mahindra Manulife Focused Fund", 0.0561, 0.07, "Satellite position, controlled exposure")]),
        ("Tier 3", "Stability & Defensive", 0.18, [
            ("Tata Business Cycle Fund", 0.1001, 0.10, "Single thematic bet (capped), exit Bandhan duplicate"),
            ("Bandhan Ultra Short Term Fund", 0.0290, 0.08, "Safety buffer, best debt fund quality")]),
    ]
    r = 5
    data_rows = []
    for tier, name, target, funds in tiers:
        ws.merge_cells(f"A{r}:F{r}")
        put(ws, f"A{r}", f"{tier}: {name}  -  {target:.0%} allocation",
            font=Font(name=FONT, size=10, bold=True, color=WHITE), fill=BLUE, border=True)
        r += 1
        for fund, cur, prop, why in funds:
            put(ws, f"A{r}", tier, align="center", border=True)
            put(ws, f"B{r}", fund, font=SRC, border=True)
            put(ws, f"C{r}", cur, font=SRC, fmt=PCT, align="center", border=True)
            put(ws, f"D{r}", prop, font=SRC, fmt=PCT0, align="center", border=True)
            put(ws, f"E{r}", f"=$D{r}-$C{r}", font=CALC, fmt=DPCT, align="center", border=True)
            put(ws, f"F{r}", why, wrap=True, border=True)
            data_rows.append(r)
            r += 1
        sub = r
        put(ws, f"A{sub}", "", fill=GREY, border=True)
        put(ws, f"B{sub}", f"{tier} subtotal", font=BOLD, fill=GREY, border=True)
        put(ws, f"C{sub}", f"=SUM(C{sub - len(funds)}:C{sub - 1})", font=BOLD,
            fmt=PCT, fill=GREY, align="center", border=True)
        put(ws, f"D{sub}", f"=SUM(D{sub - len(funds)}:D{sub - 1})", font=BOLD,
            fmt=PCT0, fill=GREY, align="center", border=True)
        put(ws, f"E{sub}", f"=$D{sub}-$C{sub}", font=BOLD, fmt=DPCT, fill=GREY,
            align="center", border=True)
        put(ws, f"F{sub}", f"Target {target:.0%}", font=ITAL, fill=GREY, border=True)
        r += 2

    put(ws, f"A{r}", "", fill=NAVY, border=True)
    put(ws, f"B{r}", "TOTAL  (7 funds)", font=HEAD, fill=NAVY, border=True)
    cur_sum = "+".join(f"C{x}" for x in data_rows)
    prop_sum = "+".join(f"D{x}" for x in data_rows)
    put(ws, f"C{r}", f"={cur_sum}", font=HEAD, fmt=PCT, fill=NAVY, align="center", border=True)
    put(ws, f"D{r}", f"={prop_sum}", font=HEAD, fmt=PCT0, fill=NAVY, align="center", border=True)
    put(ws, f"E{r}", f"=$D{r}-$C{r}", font=HEAD, fmt=DPCT, fill=NAVY, align="center", border=True)
    put(ws, f"F{r}", "The 12 funds not listed here are exited.", font=HEAD, fill=NAVY, border=True)
    return ws


# -------------------------------------------------------------- Action Plan --
def build_action_plan(wb):
    ws = wb.create_sheet("Action Plan")
    ws.sheet_view.showGridLines = False
    banner(ws, "STRATEGIC RECOMMENDATIONS & 30-DAY ACTION PLAN",
           "Reproduced from the report. Owner and Done columns are blank for you to fill in.", "E")
    for col, width in zip("ABCDE", (18, 34, 62, 18, 10)):
        ws.column_dimensions[col].width = width

    r = 4
    for tag, title, intro, bullets in RECS:
        ws.merge_cells(f"A{r}:E{r}")
        put(ws, f"A{r}", f"{tag}: {title}",
            font=Font(name=FONT, size=10, bold=True, color=WHITE), fill=BLUE, border=True)
        r += 1
        ws.merge_cells(f"A{r}:E{r}")
        put(ws, f"A{r}", intro, font=ITAL, fill=GREY, border=True)
        r += 1
        for b in bullets:
            put(ws, f"A{r}", "", border=True)
            ws.merge_cells(f"B{r}:C{r}")
            put(ws, f"B{r}", b, wrap=True, border=True)
            put(ws, f"D{r}", None, font=SRC, border=True)
            put(ws, f"E{r}", None, font=SRC, align="center", border=True)
            ws.row_dimensions[r].height = 22
            r += 1
        r += 1

    put(ws, f"A{r}", "ACTION PLAN: NEXT 30 DAYS",
        font=Font(name=FONT, size=11, bold=True, color=NAVY))
    r += 1
    for i, label in enumerate(["When", "Action", "", "Owner", "Done"]):
        col = get_column_letter(1 + i)
        put(ws, f"{col}{r}", label, font=HEAD, fill=BLUE, align="center", border=True)
    r += 1
    for when, what in PLAN:
        put(ws, f"A{r}", when, font=BOLD, align="center", border=True)
        ws.merge_cells(f"B{r}:C{r}")
        put(ws, f"B{r}", what, wrap=True, border=True)
        put(ws, f"D{r}", None, font=SRC, border=True)
        put(ws, f"E{r}", None, font=SRC, align="center", border=True)
        ws.row_dimensions[r].height = 24
        r += 1

    r += 1
    ws.merge_cells(f"A{r}:E{r + 1}")
    put(ws, f"A{r}", "Disclaimer (from the report): This analysis is based on current "
                     "allocation data and general investment principles. Before implementing "
                     "any changes, consult your financial advisor and CA for tax planning and "
                     "suitability assessment for your specific financial situation, goals, and "
                     "time horizon.   Report generated: September 2, 2026.",
        font=ITAL, wrap=True, border=True)
    return ws


# ------------------------------------------------------------------ Summary --
def build_summary(wb):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    banner(ws, "PORTFOLIO FUND ANALYSIS REPORT",
           "Converted from the 9-page PDF dated 2 September 2026. Blue figures are read "
           "off the report; black figures are computed here from the Holdings sheet.", "E")
    for col, width in zip("ABCDE", (34, 15, 15, 15, 62)):
        ws.column_dimensions[col].width = width

    put(ws, "A4", "HEADLINE NUMBERS", font=Font(name=FONT, size=11, bold=True, color=NAVY))
    for i, label in enumerate(["Metric", "As stated", "Computed", "Difference", "Comment"]):
        col = get_column_letter(1 + i)
        put(ws, f"{col}5", label, font=HEAD, fill=BLUE, align="center", border=True)

    top1 = f"LARGE(({H_ALLOC}),1)"
    top3 = "+".join(f"LARGE(({H_ALLOC}),{i})" for i in (1, 2, 3))
    top5 = "+".join(f"LARGE(({H_ALLOC}),{i})" for i in range(1, 6))
    metrics = [
        ("Total funds", 19, f"=COUNTA(Holdings!$B${H_FIRST}:$B${H_LAST})", CNT,
         "Matches."),
        ("Top fund holding", 0.1519, f"={top1}", PCT,
         "TRUSTMF Small Cap. Matches."),
        ("Top 3 funds", 0.3616, f"={top3}", PCT,
         "Report states 36.16%; the three largest holdings add to 37.16%."),
        ("Top 5 funds", 0.4817, f"={top5}", PCT,
         "Report states 48.17%; the five largest holdings add to 54.16%."),
        ("Allocation total", 1.0, f"=SUM({H_ALLOC})", PCT,
         "The 19 printed allocations sum to exactly 100%."),
        ("Funds kept in target", 7, f'=COUNTIF(Holdings!$G${H_FIRST}:$G${H_LAST},"KEEP")', CNT,
         "Matches the 7-fund proposed portfolio."),
        ("Funds exited", 12, f'=COUNTIF(Holdings!$G${H_FIRST}:$G${H_LAST},"EXIT")', CNT,
         "Matches the report's \"12 redundant funds\"."),
    ]
    r = 6
    for label, stated, formula, fmt, comment in metrics:
        put(ws, f"A{r}", label, font=BOLD, border=True)
        put(ws, f"B{r}", stated, font=SRC, fmt=fmt, align="center", border=True)
        put(ws, f"C{r}", formula, font=CALC, fmt=fmt, align="center", border=True)
        diff_fmt = DPCT if fmt == PCT else '+#,##0;-#,##0;-'
        put(ws, f"D{r}", f"=$C{r}-$B{r}", font=CALC, fmt=diff_fmt,
            align="center", border=True)
        put(ws, f"E{r}", comment, wrap=True, border=True)
        ws.row_dimensions[r].height = 24
        r += 1

    r += 1
    put(ws, f"A{r}", "CRITICAL ISSUE  (as printed)", font=Font(name=FONT, size=11, bold=True, color=NAVY))
    r += 1
    ws.merge_cells(f"A{r}:E{r + 1}")
    put(ws, f"A{r}", "Your portfolio shows severe concentration risk with the top 3 funds "
                     "controlling 36.16% of your portfolio. This creates unnecessary risk and "
                     "limits diversification benefits.",
        font=RED, fill=RED_FILL, wrap=True, border=True)
    r += 3

    put(ws, f"A{r}", "KEY FINDINGS & ISSUES", font=Font(name=FONT, size=11, bold=True, color=NAVY))
    r += 1
    for i, label in enumerate(["Finding", "", "", "", "Detail"]):
        col = get_column_letter(1 + i)
        put(ws, f"{col}{r}", label, font=HEAD, fill=BLUE, border=True)
    ws.merge_cells(f"A{r}:D{r}")
    r += 1
    for title, detail in FINDINGS:
        ws.merge_cells(f"A{r}:D{r}")
        put(ws, f"A{r}", title, font=BOLD, border=True)
        put(ws, f"E{r}", detail, wrap=True, border=True)
        ws.row_dimensions[r].height = 26
        r += 1

    r += 1
    put(ws, f"A{r}", "CONCENTRATION RISK  (report's own limits)",
        font=Font(name=FONT, size=11, bold=True, color=NAVY))
    r += 1
    for i, label in enumerate(["Level", "Stated", "Computed", "Limit", "Verdict"]):
        col = get_column_letter(1 + i)
        put(ws, f"{col}{r}", label, font=HEAD, fill=BLUE, align="center", border=True)
    r += 1
    limits = [("Top 1 fund", 0.1519, f"={top1}", "max 10-12%", 0.12),
              ("Top 3 funds", 0.3616, f"={top3}", "max 25-30%", 0.30),
              ("Top 5 funds", 0.4817, f"={top5}", "healthy 40-45%", 0.45)]
    for label, stated, formula, limit, cap in limits:
        put(ws, f"A{r}", label, font=BOLD, border=True)
        put(ws, f"B{r}", stated, font=SRC, fmt=PCT, align="center", border=True)
        put(ws, f"C{r}", formula, font=CALC, fmt=PCT, align="center", border=True)
        put(ws, f"D{r}", limit, align="center", border=True)
        put(ws, f"E{r}", f'=IF($C{r}>{cap},"OVER by "&TEXT($C{r}-{cap},"0.00%"),"Within limit")',
            font=RED, fill=RED_FILL, align="center", border=True)
        r += 1

    r += 1
    put(ws, f"A{r}", "FUNDS REQUIRING IMMEDIATE ATTENTION",
        font=Font(name=FONT, size=11, bold=True, color=NAVY))
    r += 1
    for i, label in enumerate(["Issue", "Fund", "", "Allocation", "Report's comment"]):
        col = get_column_letter(1 + i)
        put(ws, f"{col}{r}", label, font=HEAD, fill=BLUE, align="center", border=True)
    ws.merge_cells(f"B{r}:C{r}")
    r += 1
    for issue, fund, alloc, comment in ATTENTION:
        font, fill = ((RED, RED_FILL) if issue.startswith("Over")
                      else (AMBER, AMBER_FILL) if issue.startswith("Micro")
                      else (BODY, YELLOW))
        put(ws, f"A{r}", issue, font=font, fill=fill, border=True)
        ws.merge_cells(f"B{r}:C{r}")
        put(ws, f"B{r}", fund, font=BOLD, border=True)
        put(ws, f"D{r}", alloc, font=SRC, fmt=PCT, align="center", border=True)
        put(ws, f"E{r}", comment, wrap=True, border=True)
        ws.row_dimensions[r].height = 26
        r += 1

    r += 1
    put(ws, f"A{r}", "NOTE ON THE SOURCE", font=Font(name=FONT, size=11, bold=True, color=NAVY))
    r += 1
    ws.merge_cells(f"A{r}:E{r + 2}")
    put(ws, f"A{r}", "The PDF is a scan with no text layer, so every figure here was read "
                     "off the rendered pages rather than extracted. Three of the report's "
                     "own headline percentages do not reconcile against its 19-fund table: "
                     "top 3 (36.16% stated vs 37.16% computed), top 5 (48.17% vs 54.16%) and "
                     "Small Cap (26.90% stated vs 20.91% for the two small-cap funds, which "
                     "is the figure the report itself uses on its own page 5). The stated "
                     "figures are kept in the \"As stated\" column and nothing has been "
                     "silently corrected - check them against the source data before acting.",
        font=ITAL, fill=YELLOW, wrap=True, border=True)
    return ws


def main(out="Portfolio_Fund_Analysis.xlsx"):
    wb = Workbook()
    wb.remove(wb.active)
    build_summary(wb)
    build_holdings(wb)
    build_category(wb)
    build_proposed(wb)
    build_action_plan(wb)
    wb.active = 0
    wb.save(out)
    print(f"wrote {out}")


if __name__ == "__main__":
    main(*(sys.argv[1:2] or []))

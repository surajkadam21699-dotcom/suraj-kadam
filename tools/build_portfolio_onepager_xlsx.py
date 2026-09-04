"""Build the one-page Portfolio Concentration Review as a single-sheet workbook.

Usage: python tools/build_portfolio_onepager_xlsx.py [output.xlsx]

The Excel twin of the published one-pager: concentration against limits, all
19 holdings drawn to scale with data bars, what is wrong, and the 7-fund
target. Everything on one sheet.

Figures are recomputed from the 19 printed allocations rather than taken
from the source report's headline percentages, three of which do not
reconcile against its own holdings table. The footer records the gaps.
Direct Plan migration is omitted at the client's instruction.
"""

import sys

from openpyxl import Workbook
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
NAVY, BLUE, GREY, WHITE, LINE = "1F3864", "2E75B6", "F2F2F2", "FFFFFF", "BFBFBF"
RED_FILL, AMBER_FILL, GREEN_FILL, QUIET = "FFC7CE", "FFEB9C", "C6EFCE", "E7E9E8"

BODY = Font(name=FONT, size=10)
BOLD = Font(name=FONT, size=10, bold=True)
CALC = Font(name=FONT, size=10, color="000000")
SRC = Font(name=FONT, size=10, color="0000FF")
HEAD = Font(name=FONT, size=10, bold=True, color=WHITE)
RED = Font(name=FONT, size=10, color="9C0006")
AMBER = Font(name=FONT, size=10, color="9C5700")
GREEN = Font(name=FONT, size=10, color="006100")
ITAL = Font(name=FONT, size=9, italic=True, color="404040")
SECT = Font(name=FONT, size=11, bold=True, color=NAVY)

PCT = '0.00%;(0.00%);-'
PCT0 = '0%;(0%);-'
DPTS = '+0.00"  pts";-0.00"  pts";-'
CNT = '#,##0;(#,##0);-'
THIN = Side(style="thin", color=LINE)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# rank, fund, category, allocation, band, keep
HOLDINGS = [
    (1,  "TRUSTMF Small Cap Fund - Regular (G)",                    "Small Cap",         0.1519, "Over 10%",  False),
    (2,  "Aditya Birla SL Quant Fund - Regular (G)",                "Multi-Cap/Quant",   0.1132, "Over 10%",  False),
    (3,  "Bandhan Business Cycle Fund - Regular (G)",               "Thematic",          0.1065, "Over 10%",  False),
    (4,  "Tata Business Cycle Fund (G)",                            "Thematic",          0.1001, "Over 10%",  True),
    (5,  "Mahindra Manulife Mid Cap Fund (G)",                      "Mid Cap",           0.0699, "5 - 7%",    True),
    (6,  "Mahindra Manulife Large & Mid Cap Fund-Reg (G)",          "Large & Mid Cap",   0.0573, "5 - 7%",    True),
    (7,  "Mahindra Manulife Small Cap Fund - Regular (G)",          "Small Cap",         0.0572, "5 - 7%",    True),
    (8,  "Mahindra Manulife Focused Fund (G)",                      "Focused/Multi-Cap", 0.0561, "5 - 7%",    True),
    (9,  "Mahindra Manulife Business Cycle Fund - Reg (G)",         "Thematic",          0.0523, "5 - 7%",    False),
    (10, "Tata India Innovation Fund - Regular (G)",                "Thematic",          0.0514, "5 - 7%",    False),
    (11, "TRUSTMF Flexi Cap Fund - Regular (G)",                    "Flexi Cap",         0.0451, "Under 5%",  False),
    (12, "Aditya Birla SL Ultra Short to Short Term Fund-Reg (G)",  "Debt/Short Term",   0.0311, "Under 5%",  False),
    (13, "Bandhan Ultra Short Term Fund - Reg (G)",                 "Debt/Short Term",   0.0290, "Under 5%",  True),
    (14, "Bandhan Flexi Cap Fund (G)",                              "Flexi Cap",         0.0270, "Under 5%",  False),
    (15, "Aditya Birla SL Flexi Cap Fund (G)",                      "Flexi Cap",         0.0247, "Under 5%",  True),
    (16, "Mahindra Manulife Flexi Cap Fund (G)",                    "Flexi Cap",         0.0153, "Under 5%",  False),
    (17, "Tata Ultra Short Term Fund - Regular (G)",                "Debt/Short Term",   0.0086, "Under 5%",  False),
    (18, "TRUSTMF Money Market Fund (G)",                           "Money Market",      0.0027, "Under 5%",  False),
    (19, "Mahindra Manulife Value Fund - Regular (G)",              "Value",             0.0006, "Under 5%",  False),
]

KEEP = [
    ("Aditya Birla SL Flexi Cap Fund",          "CORE",      0.0247, 0.25),
    ("Mahindra Manulife Mid Cap Fund",          "CORE",      0.0699, 0.20),
    ("Mahindra Manulife Small Cap Fund",        "CORE",      0.0572, 0.15),
    ("Mahindra Manulife Large & Mid Cap Fund",  "BALANCED",  0.0573, 0.15),
    ("Mahindra Manulife Focused Fund",          "BALANCED",  0.0561, 0.07),
    ("Tata Business Cycle Fund",                "DEFENSIVE", 0.1001, 0.10),
    ("Bandhan Ultra Short Term Fund",           "DEFENSIVE", 0.0290, 0.08),
]

FINDINGS = [
    ("Concentration", "One fund at 15.19% against a 10-12% ceiling, and the top five at 54.16%."),
    ("The same bet, twice", "Bandhan and Tata Business Cycle hold 20.66% in one identical theme; "
                            "with Mahindra Manulife BC and Tata Innovation, thematic exposure reaches 31.03%."),
    ("Four flexi caps doing one job", "11.21% split across four overlapping funds - no fund large enough to matter."),
    ("Small-cap doubling", "TRUSTMF and Mahindra Manulife Small Cap together hold 20.91%."),
    ("A tail that earns nothing", "Three holdings below 1% - the smallest is 0.06% - add tracking work and no return."),
]

WIDTHS = [(("A"), 6), (("B"), 52), (("C"), 20), (("D"), 13), (("E"), 13), (("F"), 46)]
H_FIRST = 16
H_LAST = H_FIRST + len(HOLDINGS) - 1
ALLOC = f"$D${H_FIRST}:$D${H_LAST}"


def put(ws, ref, value, *, font=BODY, fmt=None, fill=None, align=None,
        wrap=False, border=False):
    c = ws[ref]
    c.value = value
    c.font = font
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    if align or wrap:
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if border:
        c.border = BOX
    return c


def band_style(band):
    return {"Over 10%": (RED, RED_FILL), "5 - 7%": (AMBER, AMBER_FILL),
            "Under 5%": (BODY, QUIET)}[band]


def row_heads(ws, row, labels, aligns=None):
    for i, label in enumerate(labels, start=1):
        col = get_column_letter(i)
        put(ws, f"{col}{row}", label, font=HEAD, fill=BLUE,
            align=(aligns or {}).get(col, "center"), wrap=True, border=True)
    ws.row_dimensions[row].height = 22


def main(out="Portfolio_Review_1Pager.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Review"
    ws.sheet_view.showGridLines = False
    for col, width in WIDTHS:
        ws.column_dimensions[col].width = width

    # ------------------------------------------------------------ masthead --
    ws.merge_cells("A1:F1")
    put(ws, "A1", "PORTFOLIO CONCENTRATION REVIEW",
        font=Font(name=FONT, size=14, bold=True, color=WHITE), fill=NAVY, align="left")
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A2:F2")
    put(ws, "A2", "Nineteen funds, four decisions      |      As on 2 September 2026      "
                  "|      19 holdings      |      Allocation basis 100.00%",
        font=ITAL, fill=GREY, align="left")
    ws.row_dimensions[2].height = 17

    ws.merge_cells("A4:F5")
    put(ws, "A4", "The portfolio holds nineteen funds but is not diversified. Four of them take "
                  "47.17% of the money, small-cap and thematic bets together run to 51.94%, and "
                  "nine holdings at the tail contribute almost nothing. The fix is consolidation, "
                  "not addition.",
        font=Font(name=FONT, size=11), wrap=True)
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 20

    # ------------------------------------------- concentration vs limits --
    put(ws, "A7", "CONCENTRATION AGAINST LIMITS", font=SECT)
    row_heads(ws, 8, ["", "Measure", "Holding", "Actual", "Limit", "Verdict"],
              {"B": "left", "C": "left", "F": "left"})
    limits = [
        ("Largest holding", "TRUSTMF Small Cap Fund", f"=LARGE({ALLOC},1)", 0.12, "max 10-12%"),
        ("Top 3 funds", "Three largest holdings",
         f"=LARGE({ALLOC},1)+LARGE({ALLOC},2)+LARGE({ALLOC},3)", 0.30, "max 25-30%"),
        ("Top 5 funds", "Five largest holdings",
         "+".join(f"LARGE({ALLOC},{i})" for i in range(1, 6)).join(("=", "")), 0.45, "healthy 40-45%"),
    ]
    for j, (measure, holding, formula, cap, label) in enumerate(limits):
        r = 9 + j
        put(ws, f"A{r}", "", border=True)
        put(ws, f"B{r}", measure, font=BOLD, border=True)
        put(ws, f"C{r}", holding, border=True)
        put(ws, f"D{r}", formula, font=CALC, fmt=PCT, fill=RED_FILL, align="center", border=True)
        put(ws, f"E{r}", label, align="center", border=True)
        put(ws, f"F{r}", f'=IF($D{r}>{cap},"OVER by "&TEXT(($D{r}-{cap})*100,"0.00")&" pts",'
                         f'"Within limit")',
            font=RED, fill=RED_FILL, border=True)

    # ------------------------------------------------ holdings, to scale --
    put(ws, "A13", "WHERE THE MONEY SITS  -  ALL 19 HOLDINGS", font=SECT)
    ws.merge_cells("A14:F14")
    put(ws, "A14", "Bars are drawn to scale. Four blocks carry 47.17%; the nine smallest "
                   "holdings carry 18.41% between them.", font=ITAL)
    row_heads(ws, 15, ["#", "Fund", "Category", "Allocation", "Band", "Action"],
              {"B": "left", "C": "left"})

    for i, (rank, fund, cat, alloc, band, keep) in enumerate(HOLDINGS):
        r = H_FIRST + i
        font, fill = band_style(band)
        put(ws, f"A{r}", rank, fmt=CNT, align="center", border=True)
        put(ws, f"B{r}", fund, font=SRC, border=True)
        put(ws, f"C{r}", cat, border=True)
        put(ws, f"D{r}", alloc, font=SRC, fmt=PCT, align="center", border=True)
        put(ws, f"E{r}", band, font=font, fill=fill, align="center", border=True)
        put(ws, f"F{r}", "KEEP" if keep else "Exit / consolidate",
            font=GREEN if keep else RED, fill=GREEN_FILL if keep else None, border=True)

    t = H_LAST + 1
    put(ws, f"A{t}", "", fill=NAVY, border=True)
    put(ws, f"B{t}", "TOTAL", font=HEAD, fill=NAVY, border=True)
    put(ws, f"C{t}", "", fill=NAVY, border=True)
    put(ws, f"D{t}", f"=SUM({ALLOC})", font=HEAD, fmt=PCT, fill=NAVY, align="center", border=True)
    put(ws, f"E{t}", "", fill=NAVY, border=True)
    put(ws, f"F{t}", f'=COUNTIF($F${H_FIRST}:$F${H_LAST},"KEEP")&" keep  /  "&'
                     f'COUNTIF($F${H_FIRST}:$F${H_LAST},"Exit / consolidate")&" exit"',
        font=HEAD, fill=NAVY, border=True)

    ws.conditional_formatting.add(
        f"D{H_FIRST}:D{H_LAST}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=0.16,
                    color="A6362C", showValue=True))

    # ------------------------------------------------------ what is wrong --
    r = t + 2
    put(ws, f"A{r}", "WHAT IS WRONG", font=SECT)
    r += 1
    for title, detail in FINDINGS:
        put(ws, f"A{r}", "", border=True)
        put(ws, f"B{r}", title, font=BOLD, border=True)
        ws.merge_cells(f"C{r}:F{r}")
        put(ws, f"C{r}", detail, wrap=True, border=True)
        ws.row_dimensions[r].height = 26
        r += 1

    # ------------------------------------------------------------ the fix --
    r += 1
    put(ws, f"A{r}", "THE FIX  -  19 FUNDS DOWN TO 7", font=SECT)
    r += 1
    ws.merge_cells(f"A{r}:F{r + 1}")
    put(ws, f"A{r}", "Consolidate into seven holdings across three tiers. Keep one business-cycle "
                     "bet, one flexi cap, one small cap. Exit the other twelve in phases over six "
                     "months so capital-gains tax is managed rather than triggered at once. The "
                     "seven funds to keep hold just 39.43% of the portfolio today, which is why "
                     "the exit runs in phases and starts by redirecting fresh SIPs rather than "
                     "selling.", wrap=True)
    ws.row_dimensions[r].height = 20
    ws.row_dimensions[r + 1].height = 20
    r += 3

    put(ws, f"A{r}", "THE SEVEN TO KEEP", font=SECT)
    r += 1
    row_heads(ws, r, ["", "Fund", "Tier", "Now", "Target", "Change"], {"B": "left", "C": "left"})
    r += 1
    k_first = r
    for fund, tier, now, target in KEEP:
        put(ws, f"A{r}", "", border=True)
        put(ws, f"B{r}", fund, font=SRC, border=True)
        put(ws, f"C{r}", tier, font=ITAL, align="center", border=True)
        put(ws, f"D{r}", now, font=SRC, fmt=PCT, align="center", border=True)
        put(ws, f"E{r}", target, font=SRC, fmt=PCT0, fill=GREEN_FILL, align="center", border=True)
        put(ws, f"F{r}", f"=$E{r}-$D{r}", font=CALC, fmt='+0.00%;-0.00%;-',
            align="center", border=True)
        r += 1
    k_last = r - 1
    put(ws, f"A{r}", "", fill=NAVY, border=True)
    put(ws, f"B{r}", "Seven holdings", font=HEAD, fill=NAVY, border=True)
    put(ws, f"C{r}", "", fill=NAVY, border=True)
    put(ws, f"D{r}", f"=SUM(D{k_first}:D{k_last})", font=HEAD, fmt=PCT, fill=NAVY,
        align="center", border=True)
    put(ws, f"E{r}", f"=SUM(E{k_first}:E{k_last})", font=HEAD, fmt=PCT0, fill=NAVY,
        align="center", border=True)
    put(ws, f"F{r}", f"=$E{r}-$D{r}", font=HEAD, fmt='+0.00%;-0.00%;-', fill=NAVY,
        align="center", border=True)
    r += 2

    ws.merge_cells(f"A{r}:F{r + 1}")
    put(ws, f"A{r}", "FIRST MOVE, THIS MONTH:  stop every SIP going into the twelve funds being "
                     "exited and redirect them to these seven. It rebalances with new money "
                     "before a single unit is sold, so nothing is taxed.",
        font=GREEN, fill=GREEN_FILL, wrap=True, border=True)
    ws.row_dimensions[r].height = 20
    ws.row_dimensions[r + 1].height = 20
    r += 3

    # ------------------------------------------------------------- footer --
    ws.merge_cells(f"A{r}:F{r + 2}")
    put(ws, f"A{r}", "NOTE ON THE SOURCE FIGURES:  Three percentages in the underlying report do "
                     "not reconcile against its own holdings table, so this summary uses figures "
                     "recomputed from the nineteen allocations - top 3 is 37.16% (report states "
                     "36.16%), top 5 is 54.16% (report states 48.17%), and small cap is 20.91% "
                     "(report states 26.90% in one place and 20.91% in another). The report also "
                     "describes four holdings below 1%; there are three. Confirm against the "
                     "source data before acting.",
        font=Font(name=FONT, size=9, color="9C5700"), fill="FFF2CC", wrap=True, border=True)
    for k in range(3):
        ws.row_dimensions[r + k].height = 18
    r += 4

    ws.merge_cells(f"A{r}:F{r + 1}")
    put(ws, f"A{r}", "This analysis is based on current allocation data and general investment "
                     "principles. Before implementing any changes, consult your financial adviser "
                     "and CA for tax planning and suitability assessment for your specific "
                     "financial situation, goals, and time horizon. Mutual fund investments are "
                     "subject to market risks; read all scheme-related documents carefully.",
        font=ITAL, wrap=True)
    ws.row_dimensions[r].height = 18
    ws.row_dimensions[r + 1].height = 18

    ws.print_area = f"A1:F{r + 1}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    wb.save(out)
    print(f"wrote {out}")


if __name__ == "__main__":
    main(*(sys.argv[1:2] or []))
